import json
import os
from werkzeug import secure_filename
from openpyxl import load_workbook, Workbook
from flask import flash, url_for
from flask.ext.sqlalchemy import SQLAlchemy
from flask.ext.migrate import Migrate
from app import app

#Create database
db = SQLAlchemy(app)

class Seed(db.Model):
    """Seed is an SQLAlchemy db model for handling seed product data.
    
    Data Attributes:
        id -- Unique numerical ID generated by database.
        name -- Name as we want it to appear on the product page.
        genus -- Genus of seed.
        species -- Species of seed.
        description -- Product description.
        synonyms -- Other names the seed is called.
        series -- Series seed belongs to, if any. (Ex: Benary's Giant)
        variety -- Variety of seed. (Ex: zinnia, coleus, sunflower)
        category -- Type of plant. (Ex: annual flower, herb, vegetable)
        price -- Current price of the seed.
        is_active -- Whether or not the seed is to be restocked.
        in_stock -- Whether or not the seed is in stock.
        thumbnail -- Filename of thumbnail image.

    """

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(64), unique=True)
    genus = db.Column(db.String(64))
    species = db.Column(db.String(64))
    description = db.Column(db.String(64))
    synonyms = db.relationship('Synonym',
                               backref='seed',
                               lazy='dynamic',
                               cascade='all, merge, delete, delete-orphan')
    series = db.Column(db.String(32))
    variety = db.Column(db.String(32))
    category = db.Column(db.String(32))
    price = db.Column(db.Float)
    is_active = db.Column(db.Boolean)
    in_stock = db.Column(db.Boolean)
    thumbnail = db.Column(db.String(64))
    errors = []
    
    def __init__(self,
                 name=None,
                 binomen=None,
                 description=None,
                 variety=None,
                 category=None,
                 price=None,
                 is_active=None,
                 in_stock=None,
                 thumbnail=None,
                 synonyms=None,
                 series=None):
        self.name = name
        self.set_binomen(binomen)
        self.description = description
        self.set_variety(variety)
        self.set_category(category)
        self.price = price
        self.is_active = is_active
        self.in_stock = in_stock
        self.thumbnail = thumbnail
        self.add_synonyms_from_string(synonyms)
        self.series = series

    def __eq__(self, other):
        if (
                self.name == other.name and
                self.get_binomen() == other.get_binomen() and
                self.description == other.description and
                self.variety == other.variety and
                self.category == other.category and
                self.price == other.price and
                self.is_active == other.is_active and
                self.in_stock == other.in_stock and
                self.thumbnail == other.thumbnail and
                self.get_synonyms_list() == other.get_synonyms_list() and
                self.series == other.series):
            return True
        else:
            return False

    
    def populate_from_form(self, form):
        """Populates addseed from a complete form object."""
        self.name = form.name.data
        self.set_binomen(form.binomen.data)
        self.description = form.description.data
        self.set_variety(form.variety.data)
        self.set_category(form.category.data)
        self.price = form.price.data
        self.is_active = form.is_active.data
        self.in_stock = form.in_stock.data
        self.add_synonyms_from_string(form.synonyms.data)
        series = form.series.data
        if form.thumbnail.data:
            self.save_thumbnail(form.thumbnail.data)

    def add_synonym(self, synonym):
        """Adds a synonym linked to our seed object."""
        self.synonyms.append(Synonym(synonym))

    def add_synonyms_from_list(self, synonyms):
        """Adds synonyms from a list."""
        for synonym in synonyms:
            self.add_synonym(synonym)

    def add_synonyms_from_string(self, synonyms):
        if synonyms is not None:
            self.add_synonyms_from_list(string_to_list(synonyms))
        else:
            synonyms = None

    def get_synonyms_list(self):
        """Returns a list of synonyms."""
        synonyms = []
        for synonym in self.synonyms.all():
            synonyms.append(synonym.name)
        return synonyms

    def get_synonyms_string(self):
        """Returns a string containing list of synonyms."""
        return list_to_string(self.get_synonyms_list())

    def get_fullname(self):
        """Returns full seed name including series."""
        if self.series is not None:
            return self.series + ' ' + self.name
        else:
            return self.name

    def set_binomen(self, binomen):
        """Sets genus and species from a binomen string."""
        if binomen is not None:
            genspec = [nomen.strip() for nomen in binomen.split(' ')]
            self.genus = genspec[0].lower().capitalize()
            self.species = genspec[1].lower()
        else:
            self.genus = None
            self.species = None

    def get_binomen(self):
        """Returns a binomen string."""
        return self.genus + ' ' + self.species

    def set_variety(self, variety):
        """Sets variety in the proper format for use with the database."""
        if variety is not None:
            self.variety = variety.lower().strip()
        else:
            variety = None

    def set_category(self, category):
        """Sets category in db-friendly format."""
        if category is not None:
            self.category = category.lower().strip()
        else:
            self.category = None

    def get_images_directory(self):
        """Returns the path to this seed's image files."""
        return os.path.join(
            app.config['IMAGES_FOLDER'],
            self.variety.lower(),
            self.name.lower()
        )


    def create_images_directory(self):
        """Creates this seed's images directory if not present."""
        try:
            os.makedirs(self.get_images_directory())
        except OSError as error:
            if (error.errno == os.errno.EEXIST and
                    os.path.isdir(self.get_images_directory())):
                pass
            else:
                raise

    def save_image(self, image):
        """Saves an image to this seed's images directory."""
        fullpath = os.path.join(self.get_images_directory(), secure_filename(image.filename))
        self.create_images_directory()
        image.save(fullpath)

    def save_thumbnail(self, image):
        """Saves and sets a thumbnail image."""
        self.thumbnail = secure_filename(image.filename)
        self.save_image(image)

    def get_image_location(self, filename):
        """Returns the full path of the image specified."""
        return os.path.join(self.get_images_directory(), filename)

    def image_exists(self, filename):
        """Returns true if the specified image file exists."""
        if os.path.isfile(self.get_image_location(filename)):
            return True
        else:
            return False

    def thumbnail_exists(self):
        """Returns true if thumbnail image file exists."""
        if self.thumbnail is not None:
            return self.image_exists(self.thumbnail)
        else:
            return False

    def get_image_url(self, filename):
        """Returns the URL for an image as assocated with this seed."""
        relpath = ('images/' + 
                   self.variety.lower() + '/' + 
                   self.name.lower() + '/' +
                   filename)
        return url_for('static', filename=relpath)

    def get_thumbnail_url(self):
        """Returns the URL of this seed's thumbnail."""
        return self.get_image_url(self.thumbnail)
        
    def verify(self):
        """Verifies that seed is ready to be stored in the database."""
        valid = True
        if Seed.query.filter_by(name=self.name).first() is not None:
            self.errors.append('%s is already in the database! Please edit it instead of re-adding.' % self.name)
            valid = False
        return valid

    def add_variety_to_json_file(self):
        """Checks varieties.json and adds this seed's variety if needed."""
        filename = os.path.join(app.config['JSON_FOLDER'], 'varieties.json')
        changed = False
        try:
            with open(filename, 'r') as infile:
                varieties = json.loads(infile.read())
                if self.variety not in varieties:
                    varieties.append(self.variety)
                    changed = True
        except IOError:
            #File doesn't exist, prepare varieties to be written to new file.
            varieties = [self.variety]
            changed = True
        if changed == True:
            with open(filename, 'w') as outfile:
                outfile.write(json.dumps(varieties, indent=2))

    def add_category_to_json_file(self):
        """Adds seed's category and variety to categories.json if needed."""
        changed = False
        filename = os.path.join(app.config['JSON_FOLDER'], 'categories.json')
        try:
            with open(filename, 'r') as infile:
                categories = json.loads(infile.read())
                if self.category not in categories:
                    categories[self.category]=[self.variety]
                    changed = True
                elif self.variety not in categories[self.category]:
                    categories[self.category].append(self.variety)
                    changed = True
        except IOError:
            categories = {self.category: [self.variety]}
            changed = True
        if changed == True:
            with open(filename, 'w') as outfile:
                outfile.write(json.dumps(categories, indent=2))


    def save_to_database(self):
        """Saves seed to the database."""
        db.session.add(self)
        db.session.commit()

    def save(self):
        """Saves seed's JSON files and saves seed to database."""
        self.add_variety_to_json_file()
        self.add_category_to_json_file()
        self.save_to_database()
        
                    
class Synonym(db.Model):
    """A synonym for a seed.

    Data Attributes:
        id -- Unique ID for database
        name -- The synonym itself.
        seed_id -- The ID of the parent seed.
    """
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(64))
    seed_id = db.Column(db.Integer, db.ForeignKey('seed.id'))

    def __init__(self, name):
        self.name = name



def string_to_list(stringlist):
    """Splits a string by its commas and returns a list."""
    return [item.strip() for item in stringlist.split(',')]

def list_to_string(listobj):
    """Returns a list as a string w/ commas."""
    return ', '.join(listobj)

def save_seeds_to_xlsx(seeds, filename):
    """Saves a list of Seed objects to an Excel spreadsheet."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'seeds'
    sheet.append(['Name',
                  'Binomen',
                  'Variety',
                  'Category',
                  'Price',
                  'Description',
                  'Synonyms',
                  'Series',
                  'Thumbnail',
                  'Active',
                  'In Stock'])
    for seed in seeds:
        sheet.append([seed.name,
                      seed.get_binomen(),
                      seed.variety,
                      seed.category,
                      seed.price,
                      seed.description,
                      seed.get_synonyms_string(),
                      seed.series,
                      seed.thumbnail,
                      seed.is_active,
                      seed.in_stock])
    sheet.freeze_panes = sheet['A2']
    wb.save(filename)

def load_seeds_from_xlsx(filename):
    """Returns a list of Seed objects loaded from an Excel spreadsheet."""
    seeds = []
    wb = load_workbook(filename)
    sheet = wb['seeds']
    rows = sheet.iter_rows()
    firstrow = next(iter(rows))
    for idx, cell in enumerate(firstrow, start=0):
        val = cell.value.lower().strip()
        if val == 'name':
            name_col = idx
        elif val == 'binomen':
            binomen_col = idx
        elif val == 'variety':
            variety_col = idx
        elif val == 'category':
            category_col = idx
        elif val == 'price':
            price_col = idx
        elif val == 'description':
            description_col = idx
        elif val == 'synonyms':
            synonyms_col = idx
        elif val == 'series':
            series_col = idx
        elif val == 'thumbnail':
            thumbnail_col = idx
        elif val == 'active':
            active_col = idx
        elif val == 'in stock':
            in_stock_col = idx
    for row in rows:
        seeds.append(Seed(name=row[name_col].value,
                          binomen=row[binomen_col].value,
                          variety=row[variety_col].value,
                          category=row[category_col].value,
                          price=row[price_col].value,
                          description=row[description_col].value,
                          synonyms=row[synonyms_col].value,
                          series=row[series_col].value,
                          thumbnail=row[thumbnail_col].value,
                          is_active=row[active_col].value,
                          in_stock=row[in_stock_col].value))
    return seeds



