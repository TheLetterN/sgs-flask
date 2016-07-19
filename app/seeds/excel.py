# -*- coding: utf-8 -*-
# This file is part of SGS-Flask.

# SGS-Flask is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.

# SGS-Flask is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

# Copyright Swallowtail Garden Seeds, Inc


"""
    sgs-flask.app.seeds.excel

    This module implements an interface for moving data between the database
    and xlsx (Excel spreadsheet) files.
"""


import datetime
import json
import sys
import warnings

import openpyxl

from app import db
from app.db_helpers import dbify
from app.seeds.models import (
    BotanicalName,
    Section,
    CommonName,
    Cultivar,
    Image,
    Index,
    Packet,
    Quantity
)


def queryable_dicts_to_json(objects):
    """Generate a JSON string of dictionaries for easily querying.

    Note:
        All objects passed must have the property `queryable_dict`.

    Args:
        objects: A list of objects to get queryable dictionaries from.

    Raises:
        TypeError: If any item in items lacks the `queryable_dict` property.
    """
    if all(hasattr(obj, 'queryable_dict') for obj in objects):
        return json.dumps(tuple(obj.queryable_dict for obj in objects))
    else:
        raise TypeError('One or more objects lack the property '
                        '\'queryable_dict\'!')


class SeedsWorksheet(object):
    """A container for an `openpyxl` worksheet.

    Since extending openpyxl's classes seems to be an exercise in futility, it
    is easier to just encapsulate them and create an interface that's specific
    to how we want our worksheet data formatted.
    """
    def __init__(self, sheet):
        self._ws = sheet

    def __getitem__(self, x):
        """Allow direct access to keys present in worksheet."""
        return self._ws[x]

    @property
    def title(self):
        return self._ws.title

    @title.setter
    def title(self, title):
        self._ws.title = title

    @property
    def rows(self):
        """tuple: A `tuple` listing each row as a tuple of cells.

        I'm fiendishly using the protected _cells variable here because as of
        writing this the rows property in `Worksheet` returns ((),) even if
        there is data in cell A1, while `Worksheet._cells` is empty until at
        least one cell is given a value.
        """
        # TODO: Forward _ws.rows once rows is fixed in openpyxl.
        if self._ws._cells:
            return tuple(self._ws.iter_rows())
        else:
            return ((),)

    @property
    def active_row(self):
        """int: The first empty or nonexistant row in the sheet."""
        if any(cell.value for cell in self.rows[self._ws.max_row - 1]):
            return self._ws.max_row + 1
        else:
            return self._ws.max_row

    @property
    def data_rows(self):
        """tuple: All used rows except the titles (first) row."""
        return self.rows[1:]

    def has_data(self):
        """Return `True` if there is already data in the worksheet.

        This can safely be done by checking to see if any data is in cell A1,
        because there should always be data in cell A1 in sheet that has been
        set up.
        """
        return True if self._ws['A1'].value is not None else False

    def cell(self, row, column):
        """cell: The cell of the worksheet represented by (row, column).

        Since `Worksheet.cell` doesn't use integer coordinates by default, this
        method is here to save time and space when accessing cells, as
        we only use integer values for rows and columns in `SeedsWorksheet`.

        Args:
            row: Row of cell.
            column: Column of cell.
        """
        return self._ws.cell(row=row, column=column)

    def set_column_titles(self, titles):
        """Populate the first row of a worksheet with column titles.

        Args:
            titles: A list of titles to populate the top row with.
        """
        if not self.has_data():
            for c, title in enumerate(titles, start=1):
                self.cell(1, c).value = title
        else:
            raise ValueError('The worksheet \'{0}\' already has data in its '
                             'top row!'.format(self._ws.title))

    def populate_cols_dict(self):
        """Attach a dictionary 'cols' to the sheet for lookup by title.

        The resulting `dict` is formatted: {<title>: <column number>, ...}
        """
        row = self.rows[0]
        d = {cell.value: cell.col_idx for cell in row}
        if d and all(d.keys()):
            self.cols = d
        else:
            raise ValueError('The cols dictionary for the worksheet \'{0}\' '
                             'could not be set because the first row is '
                             'empty, or contains empty cells!'
                             .format(self._ws.title))

    def freeze_title_row(self):
        """Freeze the top row of the worksheet to act as column titles."""
        self._ws.freeze_panes = self._ws['A2']

    def _setup(self, titles=None):
        """Set up a worksheet appropriate for its contents.

        If worksheet is blank, this should set up the title row and instance
            attributes.
        If worksheet contains data, it should just set up the instance
            attributes, as the title rows are already present.

        Args:
            titles: A list of titles to use in the worksheet.
        """
        if not self.has_data():
            if titles:
                self.set_column_titles(titles)
            else:
                raise ValueError('Cannot set up worksheet with no titles!')
        elif titles:
            warnings.warn(
                'Column titles for this worksheet have already been set, so '
                'new titles will not be used. If you would like to set new '
                'titles, please replace this worksheet with a blank one.',
                UserWarning
            )
        self.populate_cols_dict()

    def add_one(self, obj, stream=sys.stdout):
        """Add one object to the first empty row in the worksheet.

        What type of object is valid should be defined in the implementation
        of this method in the child class, and it should raise a `TypeError` if
        given invalid data.

        It should also have the argument 'file' to pass a file-like object to,
        defaulting to `sys.stdout`.
        """
        raise NotImplementedError('This method needs to be implemented by a '
                                  'class derived from SeedsWorksheet.')

    def add(self, objects, stream=sys.stdout):
        """Add database model objects from an iterable.

        We want to add any valid data, and warn if any invalid data is present
        rather than raise an exception, that way all valid data is added.

        Note:
            This method requires use of the abstract class `add_one`; as such,
            it should only be called from child classes of `SeedsWorksheet`.

        Args:
            objects: A list of objects to add data from. The type of objects
                depends on the subclass's implementation of `add_one`.
            stream: Optional IO stream to print messages to.
        """
        print('-- BEGIN adding data to {0}. --'
              .format(self.__class__.__name__), file=stream)
        for obj in objects:
            try:
                self.add_one(obj, stream=stream)
            except TypeError as e:
                warnings.warn(e.args[0], UserWarning)
        print('-- END adding data to {0}. --'
              .format(self.__class__.__name__), file=stream)

    def save_row_to_db(self, row, stream=sys.stdout):
        """Save a row from a worksheet to the database.

        It should take the row number, and optionally an IO stream to write to.

        It should return `False` if no changes are made, otherwise `True`.
        """
        raise NotImplementedError('This method needs to be implemented by a '
                                  'class derived from SeedsWorksheet.')

    def save_to_db(self, stream=sys.stdout):
        """Save all rows of worksheet to the database.

        Note:
            Since range(start, n) goes from start to n-1, we can use
            `active_row` (the first empty row) as n because it won't be
            included in the row numbers generated.

        Args:
            stream: Optional IO stream to write messages to. Defaults to
                `sys.stdout`.
        """
        edited = False
        print('-- BEGIN saving all rows from {0} to database. --'
              .format(self.__class__.__name__), file=stream)
        for r in range(2, self.active_row):
            try:
                if self.save_row_to_db(row=r, stream=stream):
                    edited = True
            except Exception as e:
                db.session.rollback()
                raise RuntimeError('An exception occurred while saving row '
                                   '#{0} to the database, so the database '
                                   'has been rolled back. The exception that '
                                   'was raised: {1}: {2}'
                                   .format(r, e.__class__.__name__, e))
        if edited:
            db.session.commit()
            print('All changes have been committed to the database.',
                  file=stream)
        print('-- END saving all rows from {0} to database. --'
              .format(self.__class__.__name__), file=stream)

    def beautify(self, width=32, height=42):
        """Format a worksheet to be more human readable.

        Args:
            width: Optional width to set column dimensions to.
            height: Optional height to set row dimensions to.
        """
        self._ws.freeze_panes = self._ws['A2']
        for cell in self.rows[0]:
            self._ws.column_dimensions[cell.column].width = width
        a = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        for cell in self._ws.get_cell_collection():
            cell.alignment = a
        for i in range(2, len(self.rows) + 1):
            self._ws.row_dimensions[i].height = height


class IndexesWorksheet(SeedsWorksheet):
    """Class extending `SeedsWorksheet` to have Indexes-specific methods."""
    def setup(self):
        """Set up the Indexes worksheet."""
        if self.has_data():
            self._setup()
        else:
            titles = ('Index', 'Description')
            self._setup(titles)

    def add_one(self, idx, stream=sys.stdout):
        """Add a singe Index object to the Indexes worksheet.

        Args:
            idx: The `Index` object to add.
            stream: Optional IO stream to print messages to.
        """
        if isinstance(idx, Index):
            r = self.active_row
            print('Adding data from {0} to row #{1} of indexes worksheet.'
                  .format(idx, r), file=stream)
            self.cell(r, self.cols['Index']).value = idx.name
            self.cell(r, self.cols['Description']).value = idx.description
        else:
            raise TypeError('The object \'{0}\' could not be added because '
                            'it is not of type \'Index\'!'.format(idx))

    def save_row_to_db(self, row, stream=sys.stdout):
        """Save a row representing in Index to the database.

        Args:
            row: The number of the row to save.
            stream: Optional IO stream to print messages to.

        Returns:
            bool: `True` if changes have been made, `False` if not.
        """
        name = dbify(self.cell(row, self.cols['Index']).value)
        description = self.cell(row, self.cols['Description']).value

        print('-- BEGIN editing/creating Index \'{0}\' from row #{1}. --'
              .format(name, row), file=stream)
        edited = False
        idx = Index.get_or_create(name=name, stream=stream)
        if idx.created:
            edited = True
            db.session.add(idx)
        if description != idx.description:
            edited = True
            if description:
                idx.description = description
                print('Description for the Index \'{0}\' set to: {1}'
                      .format(idx.name, idx.description), file=stream)
            elif idx.description:
                idx.description = None
                print('Description for the Index \'{0}\' has been cleared.'
                      .format(idx.name), file=stream)
        if edited:
            db.session.commit()
            print('Changes to Index \'{0}\' have been flushed to the database.'
                  .format(idx.name), file=stream)
        else:
            print('No changes were made to the Index \'{0}\'.'
                  .format(idx.name), file=stream)
        print('-- END editing/creating Index \'{0}\' from row #{1}. --'
              .format(idx.name, row), file=stream)
        return edited


class CommonNamesWorksheet(SeedsWorksheet):
    """Class extending SeedsWorksheet to have Common Names-specific methods."""
    def setup(self):
        """Set up the Common Names worksheet."""
        if self.has_data():
            self._setup()
        else:
            titles = ('Index',
                      'Common Name',
                      'Description',
                      'Planting Instructions',
                      'Synonyms',
                      'Visible')
            self._setup(titles)

    def add_one(self, cn, stream=sys.stdout):
        """Add a single `CommonName` to a CommonNames worksheet.

        Args:
            cn: The `CommonName` object to add.
            stream: Optional IO stream to print messages to.
        """
        if isinstance(cn, CommonName):
            r = self.active_row
            print('Adding data from {0} to row #{1} of common names worksheet.'
                  .format(cn, r), file=stream)
            self.cell(r, self.cols['Index']).value = cn.index.name
            self.cell(r, self.cols['Common Name']).value = cn.name
            if cn.description:
                self.cell(
                    r, self.cols['Description']
                ).value = cn.description
            if cn.instructions:
                self.cell(
                    r, self.cols['Planting Instructions']
                ).value = cn.instructions
            syns = cn.synonyms_string
            if syns:
                self.cell(r, self.cols['Synonyms']).value = syns
            v_cell = self.cell(r, self.cols['Visible'])
            v_cell.value = 'True' if cn.visible else 'False'
        else:
            raise TypeError('The object \'{0}\' could not be added because '
                            'it is not of type \'CommonName\'!'.format(cn))

    def save_row_to_db(self, row, stream=sys.stdout):
        """Save a row from the Common Names sheet to the database.

        Args:
            row: The number of the row to save.
            stream: Optional IO stream to print messages to.

        Returns:
            bool: `True` if changes have been made, `False` if not.
        """
        index = dbify(self.cell(row, self.cols['Index']).value)
        name = dbify(self.cell(row, self.cols['Common Name']).value)
        description = self.cell(row, self.cols['Description']).value
        instructions = self.cell(row, self.cols['Planting Instructions']).value
        synonyms = self.cell(row, self.cols['Synonyms']).value
        if not synonyms:
            synonyms = ''  # Match result of CommonName.synonyms_string
        vis = self.cell(row, self.cols['Visible']).value
        if vis and 'true' in vis.lower():
            visible = True
        else:
            visible = False

        print('-- BEGIN editing/creating CommonName \'{0}\' from row #{1}. --'
              .format(name, row), file=stream)
        edited = False
        cn = CommonName.get_or_create(name=name, index=index, stream=stream)
        if cn.created:
            edited = True
            db.session.add(cn)
        if description != cn.description:
            edited = True
            if description:
                cn.description = description
                print('Description for the CommonName \'{0}\' set to: {1}'
                      .format(cn.name, cn.description), file=stream)
            elif cn.description:
                cn.description = None
                print('Description for the CommonName \'{0}\' has been '
                      'cleared.'.format(cn.name), file=stream)
        if instructions != cn.instructions:
            edited = True
            if instructions:
                cn.instructions = instructions
                print('Planting instructions for the CommonName \'{0}\' set '
                      'to: {1}'.format(cn.name, cn.instructions), file=stream)
            elif cn.instructions:
                cn.instructions = None
                print('Planting instructions for the CommonName \'{0}\' have '
                      'been cleared.'.format(cn.name), file=stream)
        if synonyms != cn.synonyms_string:
            edited = True
            cn.synonyms_string = synonyms
            if synonyms:
                print('Synonyms for the CommonName \'{0}\' set to: {1}'
                      .format(cn.name, cn.synonyms_string), file=stream)
            else:
                print('Synonyms for the CommonName \'{0}\' have been cleared.'
                      .format(cn.name), file=stream)
        if visible != cn.visible:
            edited = True
            cn.visible = visible
            if cn.visible:
                print('The CommonName \'{0}\' is visible on generated '
                      'pages.'.format(cn.name), file=stream)
            else:
                print('The CommonName \'{0}\' is not visible on generated '
                      'pages.'.format(cn.name), file=stream)
        if edited:
            db.session.flush()
            print('Changes to the CommonName \'{0}\' have been flushed to the '
                  'database.'.format(cn.name), file=stream)
        else:
            print('No changes were made to the CommonName \'{0}\'.'
                  .format(cn.name), file=stream)
        print('-- END editing/creating CommonName \'{0}\' from row #{1}. --'
              .format(cn.name, row), file=stream)
        return edited


class BotanicalNamesWorksheet(SeedsWorksheet):
    """Class extending SeedsWorksheet with Botanical Names-specific methods."""
    def setup(self):
        """Set up the Botanical Names worksheet."""
        if self.has_data():
            self._setup()
        else:
            titles = ('Common Names (JSON)',
                      'Botanical Name',
                      'Synonyms')
            self._setup(titles)

    def add_one(self, bn, stream=sys.stdout):
        """Add a single BotanicalName to the Botanical Names worksheet.

        Args:
            bn: The `BotanicalName` to add.
            stream: Optional IO stream to print messages to.
        """
        if isinstance(bn, BotanicalName):
            r = self.active_row
            print('Adding data from {0} to row #{1} of botanical names '
                  'worksheet.'.format(bn, r), file=stream)
            self.cell(
                r, self.cols['Common Names (JSON)']
            ).value = queryable_dicts_to_json(bn.common_names)
            self.cell(r, self.cols['Botanical Name']).value = bn.name
            syns = bn.synonyms_string
            if syns:
                self.cell(r, self.cols['Synonyms']).value = syns
        else:
            raise TypeError('The object \'{0}\' could not be added because '
                            'it is not of type \'BotanicalName\'!'.format(bn))

    def save_row_to_db(self, row, stream=sys.stdout):
        """Save a row from the Botanical Names sheet to the database.

        Args:
            row: The number of the row to save.
            stream: Optional IO stream to print messages to.

        Returns:
            bool: `True` if changes have been made, `False` if not.
        """
        botanical_name = self.cell(row, self.cols['Botanical Name']).value
        cn_json = self.cell(row, self.cols['Common Names (JSON)']).value
        cn_dicts = json.loads(cn_json)
        synonyms = self.cell(row, self.cols['Synonyms']).value
        if not synonyms:
            synonyms = ''

        if not BotanicalName.validate(botanical_name):
            print('Could not add the BotanicalName \'{0}\' because it does '
                  'not appear to be a validly formatted botanical name.'
                  .format(botanical_name), file=stream)
            return False

        print('-- BEGIN editing/creating BotanicalName \'{0}\' from row #{1}. '
              '--'.format(botanical_name, row), file=stream)
        edited = False
        bn = BotanicalName.query\
            .filter(BotanicalName.name == botanical_name)\
            .one_or_none()
        if bn:
            print('The BotanicalName \'{0}\' has been loaded from the '
                  'database.'.format(bn.name), file=stream)
        else:
            edited = True
            bn = BotanicalName(name=botanical_name)
            db.session.add(bn)
            print('The BotanicalName \'{0}\' does not yet exist in the '
                  'database, so it has been created.'.format(bn.name),
                  file=stream)
        cns = tuple(CommonName.get_or_create(
            name=dbify(d['Common Name']),
            index=dbify(d['Index']),
            stream=stream
        ) for d in cn_dicts)
        for cn in cns:
            if cn not in bn.common_names:
                edited = True
                bn.common_names.append(cn)
                print('The CommonName \'{0}\' has been added to CommonNames '
                      'for the BotanicalName \'{1}\'.'
                      .format(cn.name, bn.name), file=stream)
        for cn in list(bn.common_names):
            if cn not in cns:
                edited = True
                bn.common_names.remove(cn)
                print('The CommonName \'{0}\' has been removed from '
                      'CommonNames for the BotanicalName \'{1}\'.'
                      .format(cn.name, bn.name), file=stream)
        if synonyms != bn.synonyms_string:
            edited = True
            bn.synonyms_string = synonyms
            if synonyms:
                print('Synonyms for the BotanicalName \'{0}\' set to: {1}'
                      .format(bn.name, bn.synonyms_string), file=stream)
            else:
                print('Synonyms for the BotanicalName \'{0}\' have been '
                      'cleared.'.format(bn.name), file=stream)
        if edited:
            db.session.flush()
            print('Changes to the BotanicalName \'{0}\' have been flushed to '
                  'the database.'.format(bn.name), file=stream)
        else:
            print('No changes were made to the BotanicalName \'{0}\'.'
                  .format(bn.name), file=stream)
        print('-- END editing/creating BotanicalName \'{0}\' from row #{1}. '
              '--'.format(bn.name, row), file=stream)
        return edited


class SectionsWorksheet(SeedsWorksheet):
    """Class extending SeedsWorksheet with Section-specific methods."""
    def setup(self):
        """Set up the Section worksheet."""
        if self.has_data():
            self._setup()
        else:
            titles = ('Common Name (JSON)',
                      'Section',
                      'Description')
            self._setup(titles)

    def add_one(self, sec, stream=sys.stdout):
        """Add a single Section to the Section worksheet.

        Args:
            sec: The `Section` to add.
            stream: Optional IO stream to print messages to.
        """
        if isinstance(sec, Section):
            r = self.active_row
            print('Adding data from {0} to row #{1} of sections worksheet.'
                  .format(sec, r), file=stream)
            self.cell(
                r, self.cols['Common Name (JSON)']
            ).value = json.dumps(sec.common_name.queryable_dict)
            self.cell(r, self.cols['Section']).value = sec.name
            if sec.description:
                self.cell(r, self.cols['Description']).value = sec.description
        else:
            raise TypeError('The object \'{0}\' could not be added because '
                            'it is not of type \'Section\'!'.format(sec))

    def save_row_to_db(self, row, stream=sys.stdout):
        """Save a row from the Common Names sheet to the database.

        Args:
            row: The number of the row to save.
            stream: Optional IO stream to print messages to.

        Returns:
            bool: `True` if changes have been made, `False` if not.
        """
        cn_json = self.cell(row, self.cols['Common Name (JSON)']).value
        cn_dict = json.loads(cn_json)
        section = dbify(self.cell(row, self.cols['Section']).value)
        description = self.cell(row, self.cols['Description']).value

        print('-- BEGIN editing/creating Section \'{0}\' from row #{1}. '
              '--'.format(section, row), file=stream)
        edited = False
        cn = CommonName.get_or_create(name=dbify(cn_dict['Common Name']),
                                      index=dbify(cn_dict['Index']),
                                      stream=stream)
        sec = None
        if not cn.created:
            sec = Section.query\
                .filter(Section.name == section, Section.common_name_id == cn.id)\
                .one_or_none()
        if sec:
            print('The Section \'{0}\' has been loaded from the database.'
                  .format(sec.name), file=stream)
        else:
            edited = True
            sec = Section(name=section)
            sec.common_name = cn
            print('CommonName for the Section \'{0}\' set to: {1}'
                  .format(sec.name, cn.name), file=stream)
            db.session.add(sec)
            print('The Section \'{0}\' does not yet exist in the database, '
                  'so it has been created.'.format(sec.name), file=stream)
        if description != sec.description:
            edited = True
            if description:
                sec.description = description
                print('Description for the Section \'{0}\' set to: {1}'
                      .format(sec.name, sec.description), file=stream)
            else:
                sec.description = None
                print('Description for the Section \'{0}\' has been cleared.'
                      .format(sec.name), file=stream)
        if edited:
            db.session.flush()
            print('Changes to the Section \'{0}\' have been flushed to '
                  'the database.'.format(sec.name), file=stream)
        else:
            print('No changes were made to the Section \'{0}\'.'
                  .format(sec.name), file=stream)
        print('-- END editing/creating Section \'{0}\' from row #{1}. '
              '--'.format(sec.name, row), file=stream)
        return edited


class CultivarsWorksheet(SeedsWorksheet):
    """Class extending SeedsWorksheet with Cultivars-specific methods."""
    def setup(self):
        """Set up the Cultivars worksheet."""
        if self.has_data():
            self._setup()
        else:
            titles = ('Index',
                      'Common Name',
                      'Cultivar Name',
                      'Section',
                      'Botanical Name',
                      'Thumbnail Filename',
                      'Description',
                      'Synonyms',
                      'New Until',
                      'In Stock',
                      'Active',
                      'Visible')
            self._setup(titles)

    def add_one(self, cv, stream=sys.stdout):
        """Add a single Cultivar to the Cultivars worksheet.

        Args:
            cv: The `Cultivar` to add.
            stream: Optional IO stream to print messages to.
        """
        if isinstance(cv, Cultivar):
            r = self.active_row
            print('Adding data from {0} to row #{1} of cultivars worksheet.'
                  .format(cv, r), file=stream)
            self.cell(r, self.cols['Index']).value = cv.common_name.index.name
            self.cell(r, self.cols['Common Name']).value = cv.common_name.name
            self.cell(r, self.cols['Cultivar Name']).value = cv.name
            if cv.section:
                self.cell(r, self.cols['Section']).value = cv.section.name
            if cv.botanical_name:
                self.cell(
                    r, self.cols['Botanical Name']
                ).value = cv.botanical_name.name
            if cv.thumbnail:
                self.cell(
                    r, self.cols['Thumbnail Filename']
                ).value = cv.thumbnail.filename
            if cv.description:
                self.cell(r, self.cols['Description']).value = cv.description
            syns = cv.synonyms_string
            if syns:
                self.cell(r, self.cols['Synonyms']).value = syns
            if cv.new_until:
                self.cell(
                    r, self.cols['New Until']
                ).value = cv.new_until.strftime('%m/%d/%Y')
            is_cell = self.cell(r, self.cols['In Stock'])
            is_cell.value = 'True' if cv.in_stock else 'False'
            act_cell = self.cell(r, self.cols['Active'])
            act_cell.value = 'True' if cv.active else 'False'
            v_cell = self.cell(r, self.cols['Visible'])
            v_cell.value = 'True' if cv.visible else 'False'
        else:
            raise TypeError('The object \'{0}\' could not be added because '
                            'it is not of type \'Cultivar\'!'.format(cv))

    def save_row_to_db(self, row, stream=sys.stdout):
        """Save a row from the Cultivars sheet to the database.

        Args:
            row: The number of the row to save.
            stream: Optional IO stream to print messages to.

        Returns:
            bool: `True` if changes have been made, `False` if not.
        """
        index = dbify(self.cell(row, self.cols['Index']).value)
        common_name = dbify(self.cell(row, self.cols['Common Name']).value)
        cultivar = dbify(self.cell(row, self.cols['Cultivar Name']).value)
        section = dbify(self.cell(row, self.cols['Section']).value)
        if not section:
            section = None
        botanical_name = self.cell(row, self.cols['Botanical Name']).value
        thumbnail = self.cell(row, self.cols['Thumbnail Filename']).value
        description = self.cell(row, self.cols['Description']).value
        synonyms = self.cell(row, self.cols['Synonyms']).value
        if not synonyms:
            synonyms = ''
        nus = self.cell(row, self.cols['New Until']).value
        if nus:
            new_until = datetime.datetime.strptime(nus, '%m/%d/%Y').date()
        else:
            new_until = None
        n_stk = self.cell(row, self.cols['In Stock']).value
        if n_stk and 'true' in n_stk.lower():
            in_stock = True
        else:
            in_stock = False
        act = self.cell(row, self.cols['Active']).value
        if act and 'true' in act.lower():
            active = True
        else:
            active = False
        vis = self.cell(row, self.cols['Visible']).value
        if vis and 'true' in vis.lower():
            visible = True
        else:
            visible = False

        print('-- BEGIN editing/creating Cultivar \'{0}\' from row #{1}. '
              '--'.format(cultivar + ' ' + common_name, row), file=stream)
        edited = False
        cv = Cultivar.get_or_create(name=cultivar,
                                    index=index,
                                    common_name=common_name,
                                    stream=stream)
        if cv.created:
            edited = True
            db.session.add(cv)
            if section:  # Section already exists if cv was not created.
                sec = Section.query\
                    .join(CommonName, CommonName.id == Section.common_name_id)\
                    .join(Index, Index.id == CommonName.index_id)\
                    .filter(Section.name == section,
                            CommonName.name == common_name,
                            Index.name == index)\
                    .one_or_none()
                if sec:
                    print('The Section \'{0}\' has been loaded from the '
                          'database.'.format(sec.name), file=stream)
                else:
                    sec = Section(name=section)
                    sec.common_name = cv.common_name
                    print('The Section \'{0}\' does not yet exist, so it has '
                          'been created.'.format(sec.name), file=stream)
                cv.section = sec
                print('Section for the Cultivar \'{0}\' set to: {1}'
                      .format(cv.fullname, sec.name), file=stream)
        if botanical_name:
            if not BotanicalName.validate(botanical_name):
                obn = botanical_name
                words = botanical_name.strip().split(' ')
                words[0] = words[0].capitalize()
                botanical_name = ' '.join(words)
                print('The BotanicalName \'{0}\' does not appear to be a '
                      'validly formatted botanical name. In an attempt to fix '
                      'it, it has been changed to: \'{1}\''
                      .format(obn, botanical_name), file=stream)
            bn = BotanicalName.query\
                .filter(BotanicalName.name == botanical_name)\
                .one_or_none()
            if bn and bn is not cv.botanical_name:
                print('The BotanicalName \'{0}\' has been loaded from the '
                      'database.'.format(bn.name), file=stream)
            elif not bn:
                bn = BotanicalName(name=botanical_name)
                bn.common_names.append(cv.common_name)
                print('The BotanicalName \'{0}\' does not yet exist, so it '
                      'has been created.'.format(bn.name), file=stream)
            if bn is not cv.botanical_name:
                edited = True
                cv.botanical_name = bn
                print('BotanicalName for the Cultivar \'{0}\' set to: {1}'
                      .format(cv.fullname, bn.name), file=stream)
        if thumbnail:
            if not cv.thumbnail or cv.thumbnail.filename != thumbnail:
                edited = True
                tn = Image.query\
                    .filter(Image.filename == thumbnail)\
                    .one_or_none()
                if tn:
                    print('The Image with the filename \'{0}\' has been '
                          'loaded from the database.'.format(tn.filename),
                          file=stream)
                else:
                    tn = Image(filename=thumbnail)
                    print('The Image with the filename \'{0}\' does not yet '
                          'exist in the database, so it has been created.'
                          .format(tn.filename), file=stream)
                cv.thumbnail = tn
                print('The Image with the filename \'{0}\' has been set as '
                      'the thumbnail for the Cultivar \'{1}\'.'
                      .format(tn.filename, cv.fullname), file=stream)
                if not tn.exists():
                    print('WARNING: The image file \'{0}\' set as the '
                          'thumbnail for the Cultivar \'{1}\' does not exist! '
                          'Please make sure you add the image file to the '
                          'images directory.'.format(tn.filename, cv.fullname),
                          file=stream)
        if description != cv.description:
            edited = True
            if description:
                cv.description = description
                print('Description for the Cultivar \'{0}\' set to: {1}'
                      .format(cv.fullname, cv.description), file=stream)
            else:
                cv.description = None
                print('Description for the Cultivar \'{0}\' has been cleared.'
                      .format(cv.fullname), file=stream)
        if synonyms != cv.synonyms_string:
            edited = True
            cv.synonyms_string = synonyms
            if synonyms:
                print('Synonyms for the Cultivar \'{0}\' set to: {1}'
                      .format(cv.fullname, cv.synonyms_string),
                      file=stream)
            else:
                print('Synonyms for the Cultivar \'{0}\' have been cleared.'
                      .format(cv.fullname), file=stream)
        if new_until != cv.new_until:
            edited = True
            if new_until:
                cv.new_until = new_until
                print('The Cultivar \'{0}\' has been set as new until {1}.'
                      .format(cv.fullname, cv.new_until.strftime('%m/%d/%Y')),
                      file=stream)
            else:
                cv.new_until = None
                print('The Cultivar \'{0}\' is no longer set as new.'
                      .format(cv.fullname), file=stream)
        if in_stock != cv.in_stock:
            edited = True
            cv.in_stock = in_stock
            if cv.in_stock:
                print('The Cultivar \'{0}\' is in stock.'.format(cv.fullname),
                      file=stream)
            else:
                print('The Cultivar \'{0}\' is out of stock.'
                      .format(cv.fullname), file=stream)
        if active != cv.active:
            edited = True
            cv.active = active
            if cv.active:
                print('The Cultivar \'{0}\' is active.'.format(cv.fullname),
                      file=stream)
            else:
                print('The Cultivar \'{0}\' is inactive.'.format(cv.fullname),
                      file=stream)
        if visible != cv.visible:
            edited = True
            cv.visible = visible
            if cv.visible:
                print('The Cultivar \'{0}\' will be shown on auto-generated '
                      'pages.'.format(cv.fullname), file=stream)
            else:
                print('The Cultivar \'{0}\' will not be shown on '
                      'auto-generated pages.'.format(cv.fullname), file=stream)
        if edited:
            db.session.flush()
            print('Changes to the Cultivar \'{0}\' have been flushed to '
                  'the database.'.format(cv.fullname), file=stream)
        else:
            print('No changes were made to the Cultivar \'{0}\'.'
                  .format(cv.fullname), file=stream)
        print('-- END editing/creating Cultivar \'{0}\' from row #{1}. '
              '--'.format(cv.fullname, row), file=stream)
        return edited


class PacketsWorksheet(SeedsWorksheet):
    """Class extending SeedsWorksheet with Packets-specific methods."""
    def setup(self):
        """Set up the packets worksheet."""
        if self.has_data():
            self._setup()
        else:
            titles = ('Cultivar (JSON)',
                      'SKU',
                      'Price',
                      'Quantity',
                      'Units')
            self._setup(titles)

    def add_one(self, pkt, stream=sys.stdout):
        """Add a single `Packet` to the Packets worksheet.

        Args:
            pkt: The `Packet` to add data from
            stream: Optional IO stream to print messages to.
        """
        if isinstance(pkt, Packet):
            r = self.active_row
            print('Adding data from {0} to row #{1} of packets worksheet.'
                  .format(pkt, r), file=stream)
            self.cell(
                r, self.cols['Cultivar (JSON)']
            ).value = json.dumps(pkt.cultivar.queryable_dict)
            self.cell(r, self.cols['SKU']).value = pkt.sku
            self.cell(r, self.cols['Price']).value = str(pkt.price)
            self.cell(r, self.cols['Quantity']).value = pkt.quantity.str_value
            self.cell(r, self.cols['Units']).value = pkt.quantity.units
        else:
            raise TypeError('The object \'{0}\' could not be added because '
                            'it is not of type \'Packet\'!'.format(pkt))

    def save_row_to_db(self, row, stream=sys.stdout):
        """Save a row from the Packets sheet to the database.

        Args:
            row: The number of the row to save.
            stream: Optional IO stream to print messages to.

        Returns:
            bool: `True` if changes have been made, `False` if not.
        """
        cultivar_json = self.cell(row, self.cols['Cultivar (JSON)']).value
        cv_dict = json.loads(cultivar_json)
        sku = self.cell(row, self.cols['SKU']).value
        price = self.cell(row, self.cols['Price']).value
        quantity = self.cell(row, self.cols['Quantity']).value
        units = self.cell(row, self.cols['Units']).value

        print('-- BEGIN editing/creating Packet with the SKU \'{0}\' from row '
              '#{1}. --'.format(sku, row), file=stream)
        edited = False
        pkt = Packet.query.filter(Packet.sku == sku).one_or_none()
        if pkt:
            print('The Packet with SKU \'{0}\' has been loaded from the '
                  'database.'.format(pkt.sku), file=stream)
        else:
            edited = True
            qty = Quantity.from_queryable_values(value=quantity, units=units)
            if not qty:
                qty = Quantity(value=quantity, units=units)
            pkt = Packet(sku=sku, price=price, quantity=qty)
            db.session.add(pkt)
            pkt.cultivar = Cultivar.get_or_create(
                name=dbify(cv_dict['Cultivar Name']),
                common_name=dbify(cv_dict['Common Name']),
                index=dbify(cv_dict['Index']),
                stream=stream
            )
            print('The Packet with SKU \'{0}\' does not yet exist, so it has '
                  'been created.'.format(pkt.sku), file=stream)
        if price != str(pkt.price):
            edited = True
            pkt.price = price
            print('The price for Packet SKU \'{0}\' has been set to: ${1}.'
                  .format(pkt.sku, pkt.price), file=stream)
        qty = Quantity.from_queryable_values(value=quantity, units=units)
        if not qty:
            qty = Quantity(value=quantity, units=units)
        if qty is not pkt.quantity:
            edited = True
            pkt.quantity = qty
            print('The quantity for the Packet SKU \'{0}\' has been set to: '
                  '{1} {2}'.format(pkt.sku, qty.value, qty.units), file=stream)
        if edited:
            db.session.flush()
            print('Changes to the Packet \'{0}\' have been flushed to '
                  'the database.'.format(pkt.info), file=stream)
        else:
            print('No changes were made to the Packet \'{0}\'.'
                  .format(pkt.info), file=stream)
        print('-- END editing/creating Packet with SKU \'{0}\' from row #{1}. '
              '--'.format(pkt.sku, row), file=stream)
        return edited


class SeedsWorkbook(object):
    """A container for an `openpyxl` workbook."""
    def __init__(self):
        self._wb = openpyxl.Workbook()
        self.create_all_sheets()

    def __getitem__(self, x):
        return self._wb[x]

    def remove_all_sheets(self):
        """Remove all worksheets from the workbook.

        This generally should only be run during creation of a `SeedsWorkbook`,
        or before loading a workbook into a `SeedsWorkbook`.
        """
        for sheet in list(self._wb.worksheets):
            self._wb.remove_sheet(sheet)

    def create_all_sheets(self):
        """Create all of the worksheets in the `SeedsWorkbook`."""
        self.remove_all_sheets()
        self.indexes = IndexesWorksheet(self._wb.create_sheet(title='Indexes'))
        self.indexes.setup()
        self.common_names = CommonNamesWorksheet(
            self._wb.create_sheet(title='Common Names')
        )
        self.common_names.setup()
        self.botanical_names = BotanicalNamesWorksheet(
            self._wb.create_sheet(title='Botanical Names')
        )
        self.botanical_names.setup()
        self.section = SectionsWorksheet(
            self._wb.create_sheet(title='Section')
        )
        self.section.setup()
        self.cultivars = CultivarsWorksheet(
            self._wb.create_sheet(title='Cultivars')
        )
        self.cultivars.setup()
        self.packets = PacketsWorksheet(self._wb.create_sheet(title='Packets'))
        self.packets.setup()

    def load_all_sheets_from_workbook(self):
        """Set up all SeedsWorksheets with sheets from loaded workbook."""
        self.indexes = IndexesWorksheet(self._wb['Indexes'])
        self.indexes.setup()
        self.common_names = CommonNamesWorksheet(self._wb['Common Names'])
        self.common_names.setup()
        self.botanical_names = BotanicalNamesWorksheet(
            self._wb['Botanical Names']
        )
        self.botanical_names.setup()
        self.section = SectionsWorksheet(self._wb['Section'])
        self.section.setup()
        self.cultivars = CultivarsWorksheet(self._wb['Cultivars'])
        self.cultivars.setup()
        self.packets = PacketsWorksheet(self._wb['Packets'])
        self.packets.setup()

    def add_all_data_to_sheets(self, stream=sys.stdout):
        """Add all relevant data from the database to respective worksheets.

        Args:
            stream: Optional IO stream to print messages to.
        """
        self.indexes.add(Index.query.all(), stream=stream)
        self.common_names.add(CommonName.query.all(), stream=stream)
        self.botanical_names.add(BotanicalName.query.all(), stream=stream)
        self.section.add(Section.query.all(), stream=stream)
        self.cultivars.add(Cultivar.query.all(), stream=stream)
        self.packets.add(Packet.query.all(), stream=stream)

    def save_all_sheets_to_db(self, stream=sys.stdout):
        """Save the contents of all worksheets to the database.

        Args:
            stream: Optional IO stream to print messages to.
        """
        print('-- BEGIN saving all worksheets to database. --', file=stream)
        self.indexes.save_to_db(stream=stream)
        self.common_names.save_to_db(stream=stream)
        self.botanical_names.save_to_db(stream=stream)
        self.section.save_to_db(stream=stream)
        self.cultivars.save_to_db(stream=stream)
        self.packets.save_to_db(stream=stream)
        print('-- END saving all worksheets to database. --', file=stream)

    def beautify_all_sheets(self, width=32, height=42):
        """Run beautify on all worksheets.

        Args:
            width: Optional column width.
            height: Optional row height.
        """
        self.indexes.beautify(width=width, height=height)
        self.common_names.beautify(width=width, height=height)
        self.botanical_names.beautify(width=width, height=height)
        self.section.beautify(width=width, height=height)
        self.cultivars.beautify(width=width, height=height)
        self.packets.beautify(width=width, height=height)

    def load(self, filename):
        self._wb = openpyxl.load_workbook(filename)
        self.load_all_sheets_from_workbook()

    def save(self, filename):
        self._wb.save(filename)
