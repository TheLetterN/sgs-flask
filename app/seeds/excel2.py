import json
import openpyxl
from app.seeds.models import (
    BotanicalName,
    CommonName,
    Index
)


def lookup_dicts_to_json(items):
    """Return a JSON string of lookup_dicts from a list of objects.
    
    Obviously, the objects need the lookup_dict() method, so this function
    is only used for groups of CommonName or Cultivar objects.
    """
    return json.dumps(tuple(it.lookup_dict() for it in items))


class SeedsWorksheet(object):
    """A container for an openpyxl worksheet.
    
    Since extending openpyxl's classes seems to be an exercise in futility, it 
    is easier to just encapsulate them and create an interface that's specific
    to how we want our worksheet data formatted.
    """
    def __init__(self, sheet):
        self._ws = sheet

    def __getitem__(self, x):
        """Allow direct access to keys present in worksheet."""
        return self._ws[x]

    @property
    def title(self):
        return self._ws.title

    @title.setter
    def title(self, title):
        self._ws.title = title

    @property
    def rows(self):
        """tuple: A tuple listing each row as a tuple of cells.
        
        I'm fiendishly using the protected _cells variable here because as of
        writing this the rows property in Worksheet returns ((),) even if there
        is data in cell A1, while Worksheet._cells is empty until at least one
        cell is given a value.
        """
        # TODO: Forward _ws.rows once rows is fixed in openpyxl.
        if self._ws._cells:
            return tuple(self._ws.iter_rows())
        else:
            return ((),)

    @property
    def active_row(self):
        """int: The first empty or nonexistant row in the sheet."""
        if any(cell.value for cell in self.rows[self._ws.max_row - 1]):
            return self._ws.max_row + 1
        else:
            return self._ws.max_row

    @property
    def data_rows(self):
        """tuple: All used rows except the titles (first) row."""
        return self.rows[1:]

    def cell(self, row, column):
        """cell: The cell of the worksheet represented by (row, column).

        Since Worksheet.cell doesn't use integer coordinates by default, this
        method is here to save time and space when accessing cells, as 
        we only use integer values for rows and columns within SeedsWorksheet.

        Args:
            row (int): Row of cell.
            column (int): Column of cell.
        """
        return self._ws.cell(row=row, column=column)

    def set_column_titles(self, titles):
        """Populate the first row of a worksheet with column titles."""
        if not self._ws.rows[0]:
            self._ws.append(titles)
        else:
            raise ValueError('The worksheet \'{0}\' already has data in its '
                             'top row!'.format(self._ws.title))

    def populate_cols_dict(self):
        """Attach a dictionary 'cols' to the sheet for lookup by title.

        The resulting dict is in the format of {<title>: <column number>, ...}
        """
        # TODO: use rows instead of iter_rows once rows is fixed in openpyxl.
        row = next(self._ws.iter_rows())  # First row.
        d = {cell.value: cell.col_idx for cell in row}
        if all(d.keys()):
            self.cols = d
        else:
            raise ValueError('The cols dictionary for the worksheet \'{0}\' '
                             'could not be set because the first row is '
                             'empty, or contains empty cells!'
                             .format(self._ws.title))

    def freeze_title_row(self):
        """Freeze the top row of the worksheet to act as column titles."""
        self._ws.freeze_panes = self._ws['A2']



class SeedsWorkbook(object):
    """A container for an openpyxl workbook."""
    def __init__(self):
        self._wb = openpyxl.Workbook()
        self.create_all_sheets()

    def __getitem__(self, x):
        return self._wb[x]

    # Indexes sheet methods
    def create_indexes_sheet(self):
        """Set up the Indexes worksheet."""
        self.indexes = SeedsWorksheet(self._wb.create_sheet('Indexes'))
        self.indexes.set_column_titles(('Index', 'Description'))
        self.indexes.populate_cols_dict()

    def load_indexes_from_workbook(self):
        """Set up self.indexes with sheet from loaded workbook."""
        self.indexes = SeedsWorksheet(self._wb['Indexes'])
        self.indexes.populate_cols_dict()

    def add_index(self, idx):
        """Add an Index object to the Indexes worksheet."""
        if isinstance(idx, Index):
            sws = self.indexes
            r = sws.active_row
            sws.cell(r, sws.cols['Index']).value = idx.name
            sws.cell(r, sws.cols['Description']).value = idx.description
        else:
            raise TypeError('add_index can only take an object of type Index!')

    def add_indexes(self, indexes):
        """Add a set of indexes to the Indexes worksheet."""
        for idx in indexes:
            self.add_index(self, idx)

    # Common Names sheet methods
    def create_common_names_sheet(self):
        """Set up the Common Names worksheet."""
        self.common_names = SeedsWorksheet(
            self._wb.create_sheet(title='Common Names')
        )
        self.common_names.set_column_titles(
            ('Index',
             'Common Name',
             'Subcategory of',
             'Description',
             'Planting Instructions',
             'Synonyms',
             'Invisible',
             'Grows With Common Names (JSON)',
             'Grows With Cultivars (JSON)')
        )
        self.common_names.populate_cols_dict()

    def load_common_names_from_workbook(self):
        """Set up self.common_names with sheet from loaded workbook."""
        self.common_names = SeedsWorksheet(self._wb['Common Names'])
        self.common_names.populate_cols_dict()

    def add_common_name(self, cn):
        """Add a CommonName object to the Common Names worksheet."""
        if isinstance(cn, CommonName):
            sws = self.common_names
            r = sws.active_row
            sws.cell(r, sws.cols['Index']).value = cn.index.name
            sws.cell(r, sws.cols['Common Name']).value = cn.name
            if cn.parent:
                sws.cell(r, sws.cols['Subcategory of']).value = cn.parent.name
            if cn.description:
                sws.cell(r, sws.cols['Description']).value = cn.description
            if cn.instructions:
                sws.cell(
                    r, sws.cols['Planting Instructions']
                ).value = cn.instructions
            syns = cn.get_synonyms_string()
            if syns:
                sws.cell(r, sws.cols['Synonyms']).value = syns
            if cn.invisible:
                sws.cell(r, sws.cols['Invisible']).value = 'True'
            if cn.gw_common_names:
                sws.cell(
                    r, sws.cols['Grows With Common Names (JSON)']
                ).value = lookup_dicts_to_json(cn.gw_common_names) 
            if cn.gw_cultivars:
                sws.cell(
                    r, sws.cols['Grows With Cultivars (JSON)']
                ).value = lookup_dicts_to_json(cn.gw_cultivars)
        else:
            raise TypeError('add_common_name can only take an object of type '
                            'CommonName!')

    def add_common_names(self, cns):
        """Add multiple CommonName objects to the Common Names worksheet."""
        for cn in cns:
            self.add_common_name(cn)

    # Botanical Names sheet methods
    def create_botanical_names_sheet(self):
        """Set up the Botanical Names worksheet."""
        self.botanical_names = SeedsWorksheet(
            self._wb.create_sheet(title='Botanical Names')
        )
        self.botanical_names.set_column_titles(
            ('Common Names (JSON)',
             'Botanical Name',
             'Synonyms')
        )
        self.botanical_names.populate_cols_dict()

    def load_botanical_names_from_workbook(self):
        """Set up self.botanical_names with sheet from loaded workbook."""
        self.botanical_names = SeedsWorksheet(self._wb['Botanical Names'])
        self.botanical_names.populate_cols_dict()

    def add_botanical_name(self, bn):
        """Add a BotanicalName object to the Botanical Names sheet."""
        if isinstance(bn, BotanicalName):
            sws = self.botanical_names
            r = sws.active_row
            sws.cell(
                r, sws.cols['Common Names (JSON)']
            ).value = lookup_dicts_to_json(bn.common_names)
            sws.cell(r, sws.cols['Botanical Name']).value = bn.name
            syns = bn.get_synonyms_string()
            if syns:
                sws.cell(r, sws.cols['Synonyms']).value = syns
        else:
            raise TypeError('add_botanical_name can only take an object of '
                            'type BotanicalName!')

    def add_botanical_names(self, bns):
        """Add multiple BotanicalName objects to the Botanical Names sheet."""
        for bn in bns:
            self.add_botanical_name(bn)

    # Series sheet methods
    def create_series_sheet(self):
        """Set up the Series worksheet."""
        self.series = SeedsWorksheet(self._wb.create_sheet(title='Series'))
        self.series.set_column_titles(
            ('Common Name (JSON)',
             'Series',
             'Position',
             'Description')
        )
        self.series.populate_cols_dict()

    def load_series_from_workbook(self):
        """Set up self.series with sheet from loaded workbook."""
        self.series = SeedsWorksheet(self._wb['Series'])
        self.series.populate_cols_dict()

    # Cultivar sheet methods
    def create_cultivars_sheet(self):
        """Set up the Cultivars worksheet."""
        self.cultivars = SeedsWorksheet(
            self._wb.create_sheet(title='Cultivars')
        )
        self.cultivars.set_column_titles(
            ('Index',
             'Common Name',
             'Botanical Name',
             'Series',
             'Cultivar Name',
             'Thumbnail Filename',
             'Description',
             'Synonyms',
             'New For',
             'In Stock',
             'Active',
             'Invisible',
             'Grows With Common Names (JSON)',
             'Grows With Cultivars (JSON)')
        )
        self.cultivars.populate_cols_dict()

    def load_cultivars_from_workbook(self):
        """Set up self.cultivars with sheet from loaded workbook."""
        self.cultivars = SeedsWorksheet(self._wb['Cultivars'])
        self.series.populate_cols_dict()

    # Packets sheet methods
    def create_packets_sheet(self):
        """Set up the Packets worksheet."""
        self.packets = SeedsWorksheet(
            self._wb.create_sheet(title='Packets')
        )
        self.packets.set_column_titles(
            ('Cultivar (JSON)',
             'SKU',
             'Price',
             'Quantity',
             'Units')
        )
        self.packets.populate_cols_dict()

    def load_packets_from_workbook(self):
        """Set up self.packets with sheet from loaded workbook."""
        self.packets = SeedsWorksheet(self._wb['Packets'])
        self.packets.populate_cols_dict()

    # All sheets methods
    def remove_all_sheets(self):
        """Remove all worksheets from the workbook.

        This generally should only be run during creation of a SeedsWorkbook,
        or before loading a workbook into a SeedsWorkbook.
        """
        for sheet in list(self._wb.worksheets):
            self._wb.remove_sheet(sheet)

    def create_all_sheets(self):
        """Create all of the worksheets in the SeedsWorkbook."""
        self.remove_all_sheets()
        self.create_indexes_sheet()
        self.create_common_names_sheet()
        self.create_botanical_names_sheet()
        self.create_series_sheet()
        self.create_cultivars_sheet()
        self.create_packets_sheet()

    def load_all_sheets_from_workbook(self):
        """Set up all SeedsWorksheets with sheets from loaded workbook."""
        self.load_indexes_from_workbook()
        self.load_common_names_from_workbook()
        self.load_botanical_names_from_workbook()
        self.load_series_from_workbook()
        self.load_cultivars_from_workbook()
        self.load_packets_from_workbook()

    # Workbook methods
    @classmethod
    def load(cls, filename):
        self._wb = openpyxl.load_workbook(filename)
        self.load_all_sheets_from_workbook()

    def save(self, filename):
        self._wb.save(filename)
