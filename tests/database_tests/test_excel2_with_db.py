import pytest
from io import StringIO
from unittest import mock
from openpyxl import Workbook
from app.seeds.excel2 import (
    BotanicalNamesWorksheet,
    CommonNamesWorksheet,
    get_or_create_common_name,
    get_or_create_cultivar,
    get_or_create_index,
    IndexesWorksheet,
    SeedsWorksheet,
    SeriesWorksheet
)
from app.seeds.models import (
    BotanicalName,
    CommonName,
    Cultivar,
    Index,
    Series
)


class TestExcel2WithDB:
    """Test module-level functions of excel2 which utilize the database."""
    def test_get_or_create_index_create(self, db):
        """Create a new Index if no Index exists with given name."""
        messages = StringIO()
        idx = get_or_create_index(name='Perennial', file=messages)
        assert idx not in Index.query.all()
        assert idx.created
        messages.seek(0)
        assert 'does not yet exist' in messages.read()

    def test_get_or_create_index_get(self, db):
        """Return an Index loaded from db if it exists."""
        messages = StringIO()
        idx = Index(name='Perennial')
        db.session.add(idx)
        db.session.commit()
        idx2 = get_or_create_index(name='Perennial', file=messages)
        assert idx2 is idx
        assert not idx.created
        messages.seek(0)
        assert 'loaded from the database' in messages.read()

    def test_get_or_create_common_name_create_cn_and_index(self, db):
        """Create a new CommonName and Index if not in db."""
        messages = StringIO()
        cn = get_or_create_common_name(name='Foxglove',
                                       index='Perennial',
                                       file=messages)
        assert cn not in CommonName.query.all()
        assert cn.created
        assert cn.index not in Index.query.all()
        assert cn.index.created
        messages.seek(0)
        msgs = messages.read()
        assert 'The CommonName \'Foxglove\' does not yet exist' in msgs
        assert 'The Index \'Perennial\' does not yet exist' in msgs

    def test_get_or_create_common_name_create_cn(self, db):
        """Create new CommonName but use existing Index."""
        messages = StringIO()
        idx = Index(name='Perennial')
        db.session.add(idx)
        db.session.commit()
        cn = get_or_create_common_name(name='Foxglove',
                                       index='Perennial',
                                       file=messages)
        assert cn.created
        assert cn.index is idx
        assert not cn.index.created
        messages.seek(0)
        msgs = messages.read()
        assert 'The CommonName \'Foxglove\' does not yet exist' in msgs
        assert 'The Index \'Perennial\' has been loaded' in msgs

    @mock.patch('app.seeds.excel2.get_or_create_index')
    def test_get_or_create_common_name_get(self, m_goci, db):
        """Load CommonName from db if it exists with given Index."""
        messages = StringIO()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cng = get_or_create_common_name(name='Foxglove',
                                        index='Perennial',
                                        file=messages)
        assert cng is cn
        assert not cng.created
        messages.seek(0)
        msgs = messages.read()
        assert 'The CommonName \'Foxglove\' has been loaded' in msgs
        assert not m_goci.called

    @mock.patch('app.seeds.excel2.get_or_create_common_name')
    def test_get_or_create_cultivar_create_all_no_series(self, m_goccn, db):
        """Create all needed parts of Cultivar."""
        messages = StringIO()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        m_goccn.return_value = cn
        cv = get_or_create_cultivar(name='Foxy',
                                    common_name='Foxglove',
                                    index='Perennial',
                                    file=messages)
        m_goccn.assert_called_with(name='Foxglove',
                                   index='Perennial',
                                   file=messages)
        assert cv.created
        assert cv.name == 'Foxy'
        assert cv.common_name.name == 'Foxglove'
        assert cv.common_name.index.name == 'Perennial'
        messages.seek(0)
        msgs = messages.read()
        assert 'The Cultivar \'Foxy Foxglove\' does not yet exist' in msgs

    def test_get_or_create_cultivar_create_all(self, db):
        """Create all needed parts of Cultivar."""
        messages = StringIO()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cv = get_or_create_cultivar(name='Petra',
                                    common_name='Foxglove',
                                    index='Perennial',
                                    series='Polkadot',
                                    file=messages)
        assert cv.series.name == 'Polkadot'
        assert cv.series.common_name is cv.common_name
        messages.seek(0)
        msgs = messages.read()
        assert 'The Series \'Polkadot\' does not yet exist' in msgs

    def test_get_or_create_cultivar_exists_no_series(self, db):
        """Return existing Cultivar instead of new one."""
        messages = StringIO()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        db.session.add(cv)
        db.session.commit()
        cvq = get_or_create_cultivar(name='Foxy',
                                     common_name='Foxglove',
                                     index='Perennial',
                                     file=messages)
        assert cvq is cv
        messages.seek(0)
        msgs = messages.read()
        assert 'The Cultivar \'Foxy Foxglove\' has been loaded' in msgs

    def test_get_or_create_cultivar_exists_with_series(self, db):
        """Return existing Cultivar instead of new one."""
        messages = StringIO()
        cv = Cultivar(name='Petra')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.series = Series(name='Polkadot')
        cv.series.common_name = cv.common_name
        db.session.add(cv)
        db.session.commit()
        cvq = get_or_create_cultivar(name='Petra',
                                     common_name='Foxglove',
                                     index='Perennial',
                                     series='Polkadot',
                                     file=messages)
        assert cvq is cv
        messages.seek(0)
        msgs = messages.read()
        assert 'The Cultivar \'Polkadot Petra Foxglove\' has been load' in msgs


class TestSeedsWorksheetWithDB:
    """Test methods of SeedsWorksheet that (normally) need to use the db."""
    @mock.patch('app.seeds.excel2.db.session.rollback')
    @mock.patch('app.seeds.excel2.SeedsWorksheet.save_row_to_db')
    def test_save_to_db_with_exception(self, m_srtdb, m_r):
        """Handle exceptions raised during the execution of save_row_to_db.

        Rollback the database, then raise a RuntimeError if save_row_to_db
        causes an exception to be raised. The RuntimeError should include the
        original exception.
        """
        m_srtdb.side_effect = Exception('Oops.')
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.cell(2, 1).value = 'One'
        sws.cell(3, 1).value = 'Two'
        sws.cell(4, 1).value = 'Three'
        with pytest.raises(RuntimeError):
            sws.save_to_db()
        assert m_r.called
        try:
            sws.save_to_db()
        except Exception as e:
            assert 'Exception: Oops.' in e.args[0]

    @mock.patch('app.seeds.excel2.db.session.commit')
    @mock.patch('app.seeds.excel2.SeedsWorksheet.save_row_to_db')
    def test_save_to_db_with_changes(self, m_srtdb, m_c):
        """Call save_one_to_db for each row in sheet, and commit changes."""
        m_srtdb.return_value = True
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.cell(2, 1).value = 'One'
        sws.cell(3, 1).value = 'Two'
        sws.cell(4, 1).value = 'Three'
        sws.save_to_db(file=messages)
        assert m_srtdb.call_count == 3
        assert m_c.called
        messages.seek(0)
        msgs = messages.read()
        assert 'All changes have been committed' in msgs


class TestIndexesWorksheetWithDB:
    """Test methods of the IndexesWorksheet that need to use the database."""
    @mock.patch('app.seeds.excel2.db.session.flush')
    def test_save_row_to_db_no_changes(self, m_f, db):
        """Make no changes if Index in row identical to one in db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        idx = Index(name='Perennial', description='Built to last.')
        db.session.add(idx)
        db.session.commit()
        iws.add_one(Index(name='Perennial', description='Built to last.'))
        assert not iws.save_row_to_db(row=2, file=messages)
        assert not m_f.called
        messages.seek(0)
        msgs = messages.read()
        assert 'No changes were made' in msgs

    @mock.patch('app.seeds.excel2.get_or_create_index')
    def test_save_row_to_db_new_no_opts(self, m_goci, db):
        """Create an Index with no optional data and flush it to the db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        idx = Index(name='Perennial')
        idx.created = True
        m_goci.return_value = idx
        iws.add_one(idx)
        assert iws.save_row_to_db(row=2, file=messages)
        m_goci.assert_called_with(name='Perennial', file=messages)
        assert Index.query.filter(Index._name == 'Perennial').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert 'Changes to Index \'Perennial\' have been flushed' in msgs

    def test_save_row_to_db_new_with_desc(self, db):
        """Create an Index with a description and flush it to db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        idx = Index(name='Perennial', description='Built to last.')
        iws.add_one(idx)
        assert iws.save_row_to_db(row=2, file=messages)
        idxq = Index.query.filter(Index._name == 'Perennial').one_or_none()
        assert idxq.description == 'Built to last.'
        messages.seek(0)
        msgs = messages.read()
        assert 'Description for the Index \'Perennial\' set to: Built' in msgs

    def test_save_row_to_db_existing_new_desc(self, db):
        """Load an Index from db and change description."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        idx = Index(name='Perennial', description='Built to last.')
        db.session.add(idx)
        db.session.commit()
        iws.add_one(Index(name='Perennial', description='Live long time.'))
        assert iws.save_row_to_db(row=2, file=messages)
        idxq = Index.query.filter(Index._name == 'Perennial').one_or_none()
        assert idxq is idx
        assert idx.description == 'Live long time.'
        messages.seek(0)
        msgs = messages.read()
        assert 'Description for the Index \'Perennial\' set to: Live' in msgs

    def test_save_row_to_db_existing_clears_desc(self, db):
        """Clear description for loaded Index if desc empty in row."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        idx = Index(name='Perennial', description='Built to last.')
        db.session.add(idx)
        db.session.commit()
        iws.add_one(Index(name='Perennial'))
        assert iws.save_row_to_db(row=2, file=messages)
        idxq = Index.query.filter(Index._name == 'Perennial').one_or_none()
        assert idxq is idx
        assert idx.description is None
        messages.seek(0)
        msgs = messages.read()
        assert 'Description for the Index \'Perennial\' has been clear' in msgs


class TestCommonNamesWorksheet:
    """Test methods of the CommonNamesWorksheet which use the db."""
    @mock.patch('app.seeds.excel2.db.session.flush')
    def test_save_row_to_db_no_changes(self, m_f, db):
        """Don't change anything if row data is identical to CN in db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cnws.add_one(cn)
        assert not cnws.save_row_to_db(row=2, file=messages)
        messages.seek(0)
        msgs = messages.read()
        assert 'No changes were made to the CommonName \'Foxglove\'' in msgs

    @mock.patch('app.seeds.excel2.get_or_create_common_name')
    def test_save_row_to_db_new_no_optionals(self, m_goccn, db):
        """Create a new CommonName and flush to db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cn.created = True
        m_goccn.return_value = cn
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, file=messages)
        assert CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        m_goccn.assert_called_with(name='Foxglove',
                                   index='Perennial',
                                   file=messages)
        messages.seek(0)
        msgs = messages.read()
        assert 'Changes to the CommonName \'Foxglove\' have been flush' in msgs

    def test_save_row_to_db_new_with_parent_new(self, db):
        """Create a new CommonName with a new parent CN."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Sauce Tomato')
        cn.index = Index(name='Vegetable')
        cn.parent = CommonName(name='Tomato')
        cn.parent.index = Index(name='Vegetable')
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Sauce Tomato')\
            .one_or_none()
        assert cnq.parent.name == 'Tomato'
        assert cnq.parent.index.name == 'Vegetable'
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Sauce Tomato\' has been set as a '
                'subcategory of \'Tomato\'.') in msgs

    def test_save_row_to_db_new_with_existing_parent(self, db):
        """Create a new CommonName with an existing parent from the db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cnp = CommonName(name='Tomato')
        cnp.index = Index(name='Vegetable')
        db.session.add(cnp)
        db.session.commit()
        cn = CommonName(name='Sauce Tomato')
        cn.index = Index(name='Vegetable')
        cn.parent = CommonName(name='Tomato')
        cn.parent.index = Index(name='Vegetable')
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Sauce Tomato')\
            .one_or_none()
        assert cnq.parent is cnp

    def test_save_row_to_db_new_with_desc(self, db):
        """Create a new CommonName with a description."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove', description='A bit spotty.')
        cn.index = Index(name='Perennial')
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq.description == 'A bit spotty.'
        messages.seek(0)
        msgs = messages.read()
        assert 'Description for the CommonName \'Foxglove\' set to' in msgs

    def test_save_row_to_db_existing_with_new_desc(self, db):
        """Edit desc of existing CommonName."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove', description='A bit spotty.')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove', description='More dots!')
        cn2.index = Index(name='Perennial')
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert cn.description == 'More dots!'
        messages.seek(0)
        msgs = messages.read()
        assert ('Description for the CommonName \'Foxglove\' set to: '
                'More dots!') in msgs

    def test_save_row_to_db_existing_clears_desc(self, db):
        """Clear description of existing CommonName if row has none."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove', description='A bit spotty.')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert cn.description is None
        messages.seek(0)
        msgs = messages.read()
        assert ('Description for the CommonName \'Foxglove\' has been '
                'cleared.') in msgs

    def test_save_row_to_db_new_with_instructions(self, db):
        """Create a new CommonName with planting instructions."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove', instructions='Just add water!')
        cn.index = Index(name='Perennial')
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq.instructions == 'Just add water!'
        messages.seek(0)
        msgs = messages.read()
        assert ('Planting instructions for the CommonName \'Foxglove\' set '
                'to: Just add water!') in msgs

    def test_save_row_to_db_existing_with_new_instructions(self, db):
        """Change instructions of existing CommonName in db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove', instructions='Just add water!')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove', instructions='Put them in soil.')
        cn2.index = Index(name='Perennial')
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert cn.instructions == 'Put them in soil.'
        messages.seek(0)
        msgs = messages.read()
        assert ('Planting instructions for the CommonName \'Foxglove\' set '
                'to: Put them in soil.') in msgs

    def test_save_row_to_db_existing_clears_instructions(self, db):
        """Clear instructions from CommonName with blank inst. in row."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove', instructions='Just add water!')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert cn.instructions is None
        messages.seek(0)
        msgs = messages.read()
        assert ('Planting instructions for the CommonName \'Foxglove\' have '
                'been cleared.') in msgs

    def test_save_row_to_db_new_with_synonyms(self, db):
        """Create a CommonName with synonyms."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cn.set_synonyms_string('Digitalis')
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq.get_synonyms_string() == 'Digitalis'
        messages.seek(0)
        msgs = messages.read()
        assert ('Synonyms for the CommonName \'Foxglove\' set to: '
                'Digitalis') in msgs

    def test_save_row_to_db_existing_with_new_synonyms(self, db):
        """Change synonyms of existing CommonName."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cn.set_synonyms_string('Digitalis')
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        cn2.set_synonyms_string('Vulpine Handwarmer')
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert cn.get_synonyms_string() == 'Vulpine Handwarmer'
        messages.seek(0)
        msgs = messages.read()
        assert ('Synonyms for the CommonName \'Foxglove\' set to: Vulpine '
                'Handwarmer') in msgs

    def test_save_row_to_db_existing_clears_synonyms(self, db):
        """Clear synonyms from existing CommonName."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cn.set_synonyms_string('Digitalis')
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert not cn.get_synonyms_string()
        messages.seek(0)
        msgs = messages.read()
        assert ('Synonyms for the CommonName \'Foxglove\' have been '
                'cleared.') in msgs

    def test_save_row_to_db_new_visible(self, db):
        """Set invisible to False if false in row."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cn.invisible = False
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert not cnq.invisible
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Foxglove\' is visible on generated '
                'pages.') in msgs
        
    def test_save_row_to_db_new_invisible(self, db):
        """Set invisible to True if true in row."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cn.invisible = True
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq.invisible
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Foxglove\' is not visible on generated '
                'pages.') in msgs

    def test_save_row_to_db_new_with_new_gwcn(self, db):
        """Create a new CommonName with Grows With Common Names."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        gwcn = CommonName(name='Butterfly Weed')
        gwcn.index = Index(name='Perennial')
        cn.gw_common_names.append(gwcn)
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq.gw_common_names[0].name == 'Butterfly Weed'
        assert cnq.gw_common_names[0].index.name == 'Perennial'
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Butterfly Weed\' has been added to Grows '
                'With Common Names for the CommonName \'Foxglove\'.') in msgs

    def test_save_row_to_db_new_with_existing_gwcn(self, db):
        """Load GWCN from db if it exists."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        gwcn = CommonName(name='Butterfly Weed')
        gwcn.index = Index(name='Perennial')
        db.session.add(gwcn)
        db.session.commit()
        gwcn2 = CommonName(name='Butterfly Weed')
        gwcn2.index = Index(name='Perennial')
        cn.gw_common_names.append(gwcn2)
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq.gw_common_names[0] is gwcn
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Butterfly Weed\' has been added to Grows '
                'With Common Names for the CommonName \'Foxglove\'.') in msgs

    def test_save_row_to_db_existing_with_new_gwcn(self, db):
        """Add gwcn to existing CommonName."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        gwcn = CommonName(name='Butterfly Weed')
        gwcn.index = Index(name='Perennial')
        cn2.gw_common_names.append(gwcn)
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert cn.gw_common_names[0].name == 'Butterfly Weed'
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Butterfly Weed\' has been added to Grows '
                'With Common Names for the CommonName \'Foxglove\'.') in msgs

    def test_save_row_to_db_existing_removes_gwcn(self, db):
        """Remove gwcn not present in row."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        gwcn = CommonName(name='Butterfly Weed')
        gwcn.index = cn.index
        cn.gw_common_names.append(gwcn)
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert not cn.gw_common_names
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Butterfly Weed\' has been removed from '
                'Grows With Common Names for the CommonName '
                '\'Foxglove\'.') in msgs

    def test_save_row_to_db_existing_gwcn_switcheroo(self, db):
        """Remove old gwcn, add new one."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        gwcn = CommonName(name='Butterfly Weed')
        gwcn.index = cn.index
        cn.gw_common_names.append(gwcn)
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        gwcn2 = CommonName(name='Coleus')
        gwcn2.index = Index(name='Annual')
        cn2.gw_common_names.append(gwcn2)
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert gwcn not in cn.gw_common_names
        assert cn.gw_common_names[0].name == 'Coleus'
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Coleus\' has been added to Grows With '
                'Common Names for the CommonName \'Foxglove\'.') in msgs
        assert ('The CommonName \'Butterfly Weed\' has been removed from '
                'Grows With Common Names for the CommonName '
                '\'Foxglove\'.') in msgs

    def test_save_row_to_db_new_with_new_gwcv(self, db):
        """Add a new CommonName with a new Grows With Cultivar."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        gwcv = Cultivar(name='King')
        gwcv.common_name = CommonName(name='Coleus')
        gwcv.common_name.index = Index(name='Annual')
        cn.gw_cultivars.append(gwcv)
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq.gw_cultivars[0].fullname == 'King Coleus'
        messages.seek(0)
        msgs = messages.read()
        assert ('The Cultivar \'King Coleus\' has been added to Grows With '
                'Cultivars for the CommonName \'Foxglove\'.') in msgs

    def test_save_row_to_db_new_with_existing_gwcv(self, db):
        """Add a new CommonName with an existing Cultivar as GWCV."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        gwcv = Cultivar(name='King')
        gwcv.common_name = CommonName(name='Coleus')
        gwcv.common_name.index = Index(name='Annual')
        db.session.add(gwcv)
        db.session.commit()
        gwcv2 = Cultivar(name='King')
        gwcv2.common_name = CommonName(name='Coleus')
        gwcv2.common_name.index = Index(name='Annual')
        cn.gw_cultivars.append(gwcv2)
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq.gw_cultivars[0] is gwcv
        messages.seek(0)
        msgs = messages.read()
        assert ('The Cultivar \'King Coleus\' has been added to Grows With '
                'Cultivars for the CommonName \'Foxglove\'.') in msgs

    def test_save_row_to_db_existing_with_new_gwcv(self, db):
        """Add a new GWCV to an existing common name."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        gwcv = Cultivar(name='King')
        gwcv.common_name = CommonName(name='Coleus')
        gwcv.common_name.index = Index(name='Annual')
        cn2.gw_cultivars.append(gwcv)
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert cn.gw_cultivars[0].fullname == 'King Coleus'
        messages.seek(0)
        msgs = messages.read()
        assert ('The Cultivar \'King Coleus\' has been added to Grows With '
                'Cultivars for the CommonName \'Foxglove\'.') in msgs

    def test_save_row_to_db_existing_removes_gwcv(self, db):
        """Remove GWCVs not present in row."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        gwcv = Cultivar(name='King')
        gwcv.common_name = CommonName(name='Coleus')
        gwcv.common_name.index = Index(name='Annual')
        cn.gw_cultivars.append(gwcv)
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert not cn.gw_cultivars
        messages.seek(0)
        msgs = messages.read()
        assert ('The Cultivar \'King Coleus\' has been removed from Grows '
                'With Cultivars for the CommonName \'Foxglove\'.') in msgs

    def test_save_row_to_db_existing_gwcv_switcheroo(self, db):
        """Remove old GWCVs and add new ones."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        gwcv = Cultivar(name='King')
        gwcv.common_name = CommonName(name='Coleus')
        gwcv.common_name.index = Index(name='Annual')
        cn.gw_cultivars.append(gwcv)
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        gwcv2 = Cultivar(name='Soulmate')
        gwcv2.common_name = CommonName(name='Butterfly Weed')
        gwcv2.common_name.index = Index(name='Perennial')
        cn2.gw_cultivars.append(gwcv2)
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, file=messages)
        cnq = CommonName.query\
            .filter(CommonName._name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert gwcv not in cn.gw_cultivars
        assert cn.gw_cultivars[0].fullname == 'Soulmate Butterfly Weed'
        messages.seek(0)
        msgs = messages.read()
        assert ('The Cultivar \'Soulmate Butterfly Weed\' has been added to '
                'Grows With Cultivars for the CommonName '
                '\'Foxglove\'.') in msgs
        assert ('The Cultivar \'King Coleus\' has been removed from Grows '
                'With Cultivars for the CommonName \'Foxglove\'.') in msgs


class TestBotanicalNamesWorksheetWithDB:
    """Test methods of BotanicalNamesWorksheet that require db use."""
    def test_save_row_to_db_bad_botanical_name(self, db):
        """Print a message and return False if bn is not valid.

        This check should happen before entering the main body of logic in
        save_row_to_db, so it shouldn't execute any code past the error check,
        but it should not raise an exception, as we just want to skip adding
        bad botanical names and carry on with program execution.
        """
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName()
        bn._name = 'Invalid Botanical Name'
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        bnws.add_one(bn)
        assert not bnws.save_row_to_db(2, file=messages)
        messages.seek(0)
        msgs = messages.read()
        assert msgs == ('Could not add the BotanicalName \'Invalid Botanical '
                        'Name\' because it does not appear to be a validly '
                        'formatted botanical name.\n')

    @mock.patch('app.seeds.excel2.db.session.flush')
    def test_save_row_to_db_no_changes(self, m_f, db):
        """Don't change existing BotanicalName if row contains same data."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        db.session.add(bn)
        db.session.commit()
        bn2 = BotanicalName(name='Digitalis purpurea')
        bn2.common_names = [CommonName(name='Foxglove')]
        bn2.common_names[0].index = Index(name='Perennial')
        bnws.add_one(bn2)
        assert not bnws.save_row_to_db(2, file=messages)
        assert not m_f.called
        messages.seek(0)
        msgs = messages.read()
        assert ('The BotanicalName \'Digitalis purpurea\' has been loaded '
                'from the database.') in msgs
        assert ('No changes were made to the BotanicalName \'Digitalis '
                'purpurea\'.') in msgs

    def test_save_row_to_db_new_no_opts(self, db):
        """Add a new BotanicalName to the database."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        bnws.add_one(bn)
        assert bnws.save_row_to_db(2, file=messages)
        bnq = BotanicalName.query\
            .filter(BotanicalName._name == 'Digitalis purpurea')\
            .one_or_none()
        assert bnq.name == 'Digitalis purpurea'
        assert bnq.common_names[0].name == 'Foxglove'
        assert bnq.common_names[0].index.name == 'Perennial'
        messages.seek(0)
        msgs = messages.read()
        assert ('The BotanicalName \'Digitalis purpurea\' does not yet exist '
                'in the database') in msgs
        assert ('The CommonName \'Foxglove\' has been added to CommonNames '
                'for the BotanicalName \'Digitalis purpurea\'.') in msgs
        assert ('Changes to the BotanicalName \'Digitalis purpurea\' have '
                'been flushed to the database.') in msgs

    def test_save_row_to_db_new_with_synonyms(self, db):
        """Add a new BotanicalName with synonyms."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        bn.set_synonyms_string('Digitalis über alles')
        bnws.add_one(bn)
        assert bnws.save_row_to_db(2, file=messages)
        bnq = BotanicalName.query\
            .filter(BotanicalName._name == 'Digitalis purpurea')\
            .one_or_none()
        assert bnq.get_synonyms_string() == 'Digitalis über alles'
        messages.seek(0)
        msgs = messages.read()
        assert ('Synonyms for the BotanicalName \'Digitalis purpurea\' set '
                'to: Digitalis über alles') in msgs

    def test_save_row_to_db_existing_with_synonyms(self, db):
        """Change synonyms for an existing BotanicalName."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        bn.set_synonyms_string('Digitalis über alles')
        db.session.add(bn)
        db.session.commit()
        bn2 = BotanicalName(name='Digitalis purpurea')
        bn2.common_names = [CommonName(name='Foxglove')]
        bn2.common_names[0].index = Index(name='Perennial')
        bn2.set_synonyms_string('Innagada davida')
        bnws.add_one(bn2)
        assert bnws.save_row_to_db(2, file=messages)
        bnq = BotanicalName.query\
            .filter(BotanicalName._name == 'Digitalis purpurea')\
            .one_or_none()
        assert bnq is bn
        assert bn.get_synonyms_string() == 'Innagada davida'
        messages.seek(0)
        msgs = messages.read()
        assert ('Synonyms for the BotanicalName \'Digitalis purpurea\' set '
                'to: Innagada davida') in msgs

    def test_save_row_to_db_existing_clears_synonyms(self, db):
        """Clear synonyms if row does not have them."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        bn.set_synonyms_string('Digitalis über alles')
        db.session.add(bn)
        db.session.commit()
        bn2 = BotanicalName(name='Digitalis purpurea')
        bn2.common_names = [CommonName(name='Foxglove')]
        bn2.common_names[0].index = Index(name='Perennial')
        bnws.add_one(bn2)
        assert bnws.save_row_to_db(2, file=messages)
        bnq = BotanicalName.query\
            .filter(BotanicalName._name == 'Digitalis purpurea')\
            .one_or_none()
        assert bnq is bn
        assert not bn.get_synonyms_string()
        messages.seek(0)
        msgs = messages.read()
        assert ('Synonyms for the BotanicalName \'Digitalis purpurea\' have '
                'been cleared.') in msgs

    def test_save_row_to_db_existing_add_common_name(self, db):
        """Add a CommonName to existing BotanicalName."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        db.session.add(bn)
        db.session.commit()
        bn2 = BotanicalName(name='Digitalis purpurea')
        bn2.common_names = [CommonName(name='Foxglove'),
                            CommonName(name='Digitalis')]
        bn2.common_names[0].index = Index(name='Perennial')
        bn2.common_names[1].index = Index(name='Perennial')
        bnws.add_one(bn2)
        assert bnws.save_row_to_db(2, file=messages)
        bnq = BotanicalName.query\
            .filter(BotanicalName._name == 'Digitalis purpurea')\
            .one_or_none()
        assert bnq is bn
        assert bn.common_names[0].name == 'Foxglove'
        assert bn.common_names[1].name == 'Digitalis'
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Digitalis\' has been added to CommonNames '
                'for the BotanicalName \'Digitalis purpurea\'.') in msgs

    def test_save_row_to_db_existing_common_name_switcheroo(self, db):
        """Replace a CommonName in an existing BotanicalName."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        db.session.add(bn)
        db.session.commit()
        bn2 = BotanicalName(name='Digitalis purpurea')
        bn2.common_names = [CommonName(name='Digitalis')]
        bn2.common_names[0].index = Index(name='Perennial')
        bnws.add_one(bn2)
        assert bnws.save_row_to_db(2, file=messages)
        bnq = BotanicalName.query\
            .filter(BotanicalName._name == 'Digitalis purpurea')\
            .one_or_none()
        assert bnq is bn
        assert len(bn.common_names) == 1
        assert bn.common_names[0].name == 'Digitalis'
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Digitalis\' has been added to CommonNames '
                'for the BotanicalName \'Digitalis purpurea\'.') in msgs
        assert ('The CommonName \'Foxglove\' has been removed from '
                'CommonNames for the BotanicalName \'Digitalis '
                'purpurea\'.') in msgs


class TestSeriesWorksheetWithDB:
    """Test methods of the SeriesWorksheet that use the database."""
    @mock.patch('app.seeds.excel2.db.session.flush')
    def test_save_row_to_db_no_changes(self, m_f, db):
        """If it ain't broke don't fix it."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        sr.position = Series.BEFORE_CULTIVAR
        db.session.add(sr)
        db.session.commit()
        sr2 = Series(name='Polkadot')
        sr2.common_name = CommonName(name='Foxglove')
        sr2.common_name.index = Index(name='Perennial')
        sr2.position = Series.BEFORE_CULTIVAR
        srws.add_one(sr2)
        assert not srws.save_row_to_db(2, file=messages)
        assert not m_f.called
        messages.seek(0)
        msgs = messages.read()
        assert 'No changes were made to the Series \'Polkadot\'.' in msgs

    def test_save_row_to_db_new_no_desc(self, db):
        """Save new Series to database."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        sr.position = Series.BEFORE_CULTIVAR
        srws.add_one(sr)
        assert srws.save_row_to_db(2, file=messages)
        srq = Series.query.filter(Series.name == 'Polkadot').one_or_none()
        assert srq.name == 'Polkadot'
        assert srq.common_name.name == 'Foxglove'
        assert srq.common_name.index.name == 'Perennial'
        messages.seek(0)
        msgs = messages.read()
        assert ('The Series \'Polkadot\' does not yet exist in the database, '
                'so it has been created.') in msgs
        assert ('The Series name \'Polkadot\' will be placed before the '
                'Cultivar name for each Cultivar in the Series.') in msgs

    def test_save_row_to_db_new_with_desc(self, db):
        """Save a new Series with a description to database."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot', description='A bit spotty.')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        sr.position = Series.BEFORE_CULTIVAR
        srws.add_one(sr)
        assert srws.save_row_to_db(2, file=messages)
        srq = Series.query.filter(Series.name == 'Polkadot').one_or_none()
        assert srq.description == 'A bit spotty.'
        messages.seek(0)
        msgs = messages.read()
        assert ('Description for the Series \'Polkadot\' set to: A bit '
                'spotty.') in msgs

    def test_save_row_to_db_existing_change_desc(self, db):
        """Change the description of an existing Series."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot', description='A bit spotty.')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        sr.position = Series.BEFORE_CULTIVAR
        db.session.add(sr)
        db.session.commit()
        sr2 = Series(name='Polkadot', description='More dots!')
        sr2.common_name = CommonName(name='Foxglove')
        sr2.common_name.index = Index(name='Perennial')
        sr2.position = Series.BEFORE_CULTIVAR
        srws.add_one(sr2)
        assert srws.save_row_to_db(2, file=messages)
        srq = Series.query.filter(Series.name == 'Polkadot').one_or_none()
        assert srq is sr
        assert sr.description == 'More dots!'
        messages.seek(0)
        msgs = messages.read()
        assert ('Description for the Series \'Polkadot\' set to: More '
                'dots!') in msgs

    def test_save_row_to_db_existing_clears_desc(self, db):
        """Clear description of existing Series if row has no desc."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot', description='A bit spotty.')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        sr.position = Series.BEFORE_CULTIVAR
        db.session.add(sr)
        db.session.commit()
        sr2 = Series(name='Polkadot')
        sr2.common_name = CommonName(name='Foxglove')
        sr2.common_name.index = Index(name='Perennial')
        sr2.position = Series.BEFORE_CULTIVAR
        srws.add_one(sr2)
        assert srws.save_row_to_db(2, file=messages)
        srq = Series.query.filter(Series.name == 'Polkadot').one_or_none()
        assert srq is sr
        assert not sr.description
        messages.seek(0)
        msgs = messages.read()
        assert ('Description for the Series \'Polkadot\' has been '
                'cleared') in msgs

    def test_save_row_to_db_existing_change_position(self, db):
        """Change the position of an existing Series."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        sr.position = Series.BEFORE_CULTIVAR
        db.session.add(sr)
        db.session.commit()
        sr2 = Series(name='Polkadot')
        sr2.common_name = CommonName(name='Foxglove')
        sr2.common_name.index = Index(name='Perennial')
        sr2.position = Series.AFTER_CULTIVAR
        srws.add_one(sr2)
        assert srws.save_row_to_db(2, file=messages)
        srq = Series.query.filter(Series.name == 'Polkadot').one_or_none()
        assert srq is sr
        assert sr.position == Series.AFTER_CULTIVAR
        messages.seek(0)
        msgs = messages.read()
        assert ('The Series name \'Polkadot\' will be placed after the '
                'Cultivar name for each Cultivar in the Series.') in msgs
