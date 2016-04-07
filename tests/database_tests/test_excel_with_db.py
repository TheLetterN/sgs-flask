import datetime
import pytest
from io import StringIO
from unittest import mock
from openpyxl import Workbook
from app.seeds.excel import (
    BotanicalNamesWorksheet,
    CommonNamesWorksheet,
    CultivarsWorksheet,
    get_or_create_common_name,
    get_or_create_cultivar,
    get_or_create_index,
    IndexesWorksheet,
    PacketsWorksheet,
    SeedsWorksheet,
    SeriesWorksheet
)
from app.seeds.models import (
    BotanicalName,
    CommonName,
    Cultivar,
    Image,
    Index,
    Packet,
    Quantity,
    Series
)


class TestExcel2WithDB:
    """Test module-level functions of excel which utilize the database."""
    def test_get_or_create_index_create(self, db):
        """Create a new Index if no Index exists with given name."""
        messages = StringIO()
        idx = get_or_create_index(name='Perennial', stream=messages)
        assert idx not in Index.query.all()
        assert idx.created
        messages.seek(0)
        assert 'does not yet exist' in messages.read()

    def test_get_or_create_index_get(self, db):
        """Return an Index loaded from db if it exists."""
        messages = StringIO()
        idx = Index(name='Perennial')
        db.session.add(idx)
        db.session.commit()
        idx2 = get_or_create_index(name='Perennial', stream=messages)
        assert idx2 is idx
        assert not idx.created
        messages.seek(0)
        assert 'loaded from the database' in messages.read()

    def test_get_or_create_common_name_create_cn_and_index(self, db):
        """Create a new CommonName and Index if not in db."""
        messages = StringIO()
        cn = get_or_create_common_name(name='Foxglove',
                                       index='Perennial',
                                       stream=messages)
        assert cn not in CommonName.query.all()
        assert cn.created
        assert cn.index not in Index.query.all()
        assert cn.index.created
        messages.seek(0)
        msgs = messages.read()
        assert 'The CommonName \'Foxglove\' does not yet exist' in msgs
        assert 'The Index \'Perennial\' does not yet exist' in msgs

    def test_get_or_create_common_name_create_cn(self, db):
        """Create new CommonName but use existing Index."""
        messages = StringIO()
        idx = Index(name='Perennial')
        db.session.add(idx)
        db.session.commit()
        cn = get_or_create_common_name(name='Foxglove',
                                       index='Perennial',
                                       stream=messages)
        assert cn.created
        assert cn.index is idx
        assert not cn.index.created
        messages.seek(0)
        msgs = messages.read()
        assert 'The CommonName \'Foxglove\' does not yet exist' in msgs
        assert 'The Index \'Perennial\' has been loaded' in msgs

    @mock.patch('app.seeds.excel.get_or_create_index')
    def test_get_or_create_common_name_get(self, m_goci, db):
        """Load CommonName from db if it exists with given Index."""
        messages = StringIO()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cng = get_or_create_common_name(name='Foxglove',
                                        index='Perennial',
                                        stream=messages)
        assert cng is cn
        assert not cng.created
        messages.seek(0)
        msgs = messages.read()
        assert 'The CommonName \'Foxglove\' has been loaded' in msgs
        assert not m_goci.called

    @mock.patch('app.seeds.excel.get_or_create_common_name')
    def test_get_or_create_cultivar_create_all_no_series(self, m_goccn, db):
        """Create all needed parts of Cultivar."""
        messages = StringIO()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        m_goccn.return_value = cn
        cv = get_or_create_cultivar(name='Foxy',
                                    common_name='Foxglove',
                                    index='Perennial',
                                    stream=messages)
        m_goccn.assert_called_with(name='Foxglove',
                                   index='Perennial',
                                   stream=messages)
        assert cv.created
        assert cv.name == 'Foxy'
        assert cv.common_name.name == 'Foxglove'
        assert cv.common_name.index.name == 'Perennial'
        messages.seek(0)
        msgs = messages.read()
        assert 'The Cultivar \'Foxy Foxglove\' does not yet exist' in msgs

    def test_get_or_create_cultivar_create_all(self, db):
        """Create all needed parts of Cultivar."""
        messages = StringIO()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cv = get_or_create_cultivar(name='Petra',
                                    common_name='Foxglove',
                                    index='Perennial',
                                    series='Polkadot',
                                    stream=messages)
        assert cv.series.name == 'Polkadot'
        assert cv.series.common_name is cv.common_name
        messages.seek(0)
        msgs = messages.read()
        assert 'The Series \'Polkadot\' does not yet exist' in msgs

    def test_get_or_create_cultivar_exists_no_series(self, db):
        """Return existing Cultivar instead of new one."""
        messages = StringIO()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        db.session.add(cv)
        db.session.commit()
        cvq = get_or_create_cultivar(name='Foxy',
                                     common_name='Foxglove',
                                     index='Perennial',
                                     stream=messages)
        assert cvq is cv
        messages.seek(0)
        msgs = messages.read()
        assert 'The Cultivar \'Foxy Foxglove\' has been loaded' in msgs

    def test_get_or_create_cultivar_exists_with_series(self, db):
        """Return existing Cultivar instead of new one."""
        messages = StringIO()
        cv = Cultivar(name='Petra')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.series = Series(name='Polkadot')
        cv.series.common_name = cv.common_name
        db.session.add(cv)
        db.session.commit()
        cvq = get_or_create_cultivar(name='Petra',
                                     common_name='Foxglove',
                                     index='Perennial',
                                     series='Polkadot',
                                     stream=messages)
        assert cvq is cv
        messages.seek(0)
        msgs = messages.read()
        assert 'The Cultivar \'Polkadot Petra Foxglove\' has been load' in msgs


class TestSeedsWorksheetWithDB:
    """Test methods of SeedsWorksheet that (normally) need to use the db."""
    @mock.patch('app.seeds.excel.db.session.rollback')
    @mock.patch('app.seeds.excel.SeedsWorksheet.save_row_to_db')
    def test_save_to_db_with_exception(self, m_srtdb, m_r):
        """Handle exceptions raised during the execution of save_row_to_db.

        Rollback the database, then raise a RuntimeError if save_row_to_db
        causes an exception to be raised. The RuntimeError should include the
        original exception.
        """
        m_srtdb.side_effect = Exception('Oops.')
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.cell(2, 1).value = 'One'
        sws.cell(3, 1).value = 'Two'
        sws.cell(4, 1).value = 'Three'
        with pytest.raises(RuntimeError):
            sws.save_to_db()
        assert m_r.called
        try:
            sws.save_to_db()
        except Exception as e:
            assert 'Exception: Oops.' in e.args[0]

    @mock.patch('app.seeds.excel.db.session.commit')
    @mock.patch('app.seeds.excel.SeedsWorksheet.save_row_to_db')
    def test_save_to_db_with_changes(self, m_srtdb, m_c):
        """Call save_one_to_db for each row in sheet, and commit changes."""
        m_srtdb.return_value = True
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.cell(2, 1).value = 'One'
        sws.cell(3, 1).value = 'Two'
        sws.cell(4, 1).value = 'Three'
        sws.save_to_db(stream=messages)
        assert m_srtdb.call_count == 3
        assert m_c.called
        messages.seek(0)
        msgs = messages.read()
        assert 'All changes have been committed' in msgs


class TestIndexesWorksheetWithDB:
    """Test methods of the IndexesWorksheet that need to use the database."""
    @mock.patch('app.seeds.excel.db.session.flush')
    def test_save_row_to_db_no_changes(self, m_f, db):
        """Make no changes if Index in row identical to one in db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        idx = Index(name='Perennial', description='Built to last.')
        db.session.add(idx)
        db.session.commit()
        iws.add_one(Index(name='Perennial', description='Built to last.'))
        assert not iws.save_row_to_db(row=2, stream=messages)
        assert not m_f.called
        messages.seek(0)
        msgs = messages.read()
        assert 'No changes were made' in msgs

    @mock.patch('app.seeds.excel.get_or_create_index')
    def test_save_row_to_db_new_no_opts(self, m_goci, db):
        """Create an Index with no optional data and flush it to the db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        idx = Index(name='Perennial')
        idx.created = True
        m_goci.return_value = idx
        iws.add_one(idx)
        assert iws.save_row_to_db(row=2, stream=messages)
        m_goci.assert_called_with(name='Perennial', stream=messages)
        assert Index.query.filter(Index.name == 'Perennial').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert 'Changes to Index \'Perennial\' have been flushed' in msgs

    def test_save_row_to_db_new_with_desc(self, db):
        """Create an Index with a description and flush it to db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        idx = Index(name='Perennial', description='Built to last.')
        iws.add_one(idx)
        assert iws.save_row_to_db(row=2, stream=messages)
        idxq = Index.query.filter(Index.name == 'Perennial').one_or_none()
        assert idxq.description == '<p>Built to last.</p>'
        messages.seek(0)
        msgs = messages.read()
        assert 'Description for the Index \'Perennial\' set to:' in msgs

    def test_save_row_to_db_existing_new_desc(self, db):
        """Load an Index from db and change description."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        idx = Index(name='Perennial', description='Built to last.')
        db.session.add(idx)
        db.session.commit()
        iws.add_one(Index(name='Perennial', description='Live long time.'))
        assert iws.save_row_to_db(row=2, stream=messages)
        idxq = Index.query.filter(Index.name == 'Perennial').one_or_none()
        assert idxq is idx
        assert idx.description == '<p>Live long time.</p>'
        messages.seek(0)
        msgs = messages.read()
        assert 'Description for the Index \'Perennial\' set to:' in msgs

    def test_save_row_to_db_existing_clears_desc(self, db):
        """Clear description for loaded Index if desc empty in row."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        idx = Index(name='Perennial', description='Built to last.')
        db.session.add(idx)
        db.session.commit()
        iws.add_one(Index(name='Perennial'))
        assert iws.save_row_to_db(row=2, stream=messages)
        idxq = Index.query.filter(Index.name == 'Perennial').one_or_none()
        assert idxq is idx
        assert idx.description is None
        messages.seek(0)
        msgs = messages.read()
        assert 'Description for the Index \'Perennial\' has been clear' in msgs


class TestCommonNamesWorksheet:
    """Test methods of the CommonNamesWorksheet which use the db."""
    @mock.patch('app.seeds.excel.db.session.flush')
    def test_save_row_to_db_no_changes(self, m_f, db):
        """Don't change anything if row data is identical to CN in db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cnws.add_one(cn)
        assert not cnws.save_row_to_db(row=2, stream=messages)
        messages.seek(0)
        msgs = messages.read()
        assert 'No changes were made to the CommonName \'Foxglove\'' in msgs

    @mock.patch('app.seeds.excel.get_or_create_common_name')
    def test_save_row_to_db_new_no_optionals(self, m_goccn, db):
        """Create a new CommonName and flush to db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cn.created = True
        m_goccn.return_value = cn
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, stream=messages)
        assert CommonName.query\
            .filter(CommonName.name == 'Foxglove')\
            .one_or_none()
        m_goccn.assert_called_with(name='Foxglove',
                                   index='Perennial',
                                   stream=messages)
        messages.seek(0)
        msgs = messages.read()
        assert 'Changes to the CommonName \'Foxglove\' have been flush' in msgs

    def test_save_row_to_db_new_with_parent_new(self, db):
        """Create a new CommonName with a new parent CN."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Sauce Tomato')
        cn.index = Index(name='Vegetable')
        cn.parent = CommonName(name='Tomato')
        cn.parent.index = Index(name='Vegetable')
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Sauce Tomato')\
            .one_or_none()
        assert cnq.parent.name == 'Tomato'
        assert cnq.parent.index.name == 'Vegetable'
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Sauce Tomato\' has been set as a '
                'subcategory of \'Tomato\'.') in msgs

    def test_save_row_to_db_new_with_existing_parent(self, db):
        """Create a new CommonName with an existing parent from the db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cnp = CommonName(name='Tomato')
        cnp.index = Index(name='Vegetable')
        db.session.add(cnp)
        db.session.commit()
        cn = CommonName(name='Sauce Tomato')
        cn.index = Index(name='Vegetable')
        cn.parent = CommonName(name='Tomato')
        cn.parent.index = Index(name='Vegetable')
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Sauce Tomato')\
            .one_or_none()
        assert cnq.parent is cnp

    def test_save_row_to_db_new_with_desc(self, db):
        """Create a new CommonName with a description."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove', description='A bit spotty.')
        cn.index = Index(name='Perennial')
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Foxglove')\
            .one_or_none()
        assert cnq.description == '<p>A bit spotty.</p>'
        messages.seek(0)
        msgs = messages.read()
        assert 'Description for the CommonName \'Foxglove\' set to' in msgs

    def test_save_row_to_db_existing_with_new_desc(self, db):
        """Edit desc of existing CommonName."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove', description='A bit spotty.')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove', description='More dots!')
        cn2.index = Index(name='Perennial')
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert cn.description == '<p>More dots!</p>'
        messages.seek(0)
        msgs = messages.read()
        assert ('Description for the CommonName \'Foxglove\' set to: '
                '<p>More dots!</p>') in msgs

    def test_save_row_to_db_existing_clears_desc(self, db):
        """Clear description of existing CommonName if row has none."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove', description='A bit spotty.')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert cn.description is None
        messages.seek(0)
        msgs = messages.read()
        assert ('Description for the CommonName \'Foxglove\' has been '
                'cleared.') in msgs

    def test_save_row_to_db_new_with_instructions(self, db):
        """Create a new CommonName with planting instructions."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove', instructions='Just add water!')
        cn.index = Index(name='Perennial')
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Foxglove')\
            .one_or_none()
        assert cnq.instructions == '<p>Just add water!</p>'
        messages.seek(0)
        msgs = messages.read()
        assert ('Planting instructions for the CommonName \'Foxglove\' set '
                'to: <p>Just add water!</p>') in msgs

    def test_save_row_to_db_existing_with_new_instructions(self, db):
        """Change instructions of existing CommonName in db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove', instructions='Just add water!')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove', instructions='Put them in soil.')
        cn2.index = Index(name='Perennial')
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert cn.instructions == '<p>Put them in soil.</p>'
        messages.seek(0)
        msgs = messages.read()
        assert ('Planting instructions for the CommonName \'Foxglove\' set '
                'to: <p>Put them in soil.</p>') in msgs

    def test_save_row_to_db_existing_clears_instructions(self, db):
        """Clear instructions from CommonName with blank inst. in row."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove', instructions='Just add water!')
        cn.index = Index(name='Perennial')
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert cn.instructions is None
        messages.seek(0)
        msgs = messages.read()
        assert ('Planting instructions for the CommonName \'Foxglove\' have '
                'been cleared.') in msgs

    def test_save_row_to_db_new_with_synonyms(self, db):
        """Create a CommonName with synonyms."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cn.synonyms_string = 'Digitalis'
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Foxglove')\
            .one_or_none()
        assert cnq.synonyms_string == 'Digitalis'
        messages.seek(0)
        msgs = messages.read()
        assert ('Synonyms for the CommonName \'Foxglove\' set to: '
                'Digitalis') in msgs

    def test_save_row_to_db_existing_with_new_synonyms(self, db):
        """Change synonyms of existing CommonName."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cn.synonyms_string = 'Digitalis'
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        cn2.synonyms_string = 'Vulpine Handwarmer'
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert cn.synonyms_string == 'Vulpine Handwarmer'
        messages.seek(0)
        msgs = messages.read()
        assert ('Synonyms for the CommonName \'Foxglove\' set to: Vulpine '
                'Handwarmer') in msgs

    def test_save_row_to_db_existing_clears_synonyms(self, db):
        """Clear synonyms from existing CommonName."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cn.synonyms_string = 'Digitalis'
        db.session.add(cn)
        db.session.commit()
        cn2 = CommonName(name='Foxglove')
        cn2.index = Index(name='Perennial')
        cnws.add_one(cn2)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Foxglove')\
            .one_or_none()
        assert cnq is cn
        assert not cn.synonyms_string
        messages.seek(0)
        msgs = messages.read()
        assert ('Synonyms for the CommonName \'Foxglove\' have been '
                'cleared.') in msgs

    def test_save_row_to_db_new_visible(self, db):
        """Set visible to False if false in row."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cn.visible = False
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Foxglove')\
            .one_or_none()
        assert not cnq.visible
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Foxglove\' is visible on generated '
                'pages.') in msgs

    def test_save_row_to_db_new_visible(self, db):
        """Set visible to True if true in row."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cn.visible = True
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Foxglove')\
            .one_or_none()
        assert cnq.visible
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Foxglove\' is visible on generated '
                'pages.') in msgs

    def test_save_row_to_db_new_null_visible(self, db):
        """Treat null value for visible as False."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cn.visible = None
        cnws.add_one(cn)
        assert cnws.save_row_to_db(row=2, stream=messages)
        cnq = CommonName.query\
            .filter(CommonName.name == 'Foxglove')\
            .one_or_none()
        assert not cnq.visible
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Foxglove\' is not visible on generated '
                'pages.') in msgs


class TestBotanicalNamesWorksheetWithDB:
    """Test methods of BotanicalNamesWorksheet that require db use."""
    def test_save_row_to_db_bad_botanical_name(self, db):
        """Print a message and return False if bn is not valid.

        This check should happen before entering the main body of logic in
        save_row_to_db, so it shouldn't execute any code past the error check,
        but it should not raise an exception, as we just want to skip adding
        bad botanical names and carry on with program execution.
        """
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName()
        bn.name = 'invalid Botanical Name'
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        bnws.add_one(bn)
        assert not bnws.save_row_to_db(2, stream=messages)
        messages.seek(0)
        msgs = messages.read()
        assert msgs == ('Could not add the BotanicalName \'invalid Botanical '
                        'Name\' because it does not appear to be a validly '
                        'formatted botanical name.\n')

    @mock.patch('app.seeds.excel.db.session.flush')
    def test_save_row_to_db_no_changes(self, m_f, db):
        """Don't change existing BotanicalName if row contains same data."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        db.session.add(bn)
        db.session.commit()
        bn2 = BotanicalName(name='Digitalis purpurea')
        bn2.common_names = [CommonName(name='Foxglove')]
        bn2.common_names[0].index = Index(name='Perennial')
        bnws.add_one(bn2)
        assert not bnws.save_row_to_db(2, stream=messages)
        assert not m_f.called
        messages.seek(0)
        msgs = messages.read()
        assert ('The BotanicalName \'Digitalis purpurea\' has been loaded '
                'from the database.') in msgs
        assert ('No changes were made to the BotanicalName \'Digitalis '
                'purpurea\'.') in msgs

    def test_save_row_to_db_new_no_opts(self, db):
        """Add a new BotanicalName to the database."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        bnws.add_one(bn)
        assert bnws.save_row_to_db(2, stream=messages)
        bnq = BotanicalName.query\
            .filter(BotanicalName.name == 'Digitalis purpurea')\
            .one_or_none()
        assert bnq.name == 'Digitalis purpurea'
        assert bnq.common_names[0].name == 'Foxglove'
        assert bnq.common_names[0].index.name == 'Perennial'
        messages.seek(0)
        msgs = messages.read()
        assert ('The BotanicalName \'Digitalis purpurea\' does not yet exist '
                'in the database') in msgs
        assert ('The CommonName \'Foxglove\' has been added to CommonNames '
                'for the BotanicalName \'Digitalis purpurea\'.') in msgs
        assert ('Changes to the BotanicalName \'Digitalis purpurea\' have '
                'been flushed to the database.') in msgs

    def test_save_row_to_db_new_with_synonyms(self, db):
        """Add a new BotanicalName with synonyms."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        bn.synonyms_string = 'Digitalis über alles'
        bnws.add_one(bn)
        assert bnws.save_row_to_db(2, stream=messages)
        bnq = BotanicalName.query\
            .filter(BotanicalName.name == 'Digitalis purpurea')\
            .one_or_none()
        assert bnq.synonyms_string == 'Digitalis über alles'
        messages.seek(0)
        msgs = messages.read()
        assert ('Synonyms for the BotanicalName \'Digitalis purpurea\' set '
                'to: Digitalis über alles') in msgs

    def test_save_row_to_db_existing_with_synonyms(self, db):
        """Change synonyms for an existing BotanicalName."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        bn.synonyms_string = 'Digitalis über alles'
        db.session.add(bn)
        db.session.commit()
        bn2 = BotanicalName(name='Digitalis purpurea')
        bn2.common_names = [CommonName(name='Foxglove')]
        bn2.common_names[0].index = Index(name='Perennial')
        bn2.synonyms_string = 'Innagada davida'
        bnws.add_one(bn2)
        assert bnws.save_row_to_db(2, stream=messages)
        bnq = BotanicalName.query\
            .filter(BotanicalName.name == 'Digitalis purpurea')\
            .one_or_none()
        assert bnq is bn
        assert bn.synonyms_string == 'Innagada davida'
        messages.seek(0)
        msgs = messages.read()
        assert ('Synonyms for the BotanicalName \'Digitalis purpurea\' set '
                'to: Innagada davida') in msgs

    def test_save_row_to_db_existing_clears_synonyms(self, db):
        """Clear synonyms if row does not have them."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        bn.synonyms_string = 'Digitalis über alles'
        db.session.add(bn)
        db.session.commit()
        bn2 = BotanicalName(name='Digitalis purpurea')
        bn2.common_names = [CommonName(name='Foxglove')]
        bn2.common_names[0].index = Index(name='Perennial')
        bnws.add_one(bn2)
        assert bnws.save_row_to_db(2, stream=messages)
        bnq = BotanicalName.query\
            .filter(BotanicalName.name == 'Digitalis purpurea')\
            .one_or_none()
        assert bnq is bn
        assert not bn.synonyms_string
        messages.seek(0)
        msgs = messages.read()
        assert ('Synonyms for the BotanicalName \'Digitalis purpurea\' have '
                'been cleared.') in msgs

    def test_save_row_to_db_existing_add_common_name(self, db):
        """Add a CommonName to existing BotanicalName."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        db.session.add(bn)
        db.session.commit()
        bn2 = BotanicalName(name='Digitalis purpurea')
        bn2.common_names = [CommonName(name='Foxglove'),
                            CommonName(name='Digitalis')]
        bn2.common_names[0].index = Index(name='Perennial')
        bn2.common_names[1].index = Index(name='Perennial')
        bnws.add_one(bn2)
        assert bnws.save_row_to_db(2, stream=messages)
        bnq = BotanicalName.query\
            .filter(BotanicalName.name == 'Digitalis purpurea')\
            .one_or_none()
        assert bnq is bn
        assert bn.common_names[0].name == 'Foxglove'
        assert bn.common_names[1].name == 'Digitalis'
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Digitalis\' has been added to CommonNames '
                'for the BotanicalName \'Digitalis purpurea\'.') in msgs

    def test_save_row_to_db_existing_common_name_switcheroo(self, db):
        """Replace a CommonName in an existing BotanicalName."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_names = [CommonName(name='Foxglove')]
        bn.common_names[0].index = Index(name='Perennial')
        db.session.add(bn)
        db.session.commit()
        bn2 = BotanicalName(name='Digitalis purpurea')
        bn2.common_names = [CommonName(name='Digitalis')]
        bn2.common_names[0].index = Index(name='Perennial')
        bnws.add_one(bn2)
        assert bnws.save_row_to_db(2, stream=messages)
        bnq = BotanicalName.query\
            .filter(BotanicalName.name == 'Digitalis purpurea')\
            .one_or_none()
        assert bnq is bn
        assert len(bn.common_names) == 1
        assert bn.common_names[0].name == 'Digitalis'
        messages.seek(0)
        msgs = messages.read()
        assert ('The CommonName \'Digitalis\' has been added to CommonNames '
                'for the BotanicalName \'Digitalis purpurea\'.') in msgs
        assert ('The CommonName \'Foxglove\' has been removed from '
                'CommonNames for the BotanicalName \'Digitalis '
                'purpurea\'.') in msgs


class TestSeriesWorksheetWithDB:
    """Test methods of the SeriesWorksheet that use the database."""
    @mock.patch('app.seeds.excel.db.session.flush')
    def test_save_row_to_db_no_changes(self, m_f, db):
        """If it ain't broke don't fix it."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        sr.position = Series.BEFORE_CULTIVAR
        db.session.add(sr)
        db.session.commit()
        sr2 = Series(name='Polkadot')
        sr2.common_name = CommonName(name='Foxglove')
        sr2.common_name.index = Index(name='Perennial')
        sr2.position = Series.BEFORE_CULTIVAR
        srws.add_one(sr2)
        assert not srws.save_row_to_db(2, stream=messages)
        assert not m_f.called
        messages.seek(0)
        msgs = messages.read()
        assert 'No changes were made to the Series \'Polkadot\'.' in msgs

    def test_save_row_to_db_new_no_desc(self, db):
        """Save new Series to database."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        sr.position = Series.BEFORE_CULTIVAR
        srws.add_one(sr)
        assert srws.save_row_to_db(2, stream=messages)
        srq = Series.query.filter(Series.name == 'Polkadot').one_or_none()
        assert srq.name == 'Polkadot'
        assert srq.common_name.name == 'Foxglove'
        assert srq.common_name.index.name == 'Perennial'
        messages.seek(0)
        msgs = messages.read()
        print(msgs)
        assert ('The Series \'Polkadot\' does not yet exist in the database, '
                'so it has been created.') in msgs
        assert ('The Series name \'Polkadot\' will be placed before the '
                'Cultivar name for each Cultivar in the Series.') in msgs

    def test_save_row_to_db_new_with_desc(self, db):
        """Save a new Series with a description to database."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot', description='A bit spotty.')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        sr.position = Series.BEFORE_CULTIVAR
        srws.add_one(sr)
        assert srws.save_row_to_db(2, stream=messages)
        srq = Series.query.filter(Series.name == 'Polkadot').one_or_none()
        assert srq.description == '<p>A bit spotty.</p>'
        messages.seek(0)
        msgs = messages.read()
        assert ('Description for the Series \'Polkadot\' set to: <p>A bit '
                'spotty.</p>') in msgs

    def test_save_row_to_db_existing_change_desc(self, db):
        """Change the description of an existing Series."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot', description='A bit spotty.')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        sr.position = Series.BEFORE_CULTIVAR
        db.session.add(sr)
        db.session.commit()
        sr2 = Series(name='Polkadot', description='More dots!')
        sr2.common_name = CommonName(name='Foxglove')
        sr2.common_name.index = Index(name='Perennial')
        sr2.position = Series.BEFORE_CULTIVAR
        srws.add_one(sr2)
        assert srws.save_row_to_db(2, stream=messages)
        srq = Series.query.filter(Series.name == 'Polkadot').one_or_none()
        assert srq is sr
        assert sr.description == '<p>More dots!</p>'
        messages.seek(0)
        msgs = messages.read()
        assert ('Description for the Series \'Polkadot\' set to: <p>More '
                'dots!</p>') in msgs

    def test_save_row_to_db_existing_clears_desc(self, db):
        """Clear description of existing Series if row has no desc."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot', description='A bit spotty.')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        sr.position = Series.BEFORE_CULTIVAR
        db.session.add(sr)
        db.session.commit()
        sr2 = Series(name='Polkadot')
        sr2.common_name = CommonName(name='Foxglove')
        sr2.common_name.index = Index(name='Perennial')
        sr2.position = Series.BEFORE_CULTIVAR
        srws.add_one(sr2)
        assert srws.save_row_to_db(2, stream=messages)
        srq = Series.query.filter(Series.name == 'Polkadot').one_or_none()
        assert srq is sr
        assert not sr.description
        messages.seek(0)
        msgs = messages.read()
        assert ('Description for the Series \'Polkadot\' has been '
                'cleared') in msgs

    def test_save_row_to_db_existing_change_position(self, db):
        """Change the position of an existing Series."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        sr.position = Series.BEFORE_CULTIVAR
        db.session.add(sr)
        db.session.commit()
        sr2 = Series(name='Polkadot')
        sr2.common_name = CommonName(name='Foxglove')
        sr2.common_name.index = Index(name='Perennial')
        sr2.position = Series.AFTER_CULTIVAR
        srws.add_one(sr2)
        assert srws.save_row_to_db(2, stream=messages)
        srq = Series.query.filter(Series.name == 'Polkadot').one_or_none()
        assert srq is sr
        assert sr.position == Series.AFTER_CULTIVAR
        messages.seek(0)
        msgs = messages.read()
        assert ('The Series name \'Polkadot\' will be placed after the '
                'Cultivar name for each Cultivar in the Series.') in msgs


class TestCultivarsWorksheetWithDB:
    """Test methods of CultivarsWorksheet which utilize the database."""
    @mock.patch('app.seeds.excel.db.session.flush')
    def test_save_row_to_db_no_changes(self, m_f, db):
        """Do not change that which does not need to be changed."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        db.session.add(cv)
        db.session.commit()
        cv2 = Cultivar(name='Foxy')
        cv2.common_name = CommonName(name='Foxglove')
        cv2.common_name.index = Index(name='Perennial')
        cv2.in_stock = True
        cv2.active = True
        cv2.visible = False
        cvws.add_one(cv2)
        assert not cvws.save_row_to_db(2, stream=messages)
        messages.seek(0)
        msgs = messages.read()
        assert not m_f.called
        assert 'No changes were made' in msgs

    @mock.patch('app.seeds.excel.db.session.flush')
    def test_save_row_to_db_no_changes_with_series(self, m_f, db):
        """Being in a Series shouldn't break the no changes scenario."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Petra')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.series = Series(name='Polkadot')
        cv.series.common_name = cv.common_name
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        db.session.add(cv)
        db.session.commit()
        cv2 = Cultivar(name='Petra')
        cv2.common_name = CommonName(name='Foxglove')
        cv2.common_name.index = Index(name='Perennial')
        cv2.series = Series(name='Polkadot')
        cv2.series.common_name = cv2.common_name
        cv2.in_stock = True
        cv2.active = True
        cv2.visible = False
        cvws.add_one(cv2)
        assert not cvws.save_row_to_db(2, stream=messages)
        messages.seek(0)
        msgs = messages.read()
        assert not m_f.called
        assert 'No changes were made' in msgs

    def test_save_row_to_db_new_no_opts(self, db):
        """Save a new Cultivar to the database using data from row."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        assert cvq.name == 'Foxy'
        assert cvq.common_name.name == 'Foxglove'
        assert cvq.common_name.index.name == 'Perennial'
        assert cvq.in_stock
        assert cvq.active
        assert not cvq.visible
        messages.seek(0)
        msgs = messages.read()
        assert ('Changes to the Cultivar \'Foxy Foxglove\' have been flushed '
                'to the database.') in msgs

    def test_save_row_to_db_new_with_series(self, db):
        """Save a new Cultivar with a Series."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Petra')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.series = Series(name='Polkadot')
        cv.series.common_name = cv.common_name
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cv = Cultivar.query.filter(Cultivar.name == 'Petra').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cv.series.name == 'Polkadot'
        assert 'The Series \'Polkadot\' does not yet exist' in msgs
        assert ('Series for the Cultivar \'Polkadot Petra Foxglove\' set to: '
                'Polkadot') in msgs

    def test_save_row_to_db_new_with_existing_series(self, db):
        """Use Series from the database if it already exists."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        sr = Series(name='Polkadot')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        db.session.add(sr)
        db.session.commit()
        cv = Cultivar(name='Petra')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.series = Series(name='Polkadot')
        cv.series.common_name = cv.common_name
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cv = Cultivar.query.filter(Cultivar.name == 'Petra').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cv.series is sr
        assert 'The Series \'Polkadot\' has been loaded from' in msgs

    def test_save_row_to_db_new_with_new_botanical_name(self, db):
        """Create a BotanicalName when creating Cultivar."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.botanical_name = BotanicalName(name='Digitalis purpurea')
        cv.botanical_name.common_names.append(cv.common_name)
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cv = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cv.botanical_name.name == 'Digitalis purpurea'
        assert 'The BotanicalName \'Digitalis purpurea\' does not yet' in msgs
        assert ('BotanicalName for the Cultivar \'Foxy Foxglove\' set to: '
                'Digitalis purpurea') in msgs

    def test_save_row_to_db_new_with_new_bad_botanical_name(self, db):
        """Fix bad BotanicalName and use fixed version."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.botanical_name = BotanicalName()
        cv.botanical_name.name = 'digitalis purpurea'
        cv.botanical_name.common_names.append(cv.common_name)
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cv = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cv.botanical_name.name == 'Digitalis purpurea'
        assert ('The BotanicalName \'digitalis purpurea\' does not appear to '
                'be a validly formatted botanical name. In an attempt to fix '
                'it, it has been changed to: \'Digitalis purpurea\'') in msgs

    def test_save_row_to_db_new_with_existing_botanical_name(self, db):
        """Use existing BotanicalName instead of new one."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        bn = BotanicalName(name='Digitalis purpurea')
        bn.common_name = CommonName(name='Foxglove')
        bn.common_name.index = Index(name='Perennial')
        db.session.add(bn)
        db.session.commit()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.botanical_name = BotanicalName(name='Digitalis purpurea')
        cv.botanical_name.common_names.append(cv.common_name)
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cv = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cv.botanical_name is bn
        assert 'The BotanicalName \'Digitalis purpurea\' has been load' in msgs

    def test_save_row_to_db_existing_with_new_botanical_name(self, db):
        """Add a new BotanicalName to an existing Cultivar."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        db.session.add(cv)
        db.session.commit()
        cv2 = Cultivar(name='Foxy')
        cv2.common_name = CommonName(name='Foxglove')
        cv2.common_name.index = Index(name='Perennial')
        cv2.botanical_name = BotanicalName(name='Digitalis purpurea')
        cv2.botanical_name.common_names.append(cv2.common_name)
        cv2.in_stock = True
        cv2.active = True
        cv2.visible = False
        cvws.add_one(cv2)
        assert cvws.save_row_to_db(row=2, stream=messages)
        messages.seek(0)
        msgs = messages.read()
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        assert cvq is cv
        assert cv.botanical_name.name == 'Digitalis purpurea'
        assert ('BotanicalName for the Cultivar \'Foxy Foxglove\' set to: '
                'Digitalis purpurea') in msgs

    @mock.patch('app.seeds.excel.Image.exists')
    def test_save_row_to_db_new_with_new_thumbnail_file_exists(self, m_i, db):
        """Add a Cultivar with a thumbnail filename."""
        m_i.return_value = True
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.thumbnail = Image(filename='foxy.jpg')
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq.thumbnail.filename == 'foxy.jpg'
        assert 'The Image with the filename \'foxy.jpg\' does not yet' in msgs

    @mock.patch('app.seeds.excel.Image.exists')
    def test_save_row_to_db_new_with_new_thumbnail_no_file(self, m_i, db):
        """Notify the user that the file doesn't exist."""
        m_i.return_value = False
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.thumbnail = Image(filename='foxy.jpg')
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq.thumbnail.filename == 'foxy.jpg'
        assert ('WARNING: The image file \'foxy.jpg\' set as the thumbnail '
                'for the Cultivar \'Foxy Foxglove\' does not exist!') in msgs

    def test_save_row_to_db_new_with_existing_thumbnail(self, db):
        """Load an existing Image from the database instead of creating it."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        tn = Image(filename='foxy.jpg')
        db.session.add(tn)
        db.session.commit()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.thumbnail = Image(filename='foxy.jpg')
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq.thumbnail is tn
        assert 'The Image with the filename \'foxy.jpg\' has been load' in msgs

    def test_save_row_to_db_existing_with_new_thumbnail(self, db):
        """Set a new thumbnail for an existing Cultivar."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        db.session.add(cv)
        db.session.commit()
        cv2 = Cultivar(name='Foxy')
        cv2.common_name = CommonName(name='Foxglove')
        cv2.common_name.index = Index(name='Perennial')
        cv2.thumbnail = Image(filename='foxy.jpg')
        cv2.in_stock = True
        cv2.active = True
        cv2.visible = False
        cvws.add_one(cv2)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq is cv
        assert cv.thumbnail.filename == 'foxy.jpg'
        assert ('The Image with the filename \'foxy.jpg\' has been set as the '
                'thumbnail for the Cultivar \'Foxy Foxglove\'.') in msgs

    def test_save_row_to_db_new_with_description(self, db):
        """Save a Cultivar with a description to db."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy', description='Like a lady.')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq.description == '<p>Like a lady.</p>'
        assert ('Description for the Cultivar \'Foxy Foxglove\' set to: <p>Like a lady.</p>') in msgs

    def test_save_row_to_db_existing_changes_description(self, db):
        """Set a new description for an existing Cultivar."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy', description='Like a lady.')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        db.session.add(cv)
        db.session.commit()
        cv2 = Cultivar(name='Foxy', description='Like a Hendrix song.')
        cv2.common_name = CommonName(name='Foxglove')
        cv2.common_name.index = Index(name='Perennial')
        cv2.in_stock = True
        cv2.active = True
        cv2.visible = False
        cvws.add_one(cv2)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq is cv
        assert cv.description == '<p>Like a Hendrix song.</p>'
        assert ('Description for the Cultivar \'Foxy Foxglove\' set to: '
                '<p>Like a Hendrix song.</p>') in msgs

    def test_save_row_to_db_existing_clears_description(self, db):
        """Clear the description for an existing Cultivar."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy', description='Like a lady.')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        db.session.add(cv)
        db.session.commit()
        cv2 = Cultivar(name='Foxy')
        cv2.common_name = CommonName(name='Foxglove')
        cv2.common_name.index = Index(name='Perennial')
        cv2.in_stock = True
        cv2.active = True
        cv2.visible = False
        cvws.add_one(cv2)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq is cv
        assert cv.description is None
        assert ('Description for the Cultivar \'Foxy Foxglove\' has been '
                'cleared.') in msgs

    def test_save_row_to_db_new_with_synonyms(self, db):
        """Save a Cultivar with synonyms."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.synonyms_string = 'Vulpine'
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq.synonyms_string == 'Vulpine'
        assert ('Synonyms for the Cultivar \'Foxy Foxglove\' set to: '
                'Vulpine') in msgs

    def test_save_row_to_db_existing_changes_synonyms(self, db):
        """Change old synonyms to new ones."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.synonyms_string = 'Vulpine'
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        db.session.add(cv)
        db.session.commit()
        cv2 = Cultivar(name='Foxy')
        cv2.common_name = CommonName(name='Foxglove')
        cv2.common_name.index = Index(name='Perennial')
        cv2.synonyms_string = 'Fauxy'
        cv2.in_stock = True
        cv2.active = True
        cv2.visible = False
        cvws.add_one(cv2)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq is cv
        assert cv.synonyms_string == 'Fauxy'
        assert ('Synonyms for the Cultivar \'Foxy Foxglove\' set to: '
                'Fauxy') in msgs

    def test_save_row_to_db_existing_clears_synonyms(self, db):
        """Clear the synonyms for an existing Cultivar."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.synonyms_string = 'Vulpine'
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        db.session.add(cv)
        db.session.commit()
        cv2 = Cultivar(name='Foxy')
        cv2.common_name = CommonName(name='Foxglove')
        cv2.common_name.index = Index(name='Perennial')
        cv2.in_stock = True
        cv2.active = True
        cv2.visible = False
        cvws.add_one(cv2)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq is cv
        assert not cv.synonyms
        assert ('Synonyms for the Cultivar \'Foxy Foxglove\' have been '
                'cleared.') in msgs

    def test_save_row_to_db_new_with_new_until(self, db):
        """Create a Cultivar with new_until data."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.new_until = datetime.date(2012, 12, 21)
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq.new_until == datetime.date(2012, 12, 21)
        assert ('The Cultivar \'Foxy Foxglove\' has been set as new until '
                '12/21/2012') in msgs

    def test_save_row_to_db_existing_with_new_until(self, db):
        """Add a new_until value to an existing Cultivar."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        db.session.add(cv)
        db.session.commit()
        cv2 = Cultivar(name='Foxy')
        cv2.common_name = CommonName(name='Foxglove')
        cv2.common_name.index = Index(name='Perennial')
        cv2.new_until = datetime.date(2012, 12, 21)
        cv2.in_stock = True
        cv2.active = True
        cv2.visible = False
        cvws.add_one(cv2)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        assert cvq is cv
        assert cv.new_until == datetime.date(2012, 12, 21)

    def test_save_row_to_db_existing_changes_new_until(self, db):
        """Replace existing new_until with new value."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.new_until = datetime.date(2012, 12, 21)
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        db.session.add(cv)
        db.session.commit()
        cv2 = Cultivar(name='Foxy')
        cv2.common_name = CommonName(name='Foxglove')
        cv2.common_name.index = Index(name='Perennial')
        cv2.new_until = datetime.date(2012, 12, 12)
        cv2.in_stock = True
        cv2.active = True
        cv2.visible = False
        cvws.add_one(cv2)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        assert cvq is cv
        assert cv.new_until == datetime.date(2012, 12, 12)

    def test_save_row_to_db_existing_removes_new_until(self, db):
        """Remove new_until if row lacks it."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.new_until = datetime.date(2012, 12, 21)
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        db.session.add(cv)
        db.session.commit()
        cv2 = Cultivar(name='Foxy')
        cv2.common_name = CommonName(name='Foxglove')
        cv2.common_name.index = Index(name='Perennial')
        cv2.in_stock = True
        cv2.active = True
        cv2.visible = False
        cvws.add_one(cv2)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq is cv
        assert cv.new_until is None
        assert 'The Cultivar \'Foxy Foxglove\' is no longer set as new' in msgs

    def test_save_row_to_db_new_default_bools(self, db):
        """Create new Cultivar w/ defaults for in_stock, active, visible."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq.in_stock
        assert cvq.active
        assert not cvq.visible
        assert 'The Cultivar \'Foxy Foxglove\' is in stock' in msgs
        assert 'The Cultivar \'Foxy Foxglove\' is active' in msgs
        assert 'The Cultivar \'Foxy Foxglove\' will not be shown' in msgs

    def test_save_row_to_db_new_opposite_bools(self, db):
        """Create new Cultivar w/ opposites for in_stock, active, visible."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = False
        cv.active = False
        cv.visible = True
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert not cvq.in_stock
        assert not cvq.active
        assert cvq.visible
        assert 'The Cultivar \'Foxy Foxglove\' is out of stock' in msgs
        assert 'The Cultivar \'Foxy Foxglove\' is inactive' in msgs
        assert 'The Cultivar \'Foxy Foxglove\' will be shown on' in msgs

    def test_save_row_to_db_existing_flip_bools(self, db):
        """Change values of bools."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = True
        cv.active = True
        cv.visible = False
        db.session.add(cv)
        db.session.commit()
        cv2 = Cultivar(name='Foxy')
        cv2.common_name = CommonName(name='Foxglove')
        cv2.common_name.index = Index(name='Perennial')
        cv2.in_stock = False
        cv2.active = False
        cv2.visible = True
        cvws.add_one(cv2)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert cvq is cv
        assert not cv.in_stock
        assert not cv.active
        assert cv.visible
        assert 'The Cultivar \'Foxy Foxglove\' is out of stock' in msgs
        assert 'The Cultivar \'Foxy Foxglove\' is inactive' in msgs
        assert 'The Cultivar \'Foxy Foxglove\' will be shown on' in msgs
        messages = StringIO()
        cv3 = Cultivar(name='Foxy')
        cv3.common_name = CommonName(name='Foxglove')
        cv3.common_name.index = Index(name='Perennial')
        cv3.in_stock = True
        cv3.active = True
        cv3.visible = False
        cvws.add_one(cv3)
        assert cvws.save_row_to_db(row=3, stream=messages)
        messages.seek(0)
        msgs = messages.read()
        assert cv.in_stock
        assert cv.active
        assert not cv.visible
        assert 'The Cultivar \'Foxy Foxglove\' is in stock' in msgs
        assert 'The Cultivar \'Foxy Foxglove\' is active' in msgs
        assert 'The Cultivar \'Foxy Foxglove\' will not be shown' in msgs

    def test_save_row_to_db_new_null_bools(self, db):
        """Treat null booleans as False."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = None
        cv.active = None
        cv.visible = None
        cvws.add_one(cv)
        assert cvws.save_row_to_db(row=2, stream=messages)
        cvq = Cultivar.query.filter(Cultivar.name == 'Foxy').one_or_none()
        messages.seek(0)
        msgs = messages.read()
        assert not cvq.in_stock
        assert not cvq.active
        assert not cvq.visible
        assert 'The Cultivar \'Foxy Foxglove\' is out of stock' in msgs
        assert 'The Cultivar \'Foxy Foxglove\' is inactive' in msgs
        assert 'The Cultivar \'Foxy Foxglove\' will not be' in msgs


class TestPacketsWorksheetWithDB:
    """Test methods of PacketsWorksheet that use the database."""
    @mock.patch('app.seeds.excel.db.session.flush')
    def test_save_row_to_db_no_changes(self, m_f, db):
        """Don't change a Packet if row data is the same."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        pws = PacketsWorksheet(ws)
        pws.setup()
        pkt = Packet(sku='8675309',
                     price='3.50',
                     quantity=Quantity(value=100, units='seeds'))
        pkt.cultivar = Cultivar(name='Foxy')
        pkt.cultivar.common_name = CommonName(name='Foxglove')
        pkt.cultivar.common_name.index = Index(name='Perennial')
        db.session.add(pkt)
        db.session.commit()
        pkt2 = Packet(sku='8675309', price='3.50')
        pkt2.quantity = Quantity(value=100, units='seeds')
        pkt2.cultivar = Cultivar(name='Foxy')
        pkt2.cultivar.common_name = CommonName(name='Foxglove')
        pkt2.cultivar.common_name.index = Index(name='Perennial')
        pws.add_one(pkt2)
        assert not pws.save_row_to_db(row=2, stream=messages)
        messages.seek(0)
        msgs = messages.read()
        assert not m_f.called
        assert 'No changes were made to the Packet' in msgs

    def test_save_row_to_db_new(self, db):
        """Save a new Packet to the database."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        pws = PacketsWorksheet(ws)
        pws.setup()
        pkt = Packet(sku='8675309',
                     price='3.50',
                     quantity=Quantity(value=100, units='seeds'))
        pkt.cultivar = Cultivar(name='Foxy')
        pkt.cultivar.common_name = CommonName(name='Foxglove')
        pkt.cultivar.common_name.index = Index(name='Perennial')
        pws.add_one(pkt)
        assert pws.save_row_to_db(row=2, stream=messages)
        messages.seek(0)
        msgs = messages.read()
        pktq = Packet.query.filter(Packet.sku == '8675309').one_or_none()
        assert pktq.price == '3.50'
        assert pktq.quantity.value == 100
        assert pktq.quantity.units == 'seeds'
        assert 'The Packet with SKU \'8675309\' does not yet exist' in msgs
        assert 'Changes to the Packet' in msgs

    def test_save_row_to_db_new_to_existing_cultivar(self, db):
        """Add a Packet to an existing Cultivar."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        pws = PacketsWorksheet(ws)
        pws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        db.session.add(cv)
        db.session.commit()
        pkt = Packet(sku='8675309',
                     price='3.50',
                     quantity=Quantity(value=100, units='seeds'))
        pkt.cultivar = Cultivar(name='Foxy')
        pkt.cultivar.common_name = CommonName(name='Foxglove')
        pkt.cultivar.common_name.index = Index(name='Perennial')
        pws.add_one(pkt)
        assert pws.save_row_to_db(row=2, stream=messages)
        pktq = Packet.query.filter(Packet.sku == '8675309').one_or_none()
        assert pktq.cultivar is cv

    def test_save_row_to_db_existing_new_price(self, db):
        """Change the price of an existing packet."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        pws = PacketsWorksheet(ws)
        pws.setup()
        pkt = Packet(sku='8675309',
                     price='3.50',
                     quantity=Quantity(value=100, units='seeds'))
        pkt.cultivar = Cultivar(name='Foxy')
        pkt.cultivar.common_name = CommonName(name='Foxglove')
        pkt.cultivar.common_name.index = Index(name='Perennial')
        db.session.add(pkt)
        db.session.commit()
        pkt2 = Packet(sku='8675309', price='9.99')
        pkt2.quantity = Quantity(value=100, units='seeds')
        pkt2.cultivar = Cultivar(name='Foxy')
        pkt2.cultivar.common_name = CommonName(name='Foxglove')
        pkt2.cultivar.common_name.index = Index(name='Perennial')
        pws.add_one(pkt2)
        assert pws.save_row_to_db(row=2, stream=messages)
        messages.seek(0)
        msgs = messages.read()
        pktq = Packet.query.filter(Packet.sku == '8675309').one_or_none()
        assert pktq is pkt
        assert pkt.price == '9.99'
        assert ('The price for Packet SKU \'8675309\' has been set to: '
                '$9.99.') in msgs

    def test_save_row_to_db_existing_new_quantity(self, db):
        """Change quantity (value) of an existing packet."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        pws = PacketsWorksheet(ws)
        pws.setup()
        pkt = Packet(sku='8675309',
                     price='3.50',
                     quantity=Quantity(value=100, units='seeds'))
        pkt.cultivar = Cultivar(name='Foxy')
        pkt.cultivar.common_name = CommonName(name='Foxglove')
        pkt.cultivar.common_name.index = Index(name='Perennial')
        db.session.add(pkt)
        db.session.commit()
        pkt2 = Packet(sku='8675309', price='3.50')
        pkt2.quantity = Quantity(value=9001, units='seeds')
        pkt2.cultivar = Cultivar(name='Foxy')
        pkt2.cultivar.common_name = CommonName(name='Foxglove')
        pkt2.cultivar.common_name.index = Index(name='Perennial')
        pws.add_one(pkt2)
        assert pws.save_row_to_db(row=2, stream=messages)
        messages.seek(0)
        msgs = messages.read()
        pktq = Packet.query.filter(Packet.sku == '8675309').one_or_none()
        assert pktq is pkt
        assert pkt.quantity.value == 9001
        assert ('The quantity for the Packet SKU \'8675309\' has been set to: '
                '9001 seeds') in msgs

    def test_save_row_to_db_existing_new_units(self, db):
        """Change units of an existing packet."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        pws = PacketsWorksheet(ws)
        pws.setup()
        pkt = Packet(sku='8675309',
                     price='3.50',
                     quantity=Quantity(value=100, units='seeds'))
        pkt.cultivar = Cultivar(name='Foxy')
        pkt.cultivar.common_name = CommonName(name='Foxglove')
        pkt.cultivar.common_name.index = Index(name='Perennial')
        db.session.add(pkt)
        db.session.commit()
        pkt2 = Packet(sku='8675309', price='3.50')
        pkt2.quantity = Quantity(value=100, units='cubits')
        pkt2.cultivar = Cultivar(name='Foxy')
        pkt2.cultivar.common_name = CommonName(name='Foxglove')
        pkt2.cultivar.common_name.index = Index(name='Perennial')
        pws.add_one(pkt2)
        assert pws.save_row_to_db(row=2, stream=messages)
        messages.seek(0)
        msgs = messages.read()
        pktq = Packet.query.filter(Packet.sku == '8675309').one_or_none()
        assert pktq is pkt
        assert pkt.quantity.units == 'cubits'
        assert ('The quantity for the Packet SKU \'8675309\' has been set to: '
                '100 cubits') in msgs
