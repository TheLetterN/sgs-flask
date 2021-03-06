import datetime
import json
import pytest
from io import StringIO
from unittest import mock
from openpyxl import Workbook
from app.seeds.excel import (
    BotanicalNamesWorksheet,
    CommonNamesWorksheet,
    CultivarsWorksheet,
    IndexesWorksheet,
    queryable_dicts_to_json,
    PacketsWorksheet,
    SeedsWorkbook,
    SeedsWorksheet,
    SectionsWorksheet
)
from app.seeds.models import (
    BotanicalName,
    Section,
    CommonName,
    Cultivar,
    Image,
    Index,
    Packet,
    Quantity
)


class TestExcelFunctions:
    """Test module level functions."""
    def test_queryable_dicts_to_json(self):
        """Generate a JSON string for looking up Grows With cns/cvs.

        It can take either, as both have the queryable_dict method.
        """
        gwcn1 = CommonName(name='Foxglove')
        gwcn1.index = Index(name='Perennial')
        assert queryable_dicts_to_json([gwcn1]) == \
            json.dumps((gwcn1.queryable_dict,))
        gwcn2 = CommonName(name='Butterfly Weed')
        gwcn2.index = Index(name='Perennial')
        assert queryable_dicts_to_json([gwcn1, gwcn2]) == \
            json.dumps((gwcn1.queryable_dict, gwcn2.queryable_dict))
        gwcv1 = Cultivar(name='Soulmate')
        gwcv1.common_name = CommonName(name='Butterfly Weed')
        gwcv1.common_name.index = Index(name='Perennial')
        assert queryable_dicts_to_json([gwcv1]) == \
            json.dumps((gwcv1.queryable_dict,))
        gwcv2 = Cultivar(name='Petra')
        gwcv2.common_name = CommonName(name='Foxglove')
        gwcv2.common_name.index = Index(name='Perennial')
        gwcv2.section = Section(name='Polkadot')
        assert queryable_dicts_to_json([gwcv1, gwcv2]) == \
            json.dumps((gwcv1.queryable_dict, gwcv2.queryable_dict))

    def test_queryable_dicts_to_json_bad_args(self):
        """Raise a TypeError if any objects lack the lookup_dict method."""
        with pytest.raises(TypeError):
            queryable_dicts_to_json((1, 2, 3))
        cn1 = CommonName(name='Foxglove')
        cn1.index = Index(name='Perennial')
        cn2 = CommonName(name='Coleus')
        cn2.index = Index(name='Annual')
        idx = Index(name='Has no lookup_dict')
        with pytest.raises(TypeError):
            queryable_dicts_to_json((cn1, cn2, idx))


class TestSeedsWorksheet:
    """Test methods of the SeedsWorksheet container class.

    We use Workbook to create sheets instead of Worksheet because it ensures
    the worksheet is created as it would be when used in a workbook.
    """
    def test_getitem(self):
        """Forward the [] functionality of sheet to SeedsWorksheet."""
        legit_sheet = {'key': 'value'}
        sws = SeedsWorksheet(legit_sheet)
        assert sws['key'] == 'value'

    def test_title_getter(self):
        """Return title of given sheet."""
        sheet = mock.MagicMock()
        sheet.title = 'Worksheet: The Sheetening'
        sws = SeedsWorksheet(sheet)
        assert sws.title == 'Worksheet: The Sheetening'

    def test_title_setter(self):
        """Set title of contained sheet."""
        sheet = mock.MagicMock()
        sws = SeedsWorksheet(sheet)
        sws.title = 'Worksheet 2: Electric Boogaloo'
        assert sws._ws.title == 'Worksheet 2: Electric Boogaloo'

    def test_rows_property(self):
        """Return a tuple of tuples containing cells.

        It should return a tuple containing an empty tuple if no cells have
        data.
        """
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        assert sws.rows == ((),)
        a1 = sws._ws['A1']
        a1.value = 'One'
        assert sws.rows == ((a1,),)
        b1 = sws._ws['B1']
        b1.value = 'Two'
        assert sws.rows == ((a1, b1),)
        a2 = sws._ws['A2']
        a2.value = 'Three'
        # B2 is created with a null value to fill the grid.
        b2 = sws._ws['B2']
        assert sws.rows == ((a1, b1), (a2, b2))

    def test_active_row_property(self):
        """Return the row number of the first empty row."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        assert sws.active_row == 1
        sws._ws['A1'].value = 'One'
        assert sws.active_row == 2
        sws._ws['A2'].value = 'Two'
        assert sws.active_row == 3
        sws._ws['A3'].value = 'Three'
        assert sws.active_row == 4

    def test_data_rows_property(self):
        """Return all rows except the first row, which contains titles."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        assert sws.data_rows == tuple()
        ws['A1'] = 'One'
        assert sws.data_rows == tuple()
        ws['A2'] = 'Two'
        assert sws.data_rows == ((ws['A2'],),)
        ws['A3'] = 'Three'
        assert sws.data_rows == ((ws['A2'],), (ws['A3'],))

    def test_has_data(self):
        """Return True if there is data in cell A1, else False."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        assert not sws.has_data()
        sws._ws['A2'].value = 'Stuff'
        assert not sws.has_data()  # Okay; should never happen in usage.
        sws._ws['A1'].value = 'More stuff'
        assert sws.has_data()

    def test_cell(self):
        """Return a cell given integer coordinates."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        ws['A1'].value = 'A1'
        ws['B1'].value = 'B1'
        ws['A2'].value = 'A2'
        assert sws.cell(1, 1) is ws['A1']
        assert sws.cell(2, 1) is ws['A2']
        assert sws.cell(1, 2) is ws['B1']

    def test_set_column_titles(self):
        """Set the top row of the worksheet to a list of titles."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        titles = ['One', 'Two', 'Three', 'Four']
        sws.set_column_titles(titles)
        assert sws._ws['A1'].value == 'One'
        assert sws._ws['B1'].value == 'Two'
        assert sws._ws['C1'].value == 'Three'
        assert sws._ws['D1'].value == 'Four'

    def test_set_column_titles_existing_sheet(self):
        """Raise a ValueError if there is already data in the top row."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        titles = ['One', 'Two', 'Three', 'Four']
        sws._ws.append(titles)
        with pytest.raises(ValueError):
            sws.set_column_titles(titles)

    def test_populate_cols_dict(self):
        """Set cols keys to values in first row, and vals to col numbers."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        titles = ['One', 'Two', 'Three', 'Four']
        sws.set_column_titles(titles)
        sws.populate_cols_dict()
        assert sws.cols['One'] == 1
        assert sws.cols['Two'] == 2
        assert sws.cols['Three'] == 3
        assert sws.cols['Four'] == 4

    def test_populate_cols_dict_no_titles(self):
        """Raise a ValueError if there is no data in the first row."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        with pytest.raises(ValueError):
            sws.populate_cols_dict()

    def test_freeze_title_row(self):
        """Freeze the top row of the worksheet."""
        wb = Workbook()
        sws = SeedsWorksheet(wb.active)
        sws.freeze_title_row()
        assert sws._ws.freeze_panes == 'A2'

    @mock.patch('app.seeds.excel.SeedsWorksheet.set_column_titles')
    @mock.patch('app.seeds.excel.SeedsWorksheet.populate_cols_dict')
    def test_setup_new(self, m_pcd, m_sct):
        """Call set_column_titles and populate_cols_dict when sheet blank."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        titles = ('One', 'Two', 'Three')
        sws._setup(titles)
        assert m_pcd.called
        assert m_sct.called_with(titles)

    def test_setup_new_no_titles(self):
        """Raise a value error if _setup is run without titles on new sheet."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        with pytest.raises(ValueError):
            sws._setup()

    @mock.patch('app.seeds.excel.SeedsWorksheet.set_column_titles')
    @mock.patch('app.seeds.excel.SeedsWorksheet.populate_cols_dict')
    def test_setup_existing(self, m_pcd, m_sct):
        """Do not call set_column_titles if data in first row already."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws._ws['A1'].value = 'One'
        sws._setup()
        assert m_pcd.called
        assert not m_sct.called

    @mock.patch('app.seeds.excel.SeedsWorksheet.set_column_titles')
    @mock.patch('app.seeds.excel.SeedsWorksheet.populate_cols_dict')
    def test_setup_existing_with_titles(self, m_pcd, m_sct):
        """Set up as normal but discard new titles."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws._ws['A1'].value = 'One'
        titles = ('One', 'Two', 'Three')
        with pytest.warns(UserWarning):
            sws._setup(titles)
        assert m_pcd.called
        assert not m_sct.called

    def test_add_one(self):
        """add_one should be an abstract method in this class."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        with pytest.raises(NotImplementedError):
            sws.add_one(None)

    @mock.patch('app.seeds.excel.SeedsWorksheet.add_one')
    def test_add(self, m_ao):
        """add should call add_one for each item in iterable."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.add(('Test',), stream=messages)
        m_ao.assert_called_with('Test', stream=messages)
        messages.seek(0)
        msgs = messages.read()
        assert '-- BEGIN adding data to SeedsWorksheet. --' in msgs
        assert '-- END adding data to SeedsWorksheet. --' in msgs

    @mock.patch('app.seeds.excel.SeedsWorksheet.add_one')
    def test_add_bad_data(self, m_ao):
        """Warn user when iterable contains bad types instead of halting."""
        m_ao.side_effect = TypeError('Bad data, yo!')
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        with pytest.warns(UserWarning):
            sws.add((1, 2, 3, 4))
        assert m_ao.call_count == 4

    def test_add_not_iterable(self):
        """Do not suppress TypeError if given non-iterable."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        with pytest.raises(TypeError):
            sws.add(42)

    def test_save_row_to_db(self):
        """save_row_to_db should be an abstract method in this class."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        with pytest.raises(NotImplementedError):
            sws.save_row_to_db(None)

    def test_beautify(self):
        """Configure worksheet to be more human-readable."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws._ws.append(('One', 'Two', 'Three'))
        sws._ws.append(('Four', 'Five', 'Six'))
        sws.beautify(width=42, height=21)
        assert sws._ws.freeze_panes == 'A2'
        assert sws._ws.column_dimensions['A'].width == 42
        assert sws._ws.column_dimensions['B'].width == 42
        assert sws._ws.column_dimensions['C'].width == 42
        assert sws._ws.row_dimensions[2].height == 21


class TestIndexesWorksheet:
    """Test methods of the IndexesWorksheet container class."""
    @mock.patch('app.seeds.excel.IndexesWorksheet._setup')
    def test_setup_new(self, m_s):
        """Call _setup with titles for Indexes sheet."""
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        m_s.assert_called_with(('Index', 'Description'))

    @mock.patch('app.seeds.excel.IndexesWorksheet._setup')
    def test_setup_existing(self, m_s):
        """Call _setup with no data if sheet titles already populated."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.set_column_titles(('Index', 'Description'))
        iws = IndexesWorksheet(sws._ws)
        iws.setup()
        assert m_s.call_args_list == [mock.call()]

    def test_add_one(self):
        """Add a single Index to worksheet."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        idx = Index(name='Perennial', description='Built to last.')
        iws.add_one(idx, stream=messages)
        assert iws.cell(2, iws.cols['Index']).value == 'Perennial'
        assert iws.cell(
            2, iws.cols['Description']
        ).value == 'Built to last.'
        messages.seek(0)
        msgs = messages.read()
        assert ('Adding data from <Index "Perennial"> to row #2 of indexes '
                'worksheet.') in msgs

    def test_add_one_bad_type(self):
        """Raise a TypeError given non-Index data."""
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        with pytest.raises(TypeError):
            iws.add_one('Frogs!')


class TestCommonNamesWorksheet:
    """Test methods of the CommonNamesWorksheet container class."""
    @mock.patch('app.seeds.excel.CommonNamesWorksheet._setup')
    def test_setup_new(self, m_s):
        """Call _setup with titles for Common Names worksheet."""
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        titles = ('Index',
                  'Common Name',
                  'Description',
                  'Planting Instructions',
                  'Synonyms',
                  'Visible')
        m_s.assert_called_with(titles)

    @mock.patch('app.seeds.excel.CommonNamesWorksheet._setup')
    def test_setup_existing(self, m_s):
        """Call _setup with no data if titles row already populated."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.set_column_titles(('One', 'Two', 'Three'))
        cnws = CommonNamesWorksheet(sws._ws)
        cnws.setup()
        assert m_s.call_args_list == [mock.call()]

    def test_add_one_no_optionals(self):
        """Add a common name (with no optional data) to Common Names sheet."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cnws.add_one(cn, stream=messages)
        assert cnws.cell(2, cnws.cols['Index']).value == 'Perennial'
        assert cnws.cell(2, cnws.cols['Common Name']).value == 'Foxglove'
        assert cnws.cell(2, cnws.cols['Description']).value is None
        assert cnws.cell(2, cnws.cols['Planting Instructions']).value is None
        assert cnws.cell(2, cnws.cols['Synonyms']).value is None
        messages.seek(0)
        msgs = messages.read()
        assert ('Adding data from <CommonName "Foxglove"> to row #2 of '
                'common names worksheet.') in msgs

    def test_add_one_with_optionals(self):
        """Add a common name with optionals to Common Names sheet."""
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove',
                        description='Spotty.',
                        instructions='Just add water!')
        cn.index = Index(name='Perennial')
        cn.synonyms_string = 'Digitalis'
        cnws.add_one(cn)
        assert cnws.cell(2, cnws.cols['Description']).value == 'Spotty.'
        assert cnws.cell(
            2, cnws.cols['Planting Instructions']
        ).value == 'Just add water!'
        assert cnws.cell(2, cnws.cols['Synonyms']).value == 'Digitalis'

    def test_add_one_not_common_name(self):
        """Raise a TypeError given data that isn't a CommonName."""
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        with pytest.raises(TypeError):
            cnws.add_one(42)
        with pytest.raises(TypeError):
            cnws.add_one(Index(name='Perennial'))


class TestBotanicalNamesWorksheet:
    """Test methods of the BotanicalNamesWorksheet container class."""
    @mock.patch('app.seeds.excel.BotanicalNamesWorksheet._setup')
    def test_setup_new(self, m_s):
        """Call _setup with titles for Botanical Names worksheet."""
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        titles = ('Common Names (JSON)',
                  'Botanical Name',
                  'Synonyms')
        m_s.assert_called_with(titles)

    @mock.patch('app.seeds.excel.BotanicalNamesWorksheet._setup')
    def test_setup_existing(self, m_s):
        """Call _setup with no data if titles already present."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.set_column_titles(('One', 'Two', 'Three'))
        bnws = BotanicalNamesWorksheet(sws._ws)
        bnws.setup()
        assert m_s.call_args_list == [mock.call()]

    def test_add_one_no_optionals(self):
        """Add a BotanicalName object to Botanical Names sheet."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Innagada davida')
        cn = CommonName(name='Rock')
        cn.index = Index(name='Music')
        bn.common_names = [cn]
        bnws.add_one(bn, stream=messages)
        assert bnws.cell(
            2, bnws.cols['Common Names (JSON)']
        ).value == queryable_dicts_to_json([cn])
        assert bnws.cell(
            2, bnws.cols['Botanical Name']
        ).value == 'Innagada davida'
        assert bnws.cell(2, bnws.cols['Synonyms']).value is None
        messages.seek(0)
        msgs = messages.read()
        assert ('Adding data from <BotanicalName "Innagada davida"> to row '
                '#2 of botanical names worksheet.') in msgs

    def test_add_one_with_synonyms(self):
        """Add a BotanicalName with synonyms to Botanical Names sheet."""
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Innagada davida')
        cn = CommonName(name='Rock')
        cn.index = Index(name='Music')
        bn.common_names = [cn]
        bn.synonyms_string = 'Iron butterfly'
        bnws.add_one(bn)
        assert bnws.cell(2, bnws.cols['Synonyms']).value == 'Iron butterfly'

    def test_add_one_not_botanical_name(self):
        """Raise a TypeError given non-BotanicalName data."""
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        with pytest.raises(TypeError):
            bnws.add_one(42)
        with pytest.raises(TypeError):
            bnws.add_one(CommonName(name='Spurious'))


class TestSectionsWorksheet:
    """Test methods of the SectionsWorksheet container class."""
    @mock.patch('app.seeds.excel.SectionsWorksheet._setup')
    def test_setup_new(self, m_s):
        """Run _setup with the titles for the Section worksheet."""
        wb = Workbook()
        ws = wb.active
        srws = SectionsWorksheet(ws)
        srws.setup()
        titles = ('Common Name (JSON)', 'Section', 'Description')
        m_s.assert_called_with(titles)

    @mock.patch('app.seeds.excel.SectionsWorksheet._setup')
    def test_setup_existing(self, m_s):
        """Run _setup with no arguments if titles apready present."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.set_column_titles(('One', 'Two', 'Three'))
        srws = SectionsWorksheet(sws._ws)
        srws.setup()
        assert m_s.call_args_list == [mock.call()]

    def test_add_one_no_optionals(self):
        """Add a Section object to the Section worksheet."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        srws = SectionsWorksheet(ws)
        srws.setup()
        sr = Section(name='Polkadot')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        srws.add_one(sr, stream=messages)
        assert srws.cell(
            2, srws.cols['Common Name (JSON)']
        ).value == json.dumps(sr.common_name.queryable_dict)
        assert srws.cell(2, srws.cols['Section']).value == 'Polkadot'
        assert srws.cell(2, srws.cols['Description']).value is None
        messages.seek(0)
        msgs = messages.read()
        assert ('Adding data from <Section "Polkadot Foxglove"> to row #2 '
                'of sections worksheet') in msgs

    def test_add_one_with_description(self):
        """Set Description column's cell with Section desc."""
        wb = Workbook()
        ws = wb.active
        srws = SectionsWorksheet(ws)
        srws.setup()
        sr = Section(name='Polkadot', description='A bit spotty.')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        srws.add_one(sr)
        assert srws.cell(
            2, srws.cols['Description']
        ).value == 'A bit spotty.'

    def test_add_one_not_section(self):
        """Raise a TypeError if passed argument is not a Section object."""
        wb = Workbook()
        ws = wb.active
        srws = SectionsWorksheet(ws)
        srws.setup()
        with pytest.raises(TypeError):
            srws.add_one(42)
        with pytest.raises(TypeError):
            srws.add_one(Index(name='Perennial'))


class TestCultivarsWorksheet:
    """Test methods of the CultivarsWorksheet container class."""
    @mock.patch('app.seeds.excel.CultivarsWorksheet._setup')
    def test_setup_new(self, m_s):
        """Run _setup with the titles for the Cultivars worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        titles = ('Index',
                  'Common Name',
                  'Cultivar Name',
                  'Section',
                  'Botanical Name',
                  'Thumbnail Filename',
                  'Description',
                  'Synonyms',
                  'New Until',
                  'In Stock',
                  'Active',
                  'Visible')
        m_s.assert_called_with(titles)

    @mock.patch('app.seeds.excel.CultivarsWorksheet._setup')
    def test_setup_existing(self, m_s):
        """Run _setup with no arguments if titles already exist."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.set_column_titles(('One', 'Two', 'Three'))
        cvws = CultivarsWorksheet(sws._ws)
        cvws.setup()
        assert m_s.call_args_list == [mock.call()]

    def test_add_one_no_optionals(self):
        """Add a Cultivar to the Cultivars worksheet."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cvws.add_one(cv, stream=messages)
        assert cvws.cell(2, cvws.cols['Index']).value == 'Perennial'
        assert cvws.cell(2, cvws.cols['Common Name']).value == 'Foxglove'
        assert cvws.cell(2, cvws.cols['Cultivar Name']).value == 'Foxy'
        assert cvws.cell(2, cvws.cols['Section']).value is None
        assert cvws.cell(2, cvws.cols['Botanical Name']).value is None
        assert cvws.cell(2, cvws.cols['Thumbnail Filename']).value is None
        assert cvws.cell(2, cvws.cols['Description']).value is None
        assert cvws.cell(2, cvws.cols['Synonyms']).value is None
        assert cvws.cell(2, cvws.cols['New Until']).value is None
        assert cvws.cell(2, cvws.cols['In Stock']).value == 'False'
        assert cvws.cell(2, cvws.cols['Active']).value == 'False'
        messages.seek(0)
        msgs = messages.read()
        assert ('Adding data from <Cultivar "Foxy Foxglove"> to row #2 of '
                'cultivars worksheet.') in msgs

    def test_add_one_with_section(self):
        """Add a Cultivar with Section to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Petra')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.section = Section(name='Polkadot')
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['Section']).value == 'Polkadot'

    def test_add_one_with_botanical_name(self):
        """Add a Cultivar with a Botanical Name to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.botanical_name = BotanicalName(name='Digitalis purpurea')
        cvws.add_one(cv)
        assert cvws.cell(
            2, cvws.cols['Botanical Name']
        ).value == 'Digitalis purpurea'

    def test_add_one_with_thumbnail_filename(self):
        """Add a Cultivar with a Thumbnail Filename to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.thumbnail = Image(filename='foo.jpg')
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['Thumbnail Filename']).value == 'foo.jpg'

    def test_add_one_with_description(self):
        """Add a Cultivar with a description to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.description = 'Like a lady!'
        cvws.add_one(cv)
        assert cvws.cell(
            2, cvws.cols['Description']
        ).value == 'Like a lady!'

    def test_add_one_with_synonyms(self):
        """Add a Cultivar with synonyms to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.synonyms_string = 'Vulpine'
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['Synonyms']).value == 'Vulpine'

    def test_add_one_with_new_until(self):
        """Add a Cultivar with New Until to worksheet.

        New Until values should be dates formatted MM/DD/YYYY because 'murica.
        """
        dt = datetime.date(2012, 12, 21)
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.new_until = dt
        cvws.add_one(cv)
        assert cvws.cell(
            2, cvws.cols['New Until']
        ).value == dt.strftime('%m/%d/%Y')

    def test_add_one_in_stock(self):
        """Add an in-stock Cultivar to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = True
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['In Stock']).value == 'True'
        cv2 = Cultivar(name='Soulmate')
        cv2.common_name = CommonName(name='Butterfly Weed')
        cv2.common_name.index = Index(name='Perennial')
        cv2.in_stock = False
        cvws.add_one(cv2)
        assert cvws.cell(3, cvws.cols['In Stock']).value == 'False'

    def test_add_one_active(self):
        """Add an active Cultivar to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.active = True
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['Active']).value == 'True'
        cv2 = Cultivar(name='Soulmate')
        cv2.common_name = CommonName(name='Butterfly Weed')
        cv2.common_name.index = Index(name='Perennial')
        cv2.active = False
        cvws.add_one(cv2)
        assert cvws.cell(3, cvws.cols['Active']).value == 'False'

    def test_add_one_not_cultivar(self):
        """Raise a TypeError given non-Cultivar data."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        with pytest.raises(TypeError):
            cvws.add_one(42)
        with pytest.raises(TypeError):
            cvws.add_one(Section(name='Spurious'))


class TestPacketsWorksheet:
    """Test methods of the PacketsWorksheet container class."""
    @mock.patch('app.seeds.excel.PacketsWorksheet._setup')
    def test_setup_new(self, m_s):
        """Run _setup with titles on setup of new worksheet."""
        wb = Workbook()
        ws = wb.active
        pws = PacketsWorksheet(ws)
        pws.setup()
        titles = ('Cultivar (JSON)', 'SKU', 'Price', 'Quantity', 'Units')
        m_s.assert_called_with(titles)

    @mock.patch('app.seeds.excel.PacketsWorksheet._setup')
    def test_setup_existing(self, m_s):
        """Run _setup with no args on setup of sheet with titles."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.set_column_titles(('One', 'Two', 'Three'))
        pws = PacketsWorksheet(sws._ws)
        pws.setup()
        assert m_s.call_args_list == [mock.call()]

    def test_add_one(self):
        """Add a Packet to the Packets worksheet."""
        messages = StringIO()
        wb = Workbook()
        ws = wb.active
        pws = PacketsWorksheet(ws)
        pws.setup()
        pkt = Packet(sku='8675309', price='3.50')
        pkt.quantity = Quantity(value=100, units='seeds')
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        pkt.cultivar = cv
        pws.add_one(pkt, stream=messages)
        assert pws.cell(
            2, pws.cols['Cultivar (JSON)']
        ).value == json.dumps(cv.queryable_dict)
        assert pws.cell(2, pws.cols['SKU']).value == '8675309'
        assert pws.cell(2, pws.cols['Price']).value == '3.50'
        assert pws.cell(2, pws.cols['Quantity']).value == '100'
        assert pws.cell(2, pws.cols['Units']).value == 'seeds'
        messages.seek(0)
        msgs = messages.read()
        assert ('Adding data from <Packet SKU #8675309> to row #2 of packets '
                'worksheet.') in msgs

    def test_add_one_not_packet(self):
        """Raise TypeError given non-Packet data."""
        wb = Workbook()
        ws = wb.active
        pws = PacketsWorksheet(ws)
        pws.setup()
        with pytest.raises(TypeError):
            pws.add_one(42)
        with pytest.raises(TypeError):
            pws.add_one(Cultivar(name='Foxy'))


class TestSeedsWorkbook:
    """Test methods of the SeedsWorkbook container class."""
    def test_getitem(self):
        """Forward contained workbook when accessing via []."""
        swb = SeedsWorkbook()
        swb._wb = mock.MagicMock()
        swb._wb = {'key': 'value'}
        assert swb['key'] == 'value'

    @mock.patch('app.seeds.excel.SeedsWorkbook.create_all_sheets')
    def test_remove_all_sheets(self, m):
        """Remove all worksheets from the workbook."""
        swb = SeedsWorkbook()
        swb._wb.remove_sheet(swb._wb.active)
        assert not swb._wb.worksheets
        titles = ['One', 'Two', 'Three']
        for title in titles:
            swb._wb.create_sheet(title=title)
        assert sorted(titles) == sorted(swb._wb.sheetnames)

    @mock.patch('app.seeds.excel.SeedsWorkbook.remove_all_sheets')
    @mock.patch('app.seeds.excel.IndexesWorksheet.setup')
    @mock.patch('app.seeds.excel.CommonNamesWorksheet.setup')
    @mock.patch('app.seeds.excel.BotanicalNamesWorksheet.setup')
    @mock.patch('app.seeds.excel.SectionsWorksheet.setup')
    @mock.patch('app.seeds.excel.CultivarsWorksheet.setup')
    @mock.patch('app.seeds.excel.PacketsWorksheet.setup')
    def test_create_all_sheets(self,
                               m_pkt,
                               m_cv,
                               m_sr,
                               m_bn,
                               m_cn,
                               m_idx,
                               m_remove):
        """Remove all worksheets, then create all worksheets."""
        with mock.patch('app.seeds.excel.SeedsWorkbook.create_all_sheets'):
            swb = SeedsWorkbook()
        swb.create_all_sheets()
        assert m_remove.called
        assert swb.indexes._ws is swb._wb['Indexes']
        assert m_idx.called
        assert swb.common_names._ws is swb._wb['Common Names']
        assert m_cn.called
        assert swb.botanical_names._ws is swb._wb['Botanical Names']
        assert m_bn.called
        assert swb.section._ws is swb._wb['Section']
        assert m_sr.called
        assert swb.cultivars._ws is swb._wb['Cultivars']
        assert m_cv.called
        assert swb.packets._ws is swb._wb['Packets']
        assert m_pkt.called

    @mock.patch('app.seeds.excel.IndexesWorksheet.setup')
    @mock.patch('app.seeds.excel.CommonNamesWorksheet.setup')
    @mock.patch('app.seeds.excel.BotanicalNamesWorksheet.setup')
    @mock.patch('app.seeds.excel.SectionsWorksheet.setup')
    @mock.patch('app.seeds.excel.CultivarsWorksheet.setup')
    @mock.patch('app.seeds.excel.PacketsWorksheet.setup')
    def test_load_all_sheets_from_workbook(self,
                                           m_pkt,
                                           m_cv,
                                           m_sr,
                                           m_bn,
                                           m_cn,
                                           m_idx):
        """Load all sheets from self._wb into appropriate attributes."""
        swb = SeedsWorkbook()
        swb.load_all_sheets_from_workbook()
        assert swb.indexes._ws is swb._wb['Indexes']
        assert m_idx.called
        assert swb.common_names._ws is swb._wb['Common Names']
        assert m_cn.called
        assert swb.botanical_names._ws is swb._wb['Botanical Names']
        assert m_bn.called
        assert swb.section._ws is swb._wb['Section']
        assert m_sr.called
        assert swb.cultivars._ws is swb._wb['Cultivars']
        assert m_cv.called
        assert swb.packets._ws is swb._wb['Packets']
        assert m_pkt.called

    @mock.patch('app.seeds.excel.SeedsWorksheet.add')
    @mock.patch('app.seeds.excel.Index.query')
    @mock.patch('app.seeds.excel.CommonName.query')
    @mock.patch('app.seeds.excel.BotanicalName.query')
    @mock.patch('app.seeds.excel.Section.query')
    @mock.patch('app.seeds.excel.Cultivar.query')
    @mock.patch('app.seeds.excel.Packet.query')
    def test_add_all_data_to_sheets(self,
                                    m_pkt,
                                    m_cv,
                                    m_sr,
                                    m_bn,
                                    m_cn,
                                    m_idx,
                                    m_a):
        """Call <sheet>.save_to_db(<obj>.query.all()) for each worksheet."""
        messages = StringIO()
        swb = SeedsWorkbook()
        swb.add_all_data_to_sheets(stream=messages)
        m_a.assert_any_call(m_pkt.all(), stream=messages)
        m_a.assert_any_call(m_cv.all(), stream=messages)
        m_a.assert_any_call(m_sr.all(), stream=messages)
        m_a.assert_any_call(m_bn.all(), stream=messages)
        m_a.assert_any_call(m_cn.all(), stream=messages)
        m_a.assert_any_call(m_idx.all(), stream=messages)

    @mock.patch('app.seeds.excel.IndexesWorksheet.save_to_db')
    @mock.patch('app.seeds.excel.CommonNamesWorksheet.save_to_db')
    @mock.patch('app.seeds.excel.BotanicalNamesWorksheet.save_to_db')
    @mock.patch('app.seeds.excel.SectionsWorksheet.save_to_db')
    @mock.patch('app.seeds.excel.CultivarsWorksheet.save_to_db')
    @mock.patch('app.seeds.excel.PacketsWorksheet.save_to_db')
    def test_save_all_sheets_to_db(self,
                                   m_pkt,
                                   m_cv,
                                   m_sr,
                                   m_bn,
                                   m_cn,
                                   m_idx):
        """Call save_to_db for each worksheet."""
        messages = StringIO()
        swb = SeedsWorkbook()
        swb.save_all_sheets_to_db(stream=messages)
        messages.seek(0)
        msgs = messages.read()
        m_idx.assert_called_with(stream=messages)
        m_cn.assert_called_with(stream=messages)
        m_bn.assert_called_with(stream=messages)
        m_cv.assert_called_with(stream=messages)
        m_sr.assert_called_with(stream=messages)
        m_pkt.assert_called_with(stream=messages)
        assert '-- BEGIN saving all worksheets to database. --' in msgs
        assert '-- END saving all worksheets to database. --' in msgs

    @mock.patch('app.seeds.excel.SeedsWorksheet.beautify')
    def test_beautify_all_sheets(self, m_b):
        """Call beautify on all sheets in workbook."""
        swb = SeedsWorkbook()
        swb.beautify_all_sheets(width=42, height=12)
        assert m_b.call_count == 6
        m_b.assert_any_call(width=42, height=12)

    @mock.patch('app.seeds.excel.openpyxl.load_workbook')
    @mock.patch('app.seeds.excel.SeedsWorkbook.load_all_sheets_from_workbook')
    def test_load(self, m_lasfw, m_lw):
        """Load a workbook into _wb, and load all sheets from it."""
        swb = SeedsWorkbook()
        wb = Workbook()
        m_lw.return_value = wb
        swb.load('file.xlsx')
        m_lw.assert_called_with('file.xlsx')
        assert swb._wb is wb
        assert m_lasfw.called

    @mock.patch('app.seeds.excel.openpyxl.Workbook.save')
    def test_save(self, m_s):
        """Beautify a SeedsWorkbook and to a file."""
        swb = SeedsWorkbook()
        swb.save('file.xlsx')
        m_s.assert_called_with('file.xlsx')
