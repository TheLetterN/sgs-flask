import json
import pytest
from io import StringIO
from unittest import mock
from openpyxl import Workbook
from app.seeds.excel2 import (
    BotanicalNamesWorksheet,
    CommonNamesWorksheet,
    CultivarsWorksheet,
    IndexesWorksheet,
    lookup_dicts_to_json,
    PacketsWorksheet,
    SeedsWorkbook,
    SeedsWorksheet,
    SeriesWorksheet
)
from app.seeds.models import (
    BotanicalName,
    CommonName,
    Cultivar,
    Image,
    Index,
    Packet,
    Quantity,
    Series
)


class TestExcel2Functions:
    """Test module level functions."""
    def test_lookup_dicts_to_json(self):
        """Generate a JSON string for looking up Grows With cns/cvs.

        It can take either, as both have the lookup_dict() method.
        """
        gwcn1 = CommonName(name='Foxglove')
        gwcn1.index = Index(name='Perennial')
        assert lookup_dicts_to_json([gwcn1]) == \
            json.dumps((gwcn1.lookup_dict(),))
        gwcn2 = CommonName(name='Butterfly Weed')
        gwcn2.index = Index(name='Perennial')
        assert lookup_dicts_to_json([gwcn1, gwcn2]) == \
            json.dumps((gwcn1.lookup_dict(), gwcn2.lookup_dict()))
        gwcv1 = Cultivar(name='Soulmate')
        gwcv1.common_name = CommonName(name='Butterfly Weed')
        gwcv1.common_name.index = Index(name='Perennial')
        assert lookup_dicts_to_json([gwcv1]) == \
            json.dumps((gwcv1.lookup_dict(),))
        gwcv2 = Cultivar(name='Petra')
        gwcv2.common_name = CommonName(name='Foxglove')
        gwcv2.common_name.index = Index(name='Perennial')
        gwcv2.series = Series(name='Polkadot')
        assert lookup_dicts_to_json([gwcv1, gwcv2]) == \
            json.dumps((gwcv1.lookup_dict(), gwcv2.lookup_dict()))


class TestSeedsWorksheet:
    """Test methods of the SeedsWorksheet container class.

    We use Workbook to create sheets instead of Worksheet because it ensures
    the worksheet is created as it would be when used in a workbook.
    """
    def test_rows_property(self):
        """Return a tuple of tuples containing cells.

        It should return a tuple containing an empty tuple if no cells have
        data.
        """
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        assert sws.rows == ((),)
        a1 = sws._ws['A1']
        a1.value = 'One'
        assert sws.rows == ((a1,),)
        b1 = sws._ws['B1']
        b1.value = 'Two'
        assert sws.rows == ((a1, b1),)
        a2 = sws._ws['A2']
        a2.value = 'Three'
        # B2 is created with a null value to fill the grid.
        b2 = sws._ws['B2']
        assert sws.rows == ((a1, b1), (a2, b2))

    def test_active_row_property(self):
        """Return the row number of the first empty row."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        assert sws.active_row == 1
        sws._ws['A1'].value = 'One'
        assert sws.active_row == 2
        sws._ws['A2'].value = 'Two'
        assert sws.active_row == 3
        sws._ws['A3'].value = 'Three'
        assert sws.active_row == 4

    def test_data_rows_property(self):
        """Return all rows except the first row, which contains titles."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        assert sws.data_rows == tuple()
        ws['A1'] = 'One'
        assert sws.data_rows == tuple()
        ws['A2'] = 'Two'
        assert sws.data_rows == ((ws['A2'],),)
        ws['A3'] = 'Three'
        assert sws.data_rows == ((ws['A2'],), (ws['A3'],))

    def test_has_data(self):
        """Return True if there is data in cell A1, else False."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        assert not sws.has_data()
        sws._ws['A2'].value = 'Stuff'
        assert not sws.has_data()  # Okay; should never happen in usage.
        sws._ws['A1'].value = 'More stuff'
        assert sws.has_data()

    def test_cell(self):
        """Return a cell given integer coordinates."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        ws['A1'].value = 'A1'
        ws['B1'].value = 'B1'
        ws['A2'].value = 'A2'
        assert sws.cell(1, 1) is ws['A1']
        assert sws.cell(2, 1) is ws['A2']
        assert sws.cell(1, 2) is ws['B1']

    def test_set_column_titles(self):
        """Set the top row of the worksheet to a list of titles."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        titles = ['One', 'Two', 'Three', 'Four']
        sws.set_column_titles(titles)
        assert sws._ws['A1'].value == 'One'
        assert sws._ws['B1'].value == 'Two'
        assert sws._ws['C1'].value == 'Three'
        assert sws._ws['D1'].value == 'Four'

    def test_set_column_titles_existing_sheet(self):
        """Raise a ValueError if there is already data in the top row."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        titles = ['One', 'Two', 'Three', 'Four']
        sws._ws.append(titles)
        with pytest.raises(ValueError):
            sws.set_column_titles(titles)

    def test_populate_cols_dict(self):
        """Set cols keys to values in first row, and vals to col numbers."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        titles = ['One', 'Two', 'Three', 'Four']
        sws.set_column_titles(titles)
        sws.populate_cols_dict()
        assert sws.cols['One'] == 1
        assert sws.cols['Two'] == 2
        assert sws.cols['Three'] == 3
        assert sws.cols['Four'] == 4

    def test_populate_cols_dict_no_titles(self):
        """Raise a ValueError if there is no data in the first row."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        with pytest.raises(ValueError):
            sws.populate_cols_dict()

    @mock.patch('app.seeds.excel2.SeedsWorksheet.set_column_titles')
    @mock.patch('app.seeds.excel2.SeedsWorksheet.populate_cols_dict')
    def test_setup_new(self, m_pcd, m_sct):
        """Call set_column_titles and populate_cols_dict when sheet blank."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        titles = ('One', 'Two', 'Three')
        sws._setup(titles)
        assert m_pcd.called
        assert m_sct.called_with(titles)

    def test_setup_new_no_titles(self):
        """Raise a value error if _setup is run without titles on new sheet."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        with pytest.raises(ValueError):
            sws._setup()

    @mock.patch('app.seeds.excel2.SeedsWorksheet.set_column_titles')
    @mock.patch('app.seeds.excel2.SeedsWorksheet.populate_cols_dict')
    def test_setup_existing(self, m_pcd, m_sct):
        """Do not call set_column_titles if data in first row already."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws._ws['A1'].value = 'One'
        sws._setup()
        assert m_pcd.called
        assert not m_sct.called

    @mock.patch('app.seeds.excel2.SeedsWorksheet.set_column_titles')
    @mock.patch('app.seeds.excel2.SeedsWorksheet.populate_cols_dict')
    def test_setup_existing_with_titles(self, m_pcd, m_sct):
        """Set up as normal but discard new titles."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws._ws['A1'].value = 'One'
        titles = ('One', 'Two', 'Three')
        with pytest.warns(UserWarning):
            sws._setup(titles)
        assert m_pcd.called
        assert not m_sct.called

    def testadd_one(self):
        """add_one should be an abstract method in this class."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        with pytest.raises(NotImplementedError):
            sws.add_one(None)

    @mock.patch('app.seeds.excel2.SeedsWorksheet.add_one')
    def test_add(self, m_ao):
        """add should call add_one for each item in iterable."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.add(('Test',))
        assert m_ao.called

    @mock.patch('app.seeds.excel2.SeedsWorksheet.add_one')
    def test_add_bad_data(self, m_ao):
        """Warn user when iterable contains bad types instead of halting."""
        m_ao.side_effect = TypeError('Bad data, yo!')
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        with pytest.warns(UserWarning):
            sws.add((1, 2, 3, 4))
        assert m_ao.call_count == 4

    def test_add_not_iterable(self):
        """Do not suppress TypeError if given non-iterable."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        with pytest.raises(TypeError):
            sws.add(42)

    def test_save_row_to_db(self):
        """save_row_to_db should be an abstract method in this class."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        with pytest.raises(NotImplementedError):
            sws.save_row_to_db(None)

    def test_beautify(self):
        """Configure worksheet to be more human-readable."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws._ws.append(('One', 'Two', 'Three'))
        sws._ws.append(('Four', 'Five', 'Six'))
        sws.beautify(width=42, height=21)
        assert sws._ws.freeze_panes == 'A2'
        assert sws._ws.column_dimensions['A'].width == 42
        assert sws._ws.column_dimensions['B'].width == 42
        assert sws._ws.column_dimensions['C'].width == 42
        assert sws._ws.row_dimensions[2].height == 21


class TestIndexesWorksheet:
    """Test methods of the IndexesWorksheet container class."""
    @mock.patch('app.seeds.excel2.IndexesWorksheet._setup')
    def test_setup_new(self, m_s):
        """Call _setup with titles for Indexes sheet."""
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        m_s.assert_called_with(('Index', 'Description'))

    @mock.patch('app.seeds.excel2.IndexesWorksheet._setup')
    def test_setup_existing(self, m_s):
        """Call _setup with no data if sheet titles already populated."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.set_column_titles(('Index', 'Description'))
        iws = IndexesWorksheet(sws._ws)
        iws.setup()
        assert m_s.call_args_list == [mock.call()]

    def testadd_one(self):
        """Add a single Index to worksheet."""
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        idx = Index(name='Perennial', description='Built to last.')
        iws.add_one(idx)
        assert iws.cell(2, iws.cols['Index']).value == 'Perennial'
        assert iws.cell(2, iws.cols['Description']).value == 'Built to last.'

    def testadd_one_bad_type(self):
        """Raise a TypeError given non-Index data."""
        wb = Workbook()
        ws = wb.active
        iws = IndexesWorksheet(ws)
        iws.setup()
        with pytest.raises(TypeError):
            iws.add_one('Frogs!')


class TestCommonNamesWorksheet:
    """Test methods of the CommonNamesWorksheet container class."""
    @mock.patch('app.seeds.excel2.CommonNamesWorksheet._setup')
    def test_setup_new(self, m_s):
        """Call _setup with titles for Common Names worksheet."""
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        titles = ('Index',
                  'Common Name',
                  'Subcategory of',
                  'Description',
                  'Planting Instructions',
                  'Synonyms',
                  'Invisible',
                  'Grows With Common Names (JSON)',
                  'Grows With Cultivars (JSON)')
        m_s.assert_called_with(titles)

    @mock.patch('app.seeds.excel2.CommonNamesWorksheet._setup')
    def test_setup_existing(self, m_s):
        """Call _setup with no data if titles row already populated."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.set_column_titles(('One', 'Two', 'Three'))
        cnws = CommonNamesWorksheet(sws._ws)
        cnws.setup()
        assert m_s.call_args_list == [mock.call()]

    def testadd_one_no_optionals(self):
        """Add a common name (with no optional data) to Common Names sheet."""
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        cnws.add_one(cn)
        assert cnws.cell(2, cnws.cols['Index']).value == 'Perennial'
        assert cnws.cell(2, cnws.cols['Common Name']).value == 'Foxglove'
        assert cnws.cell(2, cnws.cols['Subcategory of']).value is None
        assert cnws.cell(2, cnws.cols['Description']).value is None
        assert cnws.cell(2, cnws.cols['Planting Instructions']).value is None
        assert cnws.cell(2, cnws.cols['Synonyms']).value is None
        assert cnws.cell(2, cnws.cols['Invisible']).value == 'False'
        assert cnws.cell(
            2, cnws.cols['Grows With Common Names (JSON)']
        ).value is None
        assert cnws.cell(
            2, cnws.cols['Grows With Cultivars (JSON)']
        ).value is None

    def testadd_one_no_gw(self):
        """Add a common name (with no Grows With) to Common Names sheet."""
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove',
                        description='Spotty.',
                        instructions='Just add water!')
        cn.index = Index(name='Perennial')
        cn.parent = CommonName(name='Fauxglove')
        cn.invisible = True
        cn.set_synonyms_string('Digitalis')
        cnws.add_one(cn)
        assert cnws.cell(2, cnws.cols['Subcategory of']).value == 'Fauxglove'
        assert cnws.cell(2, cnws.cols['Description']).value == 'Spotty.'
        assert cnws.cell(
            2, cnws.cols['Planting Instructions']
        ).value == 'Just add water!'
        assert cnws.cell(2, cnws.cols['Synonyms']).value == 'Digitalis'
        assert cnws.cell(2, cnws.cols['Invisible']).value == 'True'

    def testadd_one_with_gw_cn(self):
        """Add a common name with some Grows With Common Names."""
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        gwcn1 = CommonName(name='Tomato')
        gwcn1.index = Index(name='Vegetable')
        gwcn2 = CommonName(name='Basil')
        gwcn2.index = Index(name='Herb')
        cn.gw_common_names = [gwcn1, gwcn2]
        cnws.add_one(cn)
        assert cnws.cell(
            2, cnws.cols['Grows With Common Names (JSON)']
        ).value == lookup_dicts_to_json([gwcn1, gwcn2])

    def testadd_one_with_gw_cv(self):
        """Add a common name with some Grows With Cultivars."""
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        cn = CommonName(name='Foxglove')
        cn.index = Index(name='Perennial')
        gwcv1 = Cultivar(name='Soulmate')
        gwcv1.common_name = CommonName(name='Butterfly Weed')
        gwcv1.common_name.index = Index(name='Perennial')
        gwcv2 = Cultivar(name='Petra')
        gwcv2.common_name = CommonName(name='Foxglove')
        gwcv2.common_name.index = Index(name='Perennial')
        gwcv2.series = Series(name='Polkadot')
        cn.gw_cultivars = [gwcv1, gwcv2]
        cnws.add_one(cn)
        assert cnws.cell(
            2, cnws.cols['Grows With Cultivars (JSON)']
        ).value == lookup_dicts_to_json([gwcv1, gwcv2])

    def testadd_one_not_common_name(self):
        """Raise a TypeError given data that isn't a CommonName."""
        wb = Workbook()
        ws = wb.active
        cnws = CommonNamesWorksheet(ws)
        cnws.setup()
        with pytest.raises(TypeError):
            cnws.add_one(42)
        with pytest.raises(TypeError):
            cnws.add_one(Index(name='Perennial'))


class TestBotanicalNamesWorksheet:
    """Test methods of the BotanicalNamesWorksheet container class."""
    @mock.patch('app.seeds.excel2.BotanicalNamesWorksheet._setup')
    def test_setup_new(self, m_s):
        """Call _setup with titles for Botanical Names worksheet."""
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        titles = ('Common Names (JSON)',
                  'Botanical Name',
                  'Synonyms')
        m_s.assert_called_with(titles)

    @mock.patch('app.seeds.excel2.BotanicalNamesWorksheet._setup')
    def test_setup_existing(self, m_s):
        """Call _setup with no data if titles already present."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.set_column_titles(('One', 'Two', 'Three'))
        bnws = BotanicalNamesWorksheet(sws._ws)
        bnws.setup()
        assert m_s.call_args_list == [mock.call()]

    def testadd_one_no_optionals(self):
        """Add a BotanicalName object to Botanical Names sheet."""
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Innagada davida')
        cn = CommonName(name='Rock')
        cn.index = Index(name='Music')
        bn.common_names = [cn]
        bnws.add_one(bn)
        assert bnws.cell(
            2, bnws.cols['Common Names (JSON)']
        ).value == lookup_dicts_to_json([cn])
        assert bnws.cell(
            2, bnws.cols['Botanical Name']
        ).value == 'Innagada davida'
        assert bnws.cell(2, bnws.cols['Synonyms']).value is None

    def testadd_one_with_synonyms(self):
        """Add a BotanicalName with synonyms to Botanical Names sheet."""
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        bn = BotanicalName(name='Innagada davida')
        cn = CommonName(name='Rock')
        cn.index = Index(name='Music')
        bn.common_names = [cn]
        bn.set_synonyms_string('Iron butterfly')
        bnws.add_one(bn)
        assert bnws.cell(2, bnws.cols['Synonyms']).value == 'Iron butterfly'

    def testadd_one_not_botanical_name(self):
        """Raise a TypeError given non-BotanicalName data."""
        wb = Workbook()
        ws = wb.active
        bnws = BotanicalNamesWorksheet(ws)
        bnws.setup()
        with pytest.raises(TypeError):
            bnws.add_one(42)
        with pytest.raises(TypeError):
            bnws.add_one(CommonName(name='Spurious'))


class TestSeriesWorksheet:
    """Test methods of the SeriesWorksheet container class."""
    @mock.patch('app.seeds.excel2.SeriesWorksheet._setup')
    def test_setup_new(self, m_s):
        """Run _setup with the titles for the Series worksheet."""
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        titles = ('Common Name (JSON)', 'Series', 'Position', 'Description')
        m_s.assert_called_with(titles)

    @mock.patch('app.seeds.excel2.SeriesWorksheet._setup')
    def test_setup_existing(self, m_s):
        """Run _setup with no arguments if titles apready present."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.set_column_titles(('One', 'Two', 'Three'))
        srws = SeriesWorksheet(sws._ws)
        srws.setup()
        assert m_s.call_args_list == [mock.call()]

    def testadd_one_no_optionals(self):
        """Add a Series object to the Series worksheet."""
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        srws.add_one(sr)
        assert srws.cell(
            2, srws.cols['Common Name (JSON)']
        ).value == json.dumps(sr.common_name.lookup_dict())
        assert srws.cell(2, srws.cols['Series']).value == 'Polkadot'
        assert srws.cell(2, srws.cols['Position']).value == 'before cultivar'
        assert srws.cell(2, srws.cols['Description']).value is None

    def testadd_one_with_position(self):
        """Set Position column's cell with relevant position description."""
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        sr.position = Series.BEFORE_CULTIVAR
        srws.add_one(sr)
        assert srws.cell(2, srws.cols['Position']).value == 'before cultivar'
        sr2 = Series(name='Queen')
        sr2.common_name = CommonName(name='Cleome')
        sr2.common_name.index = Index(name='Annual')
        sr2.position = Series.AFTER_CULTIVAR
        srws.add_one(sr2)
        assert srws.cell(3, srws.cols['Position']).value == 'after cultivar'

    def testadd_one_with_description(self):
        """Set Description column's cell with Series desc."""
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        sr = Series(name='Polkadot', description='A bit spotty.')
        sr.common_name = CommonName(name='Foxglove')
        sr.common_name.index = Index(name='Perennial')
        srws.add_one(sr)
        assert srws.cell(2, srws.cols['Description']).value == 'A bit spotty.'

    def testadd_one_not_series(self):
        """Raise a TypeError if passed argument is not a Series object."""
        wb = Workbook()
        ws = wb.active
        srws = SeriesWorksheet(ws)
        srws.setup()
        with pytest.raises(TypeError):
            srws.add_one(42)
        with pytest.raises(TypeError):
            srws.add_one(Index(name='Perennial'))


class TestCultivarsWorksheet:
    """Test methods of the CultivarsWorksheet container class."""
    @mock.patch('app.seeds.excel2.CultivarsWorksheet._setup')
    def test_setup_new(self, m_s):
        """Run _setup with the titles for the Cultivars worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        titles = ('Index',
                  'Common Name',
                  'Cultivar Name',
                  'Series',
                  'Botanical Name',
                  'Thumbnail Filename',
                  'Description',
                  'Synonyms',
                  'New For',
                  'In Stock',
                  'Active',
                  'Invisible',
                  'Grows With Common Names (JSON)',
                  'Grows With Cultivars (JSON)')
        m_s.assert_called_with(titles)

    @mock.patch('app.seeds.excel2.CultivarsWorksheet._setup')
    def test_setup_existing(self, m_s):
        """Run _setup with no arguments if titles already exist."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.set_column_titles(('One', 'Two', 'Three'))
        cvws = CultivarsWorksheet(sws._ws)
        cvws.setup()
        assert m_s.call_args_list == [mock.call()]

    def testadd_one_no_optionals(self):
        """Add a Cultivar to the Cultivars worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['Index']).value == 'Perennial'
        assert cvws.cell(2, cvws.cols['Common Name']).value == 'Foxglove'
        assert cvws.cell(2, cvws.cols['Cultivar Name']).value == 'Foxy'
        assert cvws.cell(2, cvws.cols['Series']).value is None
        assert cvws.cell(2, cvws.cols['Botanical Name']).value is None
        assert cvws.cell(2, cvws.cols['Thumbnail Filename']).value is None
        assert cvws.cell(2, cvws.cols['Description']).value is None
        assert cvws.cell(2, cvws.cols['Synonyms']).value is None
        assert cvws.cell(2, cvws.cols['New For']).value is None
        assert cvws.cell(2, cvws.cols['In Stock']).value == 'False'
        assert cvws.cell(2, cvws.cols['Active']).value == 'False'
        assert cvws.cell(2, cvws.cols['Invisible']).value == 'False'
        assert cvws.cell(
            2, cvws.cols['Grows With Common Names (JSON)']
        ).value is None
        assert cvws.cell(
            2, cvws.cols['Grows With Cultivars (JSON)']
        ).value is None

    def testadd_one_with_series(self):
        """Add a Cultivar with Series to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Petra')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.series = Series(name='Polkadot')
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['Series']).value == 'Polkadot'

    def testadd_one_with_botanical_name(self):
        """Add a Cultivar with a Botanical Name to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.botanical_name = BotanicalName(name='Digitalis purpurea')
        cvws.add_one(cv)
        assert cvws.cell(
            2, cvws.cols['Botanical Name']
        ).value == 'Digitalis purpurea'

    def testadd_one_with_thumbnail_filename(self):
        """Add a Cultivar with a Thumbnail Filename to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.thumbnail = Image(filename='foo.jpg')
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['Thumbnail Filename']).value == 'foo.jpg'

    def testadd_one_with_description(self):
        """Add a Cultivar with a description to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.description = 'Like a lady!'
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['Description']).value == 'Like a lady!'

    def testadd_one_with_synonyms(self):
        """Add a Cultivar with synonyms to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.set_synonyms_string('Vulpine')
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['Synonyms']).value == 'Vulpine'

    def testadd_one_with_new_for(self):
        """Add a Cultivar with New For to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.new_for = 1984
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['New For']).value == 1984

    def testadd_one_in_stock(self):
        """Add an in-stock Cultivar to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.in_stock = True
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['In Stock']).value == 'True'
        cv2 = Cultivar(name='Soulmate')
        cv2.common_name = CommonName(name='Butterfly Weed')
        cv2.common_name.index = Index(name='Perennial')
        cv2.in_stock = False
        cvws.add_one(cv2)
        assert cvws.cell(3, cvws.cols['In Stock']).value == 'False'

    def testadd_one_active(self):
        """Add an active Cultivar to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.active = True
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['Active']).value == 'True'
        cv2 = Cultivar(name='Soulmate')
        cv2.common_name = CommonName(name='Butterfly Weed')
        cv2.common_name.index = Index(name='Perennial')
        cv2.active = False
        cvws.add_one(cv2)
        assert cvws.cell(3, cvws.cols['Active']).value == 'False'

    def testadd_one_invisible(self):
        """Add an invisible Cultivar to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        cv.invisible = True
        cvws.add_one(cv)
        assert cvws.cell(2, cvws.cols['Invisible']).value == 'True'
        cv2 = Cultivar(name='Soulmate')
        cv2.common_name = CommonName(name='Butterfly Weed')
        cv2.common_name.index = Index(name='Perennial')
        cv2.invisible = False
        cvws.add_one(cv2)
        assert cvws.cell(3, cvws.cols['Invisible']).value == 'False'

    def testadd_one_with_gw_common_names(self):
        """Add a Cultivar with Grows With Common Names to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        gwcn = CommonName(name='Butterfly Weed')
        gwcn.index = Index(name='Perennial')
        cv.gw_common_names = [gwcn]
        cvws.add_one(cv)
        assert cvws.cell(
            2, cvws.cols['Grows With Common Names (JSON)']
        ).value == lookup_dicts_to_json((gwcn,))

    def testadd_one_with_gw_cultivars(self):
        """Add a Cultivar with Grows With Cultivars to worksheet."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        gwcv = Cultivar(name='Soulmate')
        gwcv.common_name = CommonName(name='Butterfly Weed')
        gwcv.common_name.index = Index(name='Perennial')
        cv.gw_cultivars = [gwcv]
        cvws.add_one(cv)
        assert cvws.cell(
            2, cvws.cols['Grows With Cultivars (JSON)']
        ).value == lookup_dicts_to_json((gwcv,))

    def testadd_one_not_cultivar(self):
        """Raise a TypeError given non-Cultivar data."""
        wb = Workbook()
        ws = wb.active
        cvws = CultivarsWorksheet(ws)
        cvws.setup()
        with pytest.raises(TypeError):
            cvws.add_one(42)
        with pytest.raises(TypeError):
            cvws.add_one(Series(name='Spurious'))


class TestPacketsWorksheet:
    """Test methods of the PacketsWorksheet container class."""
    @mock.patch('app.seeds.excel2.PacketsWorksheet._setup')
    def test_setup_new(self, m_s):
        """Run _setup with titles on setup of new worksheet."""
        wb = Workbook()
        ws = wb.active
        pws = PacketsWorksheet(ws)
        pws.setup()
        titles = ('Cultivar (JSON)', 'SKU', 'Price', 'Quantity', 'Units')
        m_s.assert_called_with(titles)

    @mock.patch('app.seeds.excel2.PacketsWorksheet._setup')
    def test_setup_existing(self, m_s):
        """Run _setup with no args on setup of sheet with titles."""
        wb = Workbook()
        ws = wb.active
        sws = SeedsWorksheet(ws)
        sws.set_column_titles(('One', 'Two', 'Three'))
        pws = PacketsWorksheet(sws._ws)
        pws.setup()
        assert m_s.call_args_list == [mock.call()]

    def test_add_one(self):
        """Add a Packet to the Packets worksheet."""
        wb = Workbook()
        ws = wb.active
        pws = PacketsWorksheet(ws)
        pws.setup()
        pkt = Packet(sku='8675309', price='3.50')
        pkt.quantity = Quantity(value=100, units='seeds')
        cv = Cultivar(name='Foxy')
        cv.common_name = CommonName(name='Foxglove')
        cv.common_name.index = Index(name='Perennial')
        pkt.cultivar = cv
        pws.add_one(pkt)
        assert pws.cell(
            2, pws.cols['Cultivar (JSON)']
        ).value == json.dumps(cv.lookup_dict())
        assert pws.cell(2, pws.cols['SKU']).value == '8675309'
        assert pws.cell(2, pws.cols['Price']).value == '3.50'
        assert pws.cell(2, pws.cols['Quantity']).value == '100'
        assert pws.cell(2, pws.cols['Units']).value == 'seeds'

    def testadd_one_not_packet(self):
        """Raise TypeError given non-Packet data."""
        wb = Workbook()
        ws = wb.active
        pws = PacketsWorksheet(ws)
        pws.setup()
        with pytest.raises(TypeError):
            pws.add_one(42)
        with pytest.raises(TypeError):
            pws.add_one(Cultivar(name='Foxy'))


class TestSeedsWorkbook:
    """Test methods of the SeedsWorkbook container class."""
    @mock.patch('app.seeds.excel2.SeedsWorkbook.create_all_sheets')
    def test_remove_all_sheets(self, m):
        """Remove all worksheets from the workbook."""
        swb = SeedsWorkbook()
        swb._wb.remove_sheet(swb._wb.active)
        assert not swb._wb.worksheets
        titles = ['One', 'Two', 'Three']
        for title in titles:
            swb._wb.create_sheet(title=title)
        assert sorted(titles) == sorted(swb._wb.sheetnames)

    @mock.patch('app.seeds.excel2.SeedsWorkbook.remove_all_sheets')
    @mock.patch('app.seeds.excel2.IndexesWorksheet.setup')
    @mock.patch('app.seeds.excel2.CommonNamesWorksheet.setup')
    @mock.patch('app.seeds.excel2.BotanicalNamesWorksheet.setup')
    @mock.patch('app.seeds.excel2.SeriesWorksheet.setup')
    @mock.patch('app.seeds.excel2.CultivarsWorksheet.setup')
    @mock.patch('app.seeds.excel2.PacketsWorksheet.setup')
    def test_create_all_sheets(self,
                               m_pkt,
                               m_cv,
                               m_sr,
                               m_bn,
                               m_cn,
                               m_idx,
                               m_remove):
        """Remove all worksheets, then create all worksheets."""
        with mock.patch('app.seeds.excel2.SeedsWorkbook.create_all_sheets'):
            swb = SeedsWorkbook()
        swb.create_all_sheets()
        assert m_remove.called
        assert swb.indexes._ws is swb._wb['Indexes']
        assert m_idx.called
        assert swb.common_names._ws is swb._wb['Common Names']
        assert m_cn.called
        assert swb.botanical_names._ws is swb._wb['Botanical Names']
        assert m_bn.called
        assert swb.series._ws is swb._wb['Series']
        assert m_sr.called
        assert swb.cultivars._ws is swb._wb['Cultivars']
        assert m_cv.called
        assert swb.packets._ws is swb._wb['Packets']
        assert m_pkt.called

    @mock.patch('app.seeds.excel2.IndexesWorksheet.setup')
    @mock.patch('app.seeds.excel2.CommonNamesWorksheet.setup')
    @mock.patch('app.seeds.excel2.BotanicalNamesWorksheet.setup')
    @mock.patch('app.seeds.excel2.SeriesWorksheet.setup')
    @mock.patch('app.seeds.excel2.CultivarsWorksheet.setup')
    @mock.patch('app.seeds.excel2.PacketsWorksheet.setup')
    def test_load_all_sheets_from_workbook(self,
                                           m_pkt,
                                           m_cv,
                                           m_sr,
                                           m_bn,
                                           m_cn,
                                           m_idx):
        """Load all sheets from self._wb into appropriate attributes."""
        swb = SeedsWorkbook()
        swb.load_all_sheets_from_workbook()
        assert swb.indexes._ws is swb._wb['Indexes']
        assert m_idx.called
        assert swb.common_names._ws is swb._wb['Common Names']
        assert m_cn.called
        assert swb.botanical_names._ws is swb._wb['Botanical Names']
        assert m_bn.called
        assert swb.series._ws is swb._wb['Series']
        assert m_sr.called
        assert swb.cultivars._ws is swb._wb['Cultivars']
        assert m_cv.called
        assert swb.packets._ws is swb._wb['Packets']
        assert m_pkt.called

    @mock.patch('app.seeds.excel2.SeedsWorksheet.add')
    @mock.patch('app.seeds.excel2.Index.query')
    @mock.patch('app.seeds.excel2.CommonName.query')
    @mock.patch('app.seeds.excel2.BotanicalName.query')
    @mock.patch('app.seeds.excel2.Series.query')
    @mock.patch('app.seeds.excel2.Cultivar.query')
    @mock.patch('app.seeds.excel2.Packet.query')
    def test_add_all_data_to_sheets(self,
                                    m_pkt,
                                    m_cv,
                                    m_sr,
                                    m_bn,
                                    m_cn,
                                    m_idx,
                                    m_a):
        """Call <sheet>.save_to_db(<obj>.query.all()) for each worksheet."""
        swb = SeedsWorkbook()
        swb.add_all_data_to_sheets()
        m_a.assert_any_call(m_pkt.all())
        m_a.assert_any_call(m_cv.all())
        m_a.assert_any_call(m_sr.all())
        m_a.assert_any_call(m_bn.all())
        m_a.assert_any_call(m_cn.all())
        m_a.assert_any_call(m_idx.all())

    @mock.patch('app.seeds.excel2.IndexesWorksheet.save_to_db')
    @mock.patch('app.seeds.excel2.CommonNamesWorksheet.save_to_db')
    @mock.patch('app.seeds.excel2.BotanicalNamesWorksheet.save_to_db')
    @mock.patch('app.seeds.excel2.SeriesWorksheet.save_to_db')
    @mock.patch('app.seeds.excel2.CultivarsWorksheet.save_to_db')
    @mock.patch('app.seeds.excel2.PacketsWorksheet.save_to_db')
    def test_save_all_sheets_to_db(self,
                                   m_pkt,
                                   m_cv,
                                   m_sr,
                                   m_bn,
                                   m_cn,
                                   m_idx):
        """Call save_to_db for each worksheet."""
        messages = StringIO()
        swb = SeedsWorkbook()
        swb.save_all_sheets_to_db(file=messages)
        messages.seek(0)
        msgs = messages.read()
        m_idx.assert_called_with(file=messages)
        m_cn.assert_called_with(file=messages)
        m_bn.assert_called_with(file=messages)
        m_cv.assert_called_with(file=messages)
        m_sr.assert_called_with(file=messages)
        m_pkt.assert_called_with(file=messages)
        assert '-- BEGIN saving all worksheets to database. --' in msgs
        assert '-- END saving all worksheets to database. --' in msgs

    @mock.patch('app.seeds.excel2.SeedsWorksheet.beautify')
    def test_beautify_all_sheets(self, m_b):
        """Call beautify on all sheets in workbook."""
        swb = SeedsWorkbook()
        swb.beautify_all_sheets(width=42, height=12)
        assert m_b.call_count == 6
        m_b.assert_any_call(width=42, height=12)

    @mock.patch('app.seeds.excel2.openpyxl.load_workbook')
    @mock.patch('app.seeds.excel2.SeedsWorkbook.load_all_sheets_from_workbook')
    def test_load(self, m_lasfw, m_lw):
        """Load a workbook into _wb, and load all sheets from it."""
        swb = SeedsWorkbook()
        wb = Workbook()
        m_lw.return_value = wb
        swb.load('file.xlsx')
        m_lw.assert_called_with('file.xlsx')
        assert swb._wb is wb
        assert m_lasfw.called

    @mock.patch('app.seeds.excel2.openpyxl.Workbook.save')
    def test_save(self, m_s):
        """Beautify a SeedsWorkbook and to a file."""
        swb = SeedsWorkbook()
        swb.save('file.xlsx')
        m_s.assert_called_with('file.xlsx')
